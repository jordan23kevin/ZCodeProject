#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Y2 Bridge Server v2.6.4
=======================
Flask HTTP 桥接服务 — 连接 Y2 控制台与本地 Lovart 管线 + 文件系统

架构: HTML ←HTTP/JSON→ Flask Bridge ←subprocess→ Lovart-official pipeline
                                    ←文件IO→   INBOX / DX 目录 / Registry

变更 v2.6.4：
  - 修复 /api/open「📂 打开文件夹」对卫衣（HX 前缀）报"参数非法"：原正则写死 ^DX...，
    且目录固定 PROJECTS_DIR（D:\Semems WB）。改走 _upload_cat_guard + _dx_re(ctx["prefix"])，
    按 ?cat= 路由到对应品类 02_PROJECTS（hoodie→D:\Semems Hoodie），wb 行为不变。

变更 v2.6.3：
  - /api/batch-upload 启动前先自动清理仍在运行的旧 wb_listing.py 进程（_kill_stale_wb_listing，
    只杀命令行含 wb_listing.py 的 python 进程）：用户规则"点批量上传/强制重新上款后只执行最新任务"；
    顺带解决僵尸进程占用已上款记录/标题缓存导致删记录 Permission denied 的源头。
    清理放在 force 删记录之前，杀完等 0.5s 让文件句柄释放。

变更 v2.6.2：
  - 修复「强制重新上款点了没反应（开 Edge 但不传图）」：已上款记录/标题缓存文件被占用
    （Permission denied，如 2026-08-21 00:29 实发）时，旧逻辑只打印日志仍照常启动 wb_listing，
    记录没删成 → 脚本判"已上款"跳过 → 用户看到静默空跑。
    改为：①两个删除函数对 PermissionError 重试 3 次（间隔 0.5s，覆盖瞬时锁）；
    ②删除失败返回 None，/api/batch-upload 收到 None 直接 500 报错给前端、绝不启动上款。

变更 v2.5.1：
  - 多品类架构第 1 步（零回归重构）：BASE_DIR 与 WB_REGISTRY_FILE / CHECK_REM_SCRIPT
    不再写死字面量，改走品类注册表单一真源 D:\Semems\wb_category.py（root_for(cat)，
    缺省 wb=T恤）。常量值与改造前逐字节一致，已用生产解释器+生产 PYTHONPATH 实测回归。
    为 D:\Semems Hoodie（卫衣，HX 前缀）接入做准备；本版本不含卫衣业务逻辑。

变更 v2.5.0：
  - 新增「价格申报视角」批量处理子系统（Temu「待卖家确认」调价单列表）：
    * 新增 /order-price 页面（order_price.html）+ 导航按钮（lovart_control.html「📉 价格申报」）。
    * 后端新增 /api/order_price/{scan,auto,reject,status,enter} 端点：scan 只读预览、auto 自动接受(≥底价)、
      reject 批量拒绝(<底价)、status 后台任务进度轮询。
    * 复用 _ensure_edge_cdp 连接共用 Edge 调试端口 9222（绝不另开第二个 Edge），自动点「待卖家确认」+ 设每页 200 条。
    * 扫描预览按核价底价聚合：接受/拒绝/跳过 + 各站「核价底价/接受最低价/拒绝最高价」。
    * 自动接受：仅对「建议价≥底价」逐条点「调整」→ 弹窗「确认」；attempted 守卫防确认弹窗叠加；低于底价保持原样留人工。
    * 批量拒绝：逐个勾选低于底价订单（绝不点全选）→ 批量拒绝 → 填原因「价格过低」→ 点面板外「拒绝」→
      最终「拒绝调价」确认弹窗点「拒绝」真正提交。
    * 核价底价字典 ORDER_PRICE_FLOOR（权威来源 Temu 核价仓 PRICE_MAP）：波兰52/匈牙利56/立陶宛56/德国63/捷克65/
      斯洛伐克67/葡萄牙76/西班牙85/比利时85/法国70/丹麦84/斯洛文尼亚84/奥地利86/荷兰89/罗马尼亚100/瑞典134/
      芬兰142/意大利115。

变更 v2.4.2：
  - 新增遮罩生成子系统（胚衣制作 / 人物前景遮挡）：
    * 新增 /peiyi 页面 + 19 个 /api/peiyi/* 端点（upload/list/scores/material/versions/
      version_file/use_version/open/delete/reindex/meta/mask/correct_*/working_file/
      delete_version/import_manual）。
    * 「生成遮罩」按钮 → _peiyi_worker.py mask → peiyi_mask.generate_masks（BiRefNet + LAB 聚类
      + FASHN 语义分割，v1.5.2），输出存档到 03_MATERIAL/<分类>/_mask_versions/<stem>/vNNN/。
    * 联动 tpl_generator.generate_tpl_for_material 生成 _tpl 扭曲素材；贴图时 white_t_mockup
      自动传入 --occluder（即 *_occluder.png）盖到印花上层。
    * 手动校正（点选扩散）+ 导入手动 PS 遮罩合并（peiyi_correct.py）。
    * 评分总表 /api/peiyi/scores（低分排前标红）。

变更 v2.4.1：
  - 建议零售价填写新增「🔍 诊断结构」网页按钮：复用现有「👌 好了」信号机制，
    免手动建 go.signal 文件即可触发 --diagnose 模式。
  - _start_retail_price_script 增加 diagnose 参数，diagnose=True 时 node 命令附加 --diagnose。
  - 新增 /api/retail_price/start_diagnose 端点；建议零售价.js 诊断结果同时写入 建议零售价_diagnose.json。

变更 v2.4.0：
  - 刷新已上款改为增量游标模式（联动 check_online_listed.py v1.4.0）：
    * json 新增 ordered_list / last_oldest_dx 字段，日常刷新翻到上次边界款为止，集合相减自动移除下架款
    * 首次运行全量建库；深度清理模式全量覆盖重置边界
    * /api/upload/refresh-online-listed 支持 ?mode=incremental|deep
  - 修复「刷新已上款」前端轮询提前停止：停止条件改为检测 online_updated_at 变化，不再 9 秒假完成
  - 新增「🧹 深度清理」按钮（全量覆盖，移除所有下架款）
  - /api/upload/projects 返回 online_mode

变更 v2.3.23：
  - 同步 wb上款 v2.2.2：
    * 修复 EdgeService 窗口操作误匹配夸克/Chrome 等 Chromium 浏览器的问题。
    * `_find_edge_windows()` 增加 `msedge.exe` 进程名校验，不再按类名误操作夸克窗口。
    * `show_for_user()` / `prepare_for_interaction()` / `hide_for_automation()` / `hide_at_bottom()`
      全部按 Edge 自身进程树执行，避免把夸克透明窗口提到前台或恢复不透明导致遮挡屏幕。
  - Bridge 自身代码无改动，仅更新依赖版本与文档。

变更 v2.3.22：
  - 集成 Temu 报活动控制台 (`/activity`) 与报活动引擎 v4.1.3。
  - 新增 `/api/activity/*` 端点：启动报活动、停止、状态轮询。
  - 新增 `activity.html` 前端页面，支持启动/停止、状态徽章、实时日志、当前步骤与已完成步骤展示。
  - `lovart_control.html` 工具栏新增「报活动」按钮，可在新标签页打开 `/activity`。
  - `/api/activity/status` 按 contract 返回 `{status, log: [str], state_info}`，state.json 不存在时返回空 state_info。

变更 v2.3.21：
  - 修复 WB 上款页面缩略图黑白错位。
    * 根因：`_get_upload_thumb` / `_get_ai_thumb` 用 `re.sub(r'[^A-Za-z0-9_.-]', '_', filename)` 把
      文件名中的中文统一替换为下划线，导致 `DX_B_白T.jpg` 与 `DX_B_黑T.jpg` 生成同一个缓存文件名。
    * 解决：safe_name 只替换 Windows 文件系统非法字符（`\ / * ? : " < > |`），保留中文。
    * 清理：`D:\Semems WB\_upload_thumbs` 与 `_ai_review_thumbs` 中的错误缓存已清空，重新加载页面会自动重建正确缩略图。
  - 修复点击上款图片/回收站按钮后文件夹不自动前台弹出的问题。
    * 根因：`os.startfile` 打开已存在的资源管理器窗口时不会强制激活。
    * 解决：新增 `_open_folder_front()`，使用 `explorer.exe` 打开并在打开后通过 `win32gui` 查找窗口、
      `ShowWindow(SW_RESTORE)` + `SetForegroundWindow()` 强制置顶。

变更 v2.3.20：
  - 集成 Temu 核价控制台 (`/pricing`) 与 Hermes 核价引擎。
  - 新增 `/api/pricing/*` 端点：启动核价、停止、状态轮询、导出结果、下载 Excel、发送 "好了" 信号。
  - 新增 `pricing.html` 前端页面，支持完整自动核价 / 仅核价不提交 / 继续提交 / 重试指定页 / 导出结果。
  - 修复长页核价时滚动回顶导致无法完成的问题（联动 temu-hengjia-engine v5.2.1）。
  - 核价结果输出到 `C:/Users/Administrator/Desktop/核价档案`。
  - 核价页面 `pricing.html` 与 `/api/pricing/start` 支持指定页码范围（如 2-52）：前端新增「页码范围」输入框，后端透传 `--pages=A-B` 给 hengjia.py；`/api/pricing/retry` 同步支持 `2-52` 区间展开（联动 temu-hengjia-engine v5.2.2）。

变更 v2.3.19：
  - `upload.html`（WB 上款页面）新增「📋 复制未上款」按钮。
  - 一键复制当前未上款列表中的所有 DX 款号到剪贴板（逗号分隔）。
  - 兼容 `navigator.clipboard` 与 `document.execCommand('copy')` 兜底。

变更 v2.3.17：
  - `lovart_bridge.bat` 启动 Chrome 增加 `--window-size=1400,900`，避免 Bridge 面板默认最大化占据整个屏幕。
  - 同步 wb上款 v1.3.20：Edge 自动化期间默认最小化到任务栏。

变更 v2.3.16：
  - 同步 wb上款 v1.3.19：
    * Edge 窗口默认可见（WB_EDGE_VISIBLE=1），便于上款过程人工观察与调试。
    * 分类选择精确匹配当前月份，避免跨月份分类误选。

变更 v2.3.15：
  - AI 去背 贴图 OS (`engine/check_rem.py v2.2.6`)：
    * 修复 DX0339_W 等单张去背后 02_REM_BG 无输出：美图保存路径未切换时，结果会落到 `_temp_rembg/save`。
      check_rem.py 现在从 `TEMP_REMBG/{DX}/02_REM_BG`、`WB_ROOT/_temp_rembg/save`、`WB_ROOT/_temp_rembg/archive`
      三个位置收集 `_cut.png` / `_副本.png`，并把 `_副本.png` 改名为 `_cut.png`。
    * `rembg_one_file` / `batch_rembg` 暂存时额外复制 `source_map.json` 与原始配对文件（1B.png / 1W.png 等），
      让美图 `precheck_pairs` 正确识别 B/W 角色与配对完整性。
    * 修复 `/batch-rembg` 的 BW 过滤 bug：原实现按全局 `dx_files` 判断是否含 BW，导致前一个有 BW 的款会污染后续所有款；
      现在每个 DX 独立判断，只跳过该 DX 自己的 B/W。
    * `engine/_rembg_worker.py` 增加文件日志，输出写入 `D:\Semems WB\_debug\_rembg_worker_YYYYMMDD_HHMMSS.log`。

变更 v2.3.14：
  - AI 去背 贴图 OS (`engine/check_rem.py v2.2.4`)：
    * 修复单张「重新去背」点击后无响应/不生成去背图的问题
    * 补全缺失的 `engine/_rembg_worker.py`：负责在后台运行美图去背并清理锁文件
    * `rembg_one_file` 暂存时把同 DX 所有生成图都放进临时目录，避免美图配对预检跳过

变更 v2.3.12：
  - AI 去背 贴图 OS (`engine/check_rem.py v2.2.3`)：
    * 反相与贴图解耦：反相只生成黑版专用去背图，不再自动调用贴图流水线
    * 贴图由用户单独点击「贴图」或「批量贴图」触发
    * 前端提示文案同步更新，去掉"自动贴图+BW合成"表述

变更 v2.3.11：
  - AI 去背 贴图 OS (`engine/check_rem.py v2.2.2`)：
    * 单张「反相」与「批量反相」统一进入同一个后台任务队列，串行执行
    * 避免连续点击多个反相时并发驱动 Photoshop 导致冲突
    * `/invert-rem` 与 `/batch-invert-rem` 改为立即返回「已加入队列」
    * 前端 `check_rem.js` 轮询 `/batch-invert-result` 获取完成状态
  - 与 wb上款 v1.3.16 联动版本对齐（运行时在线校验 + 终检）

变更 v2.3.10：
  - WB 上款页面新增「刷新已上款」功能：
    * 调用 wb上款 v1.3.14 的 check_online_listed.py
    * 从店小秘 Temu 在线产品页抓取 SKU，提取 DX 款号
    * 在线已上款成为 /upload 页面已上款状态的唯一权威来源
    * 新增 /api/upload/refresh-online-listed 端点
    * upload.html 增加刷新按钮、在线验证徽章、进度面板在线计数
  - 与 wb上款 v1.3.14 联动版本对齐

变更 v2.3.9：
  - 文档与版本同步：更新 SKILL.md / CHANGELOG.md / ARCHITECTURE.md / REPRODUCIBILITY.md
  - 明确与 wb上款 v1.3.13 联动：Edge 透明隐藏、LoginGuard URL 兜底、豆包传图修复
  - 新增 REPRODUCIBILITY.md：一键复现、回滚到 Tag、问题与解决记录

变更 v2.3.7：
  - 修复 /api/upload/progress 计数/百分比异常：只按当前选中的款号统计 done/fail/total
    避免历史已完成记录把 done_count 撑爆 total_count，导致 "280 / 41 (683%)" 这种显示
  - upload.html 进度文案改为：已上款 X / 总 Y  失败 Z  剩余 W，信息更直观
  - AI 生图对比页 (/ai-review) 缩略图 URL 增加 mtime 参数，重新生图后浏览器自动刷新缓存
  - AI 重新生图任务输出使用 PYTHONUNBUFFERED=1，日志实时可见

变更 v2.3.6：
  - check_rem.py 启动后 1 秒自动后台预扫描，把 scan_projects 结果 warming 到缓存
  - 用户首次打开去背预览首页时即可享受热缓存，无需等待 10+ 秒扫描

变更 v2.3.5：
  - Bridge 启动时后台守护 check_rem.py（端口 8766），「去背预览」点击即开
  - 简化 /api/launch-check-rem：不再启动进程/等待扫描，只兜底确认端口就绪
  - 去背预览按钮改为直接 window.open，与 AI 对比按钮一致，瞬时响应
  - 上款页面图片增加 loading="lazy" + decoding="async"，减少初始加载压力
  - 上款页面加载时显示「加载中…」提示，避免空白等待
  - check_rem.py scan_projects 增加 30 秒缓存，大幅提升首页刷新速度

变更 v2.3.4：
  - 修复去背预览页面悬停放大图位置乱跳：
    原 JS 用固定 900x90vh 估算预览图尺寸来定位，与实际渲染尺寸不符。
  - 新逻辑：等原图加载后读取 preview 元素实际 offsetWidth/offsetHeight 再定位；
    水平默认放缩略图右侧，溢出则放左侧；垂直仅做必要平移，不再大幅跳动。

变更 v2.3.3：
  - 修复 Y2 控制台「上款」按钮打不开：原链接使用 http://localhost:8765/upload，
    在 IPv6/localhost 解析异常或 Bridge 仅监听 127.0.0.1 时触发 ERR_CONNECTION_REFUSED。
  - 改为相对路径 /upload，确保与当前 Y2 控制台同域（127.0.0.1:8765），避免 localhost 解析问题。

变更 v2.3.2：
  - 修复 check_rem.py 启动崩溃：print 语句中的 emoji（🔄）在 GBK 控制台导致 UnicodeEncodeError
  - 强制 check_rem.py stdout/stderr 使用 UTF-8，避免 Windows GBK 控制台打印生僻字符/emoji 崩溃
  - 优化「去背预览」启动速度：移除阻塞式 90 秒预扫描，端口 ready 后快速 ping 并立即打开浏览器
  - 「去背预览」尝试在已有 Chrome 窗口中以新标签页打开（new=2）

变更 v2.3.1：
  - Y2 控制台所有日期分类统一按 DX 文件夹建立日期（st_ctime）
  - /upload、/ai-review、去背预览等页面不再按 AI/去背/贴图文件最后更新时间分类
  - 移除 _load_upload_date_map，简化日期来源

变更 v2.3.0：
  - 新增 AI 生图对比页面 (/ai-review)：在同一界面并排对比原图与 AI 生成图
  - 支持单张重新生图，输出到原 DX 文件夹（新图自动命名 DXxxxx_B2.png 等，不覆盖旧图）
  - 支持批量重新生图：勾选多张原图一键并发重跑，调用 Lovart 正常并发能力
  - 重新生图使用 MD5 检测 INBOX 同名冲突，避免错用旧批次原图
  - 状态面板实时显示：款号、Key、已用时间、成功/失败张数、进度、可展开原始日志
  - 状态面板区分「已完成」「部分失败」「失败」，避免 completed + fail_count>0 误导
  - 重新生图与 Lovart 管线统一读取 config/POD AI VIRAL FACTORY v3.md 提示词文件
  - AI 生图对比页默认显示最新日期
  - Y2 控制台所有日期分类统一按 DX 文件夹建立日期（st_ctime），不再按文件最后更新时间

变更 v2.2.1：
  - 修复 /upload 页面款号日期全部归到同一天的问题

变更 v2.2.0：
  - UID/group_id 全链路溯源：从 INBOX 开始绑定唯一 UID 和组 ID
  - 生图阶段写入 .generation_uid_manifest.json 并传给 Lovart
  - 为 AI 图、去背图、贴图成品、BW 合成图生成 .meta.json sidecar
  - 每个 DX 目录维护 uid_map.json，不依赖文件名即可回溯同一组图片
  - 解决 WB去背/registry.py 与 Bridge 双写 .image_registry.json 的冲突
  - WB去背 registry 改为独立写入 .wb_rembg_registry.json

变更 v2.1.9：
  - 新增「强制重新上款」开关
  - /api/batch-upload 支持 force=true，自动从 已上款货号_wb.md 删除对应款号后再启动 wb_listing.py
  - 不修改 wb_listing.py 内部逻辑，保持 wb上款 v1.3.1-stable 稳定版本不变

变更 v2.1.8：
  - /api/batch-upload 改为 --only 精确上款：勾选哪款就上哪款，不会继续后续款
  - 新增 /api/upload/progress 端点，读取 wb_listing.py 写入的 .wb_upload_progress.json
  - /upload 页面拆分为「未上款 / 已上款」两个区域，已上款自动沉底
  - 上款页面新增进度条、当前款、已用时间、平均耗时、预计剩余时间
  - 默认选中最新日期，勾选框仅对未上款卡片生效
  - 修复缩略放大图在屏幕下方时显示不全的问题

变更 v2.1.7：
  - /upload 页面默认选中最新日期
  - 移除后台预生成缩略图（反而拖慢），改用 Flask threaded=True 并发处理缩略图请求
  - 修复批量上传仍提示未配置脚本的问题（代码已更新，需重启 Bridge 生效）
  - /api/batch-upload 改为只启动一次 wb_listing.py，以选中款中最早的 DX 为起点连续处理
  - 修复 lovart_bridge.bat：Chrome 启动时 detached，关闭 CMD 后 Chrome 不再被关闭

变更 v2.1.6：
  - /api/batch-upload 默认对接 E:\Claude code\wb上款\wb_listing.py
  - 批量上款按顺序逐个 DX 启动 wb_listing.py，避免浏览器状态冲突
  - 优化 /upload 页面缩略图加载速度：后台预生成、透明检测、Cache-Control 缓存
  - 修复 upload.html 批量上传后页面刷新逻辑

变更 v2.1.4：
  - 移除 PS贴图控制台，替换为「上款」页面 (/upload)
  - 上款页面展示每款 03_UPLOAD 成品缩略图，按 BW/B/W 分组
  - 支持勾选款号，批量上传按钮，对接 /api/batch-upload
  - 新增 /api/upload/projects、/api/upload/thumb、/api/upload/original 端点
  - 图片显示逻辑与 AI 去背 贴图 页一致：缩略图 + 鼠标悬停放大

变更 v2.1.2：
  - Bridge 内一键启动 check_rem.py / PS贴图 / BW合成 时，子进程窗口最小化，不抢焦点
  - 新增 run_minimized() 工具函数，统一 Windows 最小化启动逻辑

变更 v2.1：
  - 支持命令行参数 --port / --host，便于启动脚本自定义端口
  - 启动时写入 bridge.pid，供 lovart_bridge.bat 优雅停止服务

变更 v2.0：
  - Registry v4 / 血缘引擎 / AutoScan / Lineage API
  - lovart_control.html 控制面板 v2.0

启动: python lovart_bridge.py  →  http://127.0.0.1:8765
"""

import os
import sys
import json
import time
import shutil
import hashlib
import subprocess
import threading
import re
import io
import ctypes
import argparse
import socket
from urllib.request import urlopen
from urllib.error import URLError
from ctypes import wintypes
from pathlib import Path
from datetime import datetime, timedelta
import urllib.parse

try:
    from flask import Flask, jsonify, request, send_file, abort, make_response
except ImportError:
    print("ERROR: Flask not installed. Run: pip install flask")
    sys.exit(1)

# 加载 UID 元数据模块（Bridge 项目内 lib/ 目录）
_WB_META_PATH = Path(__file__).parent / "lib"
if str(_WB_META_PATH) not in sys.path:
    sys.path.insert(0, str(_WB_META_PATH))
try:
    import wb_meta
except Exception as e:
    print(f"WARN: wb_meta 模块加载失败: {e}")
    wb_meta = None

# ============================================================================
# 路径常量
# ============================================================================
# 便携包布局检测：本文件位于 <包根>/ZCodeProject/lovart_bridge.py，
# 若包根下存在 lovart-official/ 即为便携包（如 D:\lovart_bridge），
# 所有路径基于包根解析；否则保持本机原有绝对路径（向后兼容）。
_PKG_ROOT  = Path(__file__).resolve().parent.parent
_PORTABLE  = (_PKG_ROOT / "lovart-official").is_dir()

# 品类根目录单一真源：D:\Semems\wb_category.py + categories.json（缺省 wb=T恤，行为不变）
import sys as _sys
if r"D:\Semems" not in _sys.path:
    _sys.path.insert(0, r"D:\Semems")
from wb_category import root_for as _cat_root, DEFAULT_CAT as _DEFAULT_CAT, id_prefix_for as _cat_prefix

if _PORTABLE:
    BASE_DIR       = _PKG_ROOT / "data"                    # 相当于本机 D:/Semems WB
    LOVART_DIR     = _PKG_ROOT / "lovart-official"
    PYTHON_EXE     = "python"                              # 由启动 bat 保证在 PATH
    PYTHONPATH     = str(_PKG_ROOT / "python_packages")
else:
    BASE_DIR       = _cat_root(_DEFAULT_CAT)               # D:/Semems WB
    LOVART_DIR     = Path("E:/Claude code/lovart-official")
    PYTHON_EXE     = r"C:/Users/Administrator/AppData/Local/Programs/Python/Python311/python.exe"
    PYTHONPATH     = "E:/python_packages"

INBOX_DIR      = BASE_DIR / "01_INBOX"
PROJECTS_DIR   = BASE_DIR / "02_PROJECTS"
REGISTRY_FILE  = BASE_DIR / ".image_registry.json"
LOVART_SCRIPT  = LOVART_DIR / "run_official_v53.py"

HOVER_CACHE    = INBOX_DIR / "_hover_cache"  # 悬停预览缩略图缓存
UID_MANIFEST_FILE = BASE_DIR / ".generation_uid_manifest.json"  # 传给 Lovart 的 UID 清单

AI_TRASH_DIR   = BASE_DIR / "_ai_trash"        # AI 图回收站
AI_THUMB_DIR   = BASE_DIR / "_ai_review_thumbs"  # AI 对比页缩略图缓存

# Lovart 处理记录文件：重新生图时需要清除对应 hash，否则 Lovart 会跳过
LOVART_TRACK_FILE = LOVART_DIR / ".processed_track.json"

# ============================================================================
# Temu 核价（Hermes）项目路径
# ============================================================================
if _PORTABLE:
    PRICING_DIR        = _PKG_ROOT / "Temu自动化" / "核价"
    PRICING_OUTPUT_DIR = BASE_DIR / "核价档案"
else:
    PRICING_DIR        = Path("E:/Claude code/Temu自动化/核价")
    PRICING_OUTPUT_DIR = Path(r"C:\Users\Administrator\Desktop\核价档案")
PRICING_ENTRYPOINT = PRICING_DIR / "entrypoint"
PRICING_MAIN       = PRICING_DIR / "hengjia.py"
PRICING_STATE_FILE = PRICING_DIR / "hengjia_state.json"

# ============================================================================
# Temu 建议零售价填写项目路径
# ============================================================================
if _PORTABLE:
    RETAIL_PRICE_DIR   = _PKG_ROOT / "WB Lovart"
else:
    RETAIL_PRICE_DIR   = Path("E:/Claude code/WB Lovart")
RETAIL_PRICE_SCRIPT = RETAIL_PRICE_DIR / "建议零售价.js"

# ============================================================================
# Temu 报活动项目路径
# ============================================================================
if _PORTABLE:
    ACTIVITY_DIR        = str(_PKG_ROOT / "Temu自动化" / "报活动")
else:
    ACTIVITY_DIR        = 'E:/Claude code/Temu自动化/报活动'
ACTIVITY_ENTRYPOINT = ACTIVITY_DIR + '/entrypoint/run.py'
ACTIVITY_STATE_FILE = ACTIVITY_DIR + '/state/state.json'
ACTIVITY_SELECTION_FILE = ACTIVITY_DIR + '/state/user_selection.json'  # 面板勾选的活动（引擎轮询消费）

# ============================================================================
# 胚衣制作（素材库）路径
# ============================================================================
MATERIAL_DIR   = BASE_DIR / "03_MATERIAL"
# 四大分类：白(W正/B背) / 黑(W正/B背)
PEIYI_CATEGORIES = {
    "W白": MATERIAL_DIR / "W白",
    "B白": MATERIAL_DIR / "B白",
    "W黑": MATERIAL_DIR / "W黑",
    "B黑": MATERIAL_DIR / "B黑",
}

# ============================================================================
# 胚衣素材库品类化（多品类第3步）：按 cat 解析数据根，缺省 wb 行为与改造前一致
# ============================================================================
from wb_category import categories as _cat_all, material_categories as _cat_material_cats

def _peiyi_dirs(cat=None):
    """品类 cat 的胚衣分类目录映射 {分类名: Path}；非法品类返回 None。

    wb（缺省）下返回的目录与 PEIYI_CATEGORIES 完全一致（D:\Semems WB_MATERIAL\...）。
    """
    cat = cat or _DEFAULT_CAT
    if cat not in _cat_all():
        return None
    if _PORTABLE and cat == _DEFAULT_CAT:
        base = MATERIAL_DIR                  # 便携包布局：保持原 BASE_DIR/"03_MATERIAL"
    else:
        base = _cat_root(cat) / "03_MATERIAL"
    return {name: base / name for name in _cat_material_cats(cat)}


def _request_cat():
    """从当前请求解析品类：query string / form / JSON body 的 cat 字段，缺省 wb。"""
    cat = request.args.get('cat') or request.form.get('cat')
    if not cat:
        try:
            cat = (request.get_json(silent=True) or {}).get('cat')
        except Exception:
            cat = None
    return cat or _DEFAULT_CAT


def _peiyi_request_dirs():
    """从当前请求解析品类并返回 (cat, dirs)；cat 未注册时 dirs 为 None，端点据此返回 400。"""
    cat = _request_cat()
    return cat, _peiyi_dirs(cat)


def _peiyi_cat_qs(cat):
    """回传给前端 URL 的品类串：wb（缺省）不追加，保持旧 URL 逐字节不变。"""
    return '' if cat == _DEFAULT_CAT else '?cat=' + urllib.parse.quote(cat)


# ============================================================================
# 控制台品类化（多品类第5步）：INBOX 卡片墙 / 生图 / AI对比 / 上款 按 cat 解析。
# wb（缺省）返回值与既有全局常量完全一致，保证老流程逐字节不变。
# ============================================================================

def _cat_ctx(cat=None):
    """按品类解析数据目录上下文 dict；cat 缺省 wb。

    返回键：cat / prefix / inbox / projects / hover_cache / ai_trash / ai_thumb / upload_thumb
    wb 下全部为既有全局常量对象本身（不是重建路径），行为与改造前一致。
    """
    cat = cat or _DEFAULT_CAT
    if cat == _DEFAULT_CAT:
        return {
            "cat": cat,
            "prefix": "DX",
            "inbox": INBOX_DIR,
            "projects": PROJECTS_DIR,
            "hover_cache": HOVER_CACHE,
            "ai_trash": AI_TRASH_DIR,
            "ai_thumb": AI_THUMB_DIR,
            "upload_thumb": UPLOAD_THUMB_DIR,
        }
    root = _cat_root(cat)          # 未注册品类在此抛 KeyError，调用方先校验
    return {
        "cat": cat,
        "prefix": _cat_prefix(cat),
        "inbox": root / "01_INBOX",
        "projects": root / "02_PROJECTS",
        "hover_cache": root / "01_INBOX" / "_hover_cache",
        "ai_trash": root / "_ai_trash",
        "ai_thumb": root / "_ai_review_thumbs",
        "upload_thumb": root / "_upload_thumbs",
    }


def _resolve_request_cat():
    """解析并校验当前请求的 cat。返回 (cat, error_response)；非法 cat 时 cat=None。"""
    cat = _request_cat()
    if cat not in _cat_all():
        return None, (jsonify({
            "ok": False,
            "error": f"未知品类: {cat!r}（已注册: {sorted(_cat_all())}）",
        }), 400)
    return cat, None


def _dx_re(prefix: str) -> re.Pattern:
    """按品类编号前缀编译项目目录名正则（wb→^DX\\d+(?:BW|B|W)?$，与原有字面量一致）。"""
    return re.compile(rf"^{re.escape(prefix)}\d+(?:BW|B|W)?$")


def _cat_not_ready(cat: str, feature: str, reason: str):
    """非 wb 品类在某功能尚未接入时的统一明确报错（避免静默写错目录）。"""
    return jsonify({
        "ok": False,
        "error": f"品类 {cat} 的「{feature}」尚未接入：{reason}",
    }), 400


def _gen_paths(cat=None):
    """生图管线的品类路径上下文；cat 缺省 wb。

    返回键：cat / prefix / inbox / projects / registry / uid_manifest / wb_registry / prompt
    wb（缺省）下全部为既有全局常量对象本身，行为与改造前逐字节一致。
    """
    cat = cat or _DEFAULT_CAT
    if cat == _DEFAULT_CAT:
        return {
            "cat": cat,
            "prefix": "DX",
            "inbox": INBOX_DIR,
            "projects": PROJECTS_DIR,
            "registry": REGISTRY_FILE,
            "uid_manifest": UID_MANIFEST_FILE,
            "wb_registry": WB_REGISTRY_FILE,
            # wb 提示词：仅重新生图时注入 LOVART_PROMPT_FILE，普通生图由脚本默认读取，
            # 保持原有行为不变
            "prompt": LOVART_DIR / "config" / "POD AI VIRAL FACTORY v3.md",
            "always_prompt": False,
        }
    root = _cat_root(cat)          # 未注册品类在此抛 KeyError，调用方先校验
    return {
        "cat": cat,
        "prefix": _cat_prefix(cat),
        "inbox": root / "01_INBOX",
        "projects": root / "02_PROJECTS",
        "registry": root / ".image_registry.json",
        "uid_manifest": root / ".generation_uid_manifest.json",
        "wb_registry": root / "WB_REGISTRY" / "registry.json",
        # 卫衣提示词：LOVART_PROMPT_FILE 覆盖视为完整 prompt（不再拼接脚本内置 CONCEPT），
        # 故文件内自带 concrete request 段
        "prompt": LOVART_DIR / "config" / f"POD AI VIRAL FACTORY v3 - {'Hoodie' if cat == 'hoodie' else cat}.md",
        "always_prompt": True,
    }
# 各分类底色（JPG 输出）：白胚衣用白底、黑胚衣用黑底
PEIYI_BG = {
    "W白": (255, 255, 255),
    "B白": (255, 255, 255),
    "W黑": (0, 0, 0),
    "B黑": (0, 0, 0),
}
PEIYI_SIZE = (1340, 1785)   # 目标分辨率
PEIYI_DPI  = (72, 72)       # 目标 DPI
PEIYI_ALLOWED_EXT = ('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.tif', '.tiff')
# 遮罩功能生成的侧车文件后缀：这些不显示在素材库画廊，只在“预览遮罩”时单独查看
PEIYI_MASK_SUFFIXES = (
    '_occluder.png', '_occluder_mask.png', '_body_mask.png',
    '_parse.png', '_alpha.png',
    # 卫衣抽绳遮罩（白=绳子）及其衍生预览/分层示意：不进主画廊
    '_drawstring_mask.png', '_drawstring_preview.png', '_drawstring_layered_demo.png',
    # 手动 PS 遮罩侧车
    '_manual.png',
)
# 每张素材侧车(.meta.json)记录的5个贴图参数，与 胚衣参数表_模板.csv 第5-9列一致
PEIYI_META_FIELDS = [
    ("width", "缩放后宽(px)", 0),
    ("height", "缩放后高(px)", 0),
    ("rotation", "旋转角度(负=逆/正=顺)", 0),
    ("highest_y", "最高像素点y", 0),
    ("center_x", "中心点x", 670),   # 670 = 1340 画布宽中点
]
PEIYI_META_KEYS = [k for k, _, _ in PEIYI_META_FIELDS]

# 贴图（AI 去背贴图）相关常量
if _PORTABLE:
    MOCKUP_PY  = "python"
    MOCKUP_ROOT = _PKG_ROOT                       # white_t_mockup 所在目录（运行 -m white_t_mockup）
    PY_PACKAGES = PYTHONPATH
else:
    MOCKUP_PY = Path(r"C:/Users/Administrator/AppData/Local/Programs/Python/Python311/python.exe")
    MOCKUP_ROOT = Path(r"E:/Kimi Code")              # white_t_mockup 所在目录（运行 -m white_t_mockup）
    PY_PACKAGES = "E:/python_packages"
MOCKUP_OUT = BASE_DIR / "03_MOCKUP_OUT"          # 贴图成品输出
ZCODE_PROJECT = Path(__file__).resolve().parent  # 本文件所在目录（peiyi_mask / tpl_generator 在此）

# ── 贴图常量品类化（多品类第4步）────────────────────────────────────────────
# wb（缺省）返回值与改造前常量逐字节一致；hoodie 指向 D:\Semems Hoodie 下对应位置，
# 只解析不创建（目录可不存在，等卫衣模板标定后落地）。
def tpl_root_for(cat=None):
    """_tpl/<款名>/ 扭曲素材根（按品类）。"""
    cat = cat or _DEFAULT_CAT
    if _PORTABLE and cat == _DEFAULT_CAT:
        return _PKG_ROOT / "1胚衣" / "_tpl"
    if cat == _DEFAULT_CAT:
        return Path(r"D:\Semems\1胚衣\_tpl")
    return _cat_root(cat) / "1胚衣" / "_tpl"

def csv_path_for(cat=None):
    """胚衣参数表 CSV（按品类）。"""
    cat = cat or _DEFAULT_CAT
    if _PORTABLE and cat == _DEFAULT_CAT:
        return _PKG_ROOT / "docs" / "胚衣参数表_模板.csv"
    if cat == _DEFAULT_CAT:
        return Path(r"E:\Kimi Code\docs\胚衣参数表_模板.csv")
    return _cat_root(cat) / "docs" / "胚衣参数表_模板.csv"

def mockup_out_for(cat=None):
    """贴图成品输出目录（按品类）；wb 与 MOCKUP_OUT 完全一致。"""
    cat = cat or _DEFAULT_CAT
    if cat == _DEFAULT_CAT:
        return MOCKUP_OUT
    return _cat_root(cat) / "03_MOCKUP_OUT"

TPL_ROOT = tpl_root_for()      # 兼容旧引用：wb 值与原常量逐字节一致
CSV_PATH = csv_path_for()

# ============================================================================
# 其它散点路径（便携包/本机双套）
# ============================================================================
if _PORTABLE:
    WB_REGISTRY_FILE   = BASE_DIR / "WB_REGISTRY" / "registry.json"
    CHECK_REM_SCRIPT   = BASE_DIR / "04_OS" / "engine" / "check_rem.py"
    WB_LISTING_DIR     = _PKG_ROOT / "wb上款"
    TEMU_ANALYSIS_DIR  = _PKG_ROOT / "temu分析"
    KIMI_SCRIPTS_DIR   = _PKG_ROOT / "scripts"
else:
    WB_REGISTRY_FILE   = BASE_DIR / "WB_REGISTRY" / "registry.json"
    CHECK_REM_SCRIPT   = BASE_DIR / "04_OS" / "engine" / "check_rem.py"
    WB_LISTING_DIR     = Path(r"E:\Claude code\wb上款")
    TEMU_ANALYSIS_DIR  = Path(r"E:/Kimi Code/temu分析")
    KIMI_SCRIPTS_DIR   = Path(r"E:/Kimi Code/scripts")

def white_t_presets_for(cat=None):
    """white_t_mockup presets.json（按品类）；wb 与原 WHITE_T_PRESETS 逐字节一致。"""
    cat = cat or _DEFAULT_CAT
    if _PORTABLE and cat == _DEFAULT_CAT:
        return _PKG_ROOT / "white_t_mockup" / "presets.json"
    if cat == _DEFAULT_CAT:
        return Path(r"E:/Kimi Code/white_t_mockup/presets.json")
    return _cat_root(cat) / "white_t_mockup" / "presets.json"

WHITE_T_PRESETS = white_t_presets_for()

# ── 贴图引擎插件化（多品类第4步）────────────────────────────────────────────
if str(ZCODE_PROJECT) not in sys.path:
    sys.path.insert(0, str(ZCODE_PROJECT))
from engine.garment_plugin import plugin_for as _garment_plugin_for, MockupConfig as _MockupConfig

def _mockup_cfg(cat=None):
    """按品类组装贴图插件配置；wb 各字段与改造前常量逐字节一致。"""
    cat = cat or _DEFAULT_CAT
    return _MockupConfig(
        mockup_py=MOCKUP_PY,
        mockup_root=MOCKUP_ROOT,
        mockup_out=mockup_out_for(cat),
        py_packages=PY_PACKAGES,
        zcode_project=ZCODE_PROJECT,
        tpl_root=tpl_root_for(cat),
        presets_path=white_t_presets_for(cat),
    )


def _single_thread_env(base_env):
    """准备子进程环境：禁用 OpenMP/MKL 多线程，并清理无效 PATH 项（如 cv2 留下的裸驱动器号）。"""
    env = dict(base_env)
    # cv2 初始化会往 PATH 追加形如 "E;" 的裸驱动器号，导致子进程 Python 找不到 DLL 而静默退出
    raw_path = env.get("PATH", "")
    cleaned = []
    for p in raw_path.split(os.pathsep):
        p = p.strip()
        if not p:
            continue
        # 丢弃纯驱动器号项（如 "E" 或 "E:"）
        if len(p.rstrip(":")) == 1 and p[0].isalpha():
            continue
        cleaned.append(p)
    env["PATH"] = os.pathsep.join(cleaned)
    env.update({
        "OMP_NUM_THREADS": "1",
        "OPENBLAS_NUM_THREADS": "1",
        "MKL_NUM_THREADS": "1",
        "VECLIB_MAXIMUM_THREADS": "1",
        "NUMEXPR_NUM_THREADS": "1",
        # 彻底关闭 cv2 内部多线程：后台/服务进程里 cv2 的 warp/remap 等多线程操作
        # 偶发段错误（静默 rc!=0、无输出），关闭后所有 cv2 子进程稳定
        "OPENCV_DISABLE_THREADING": "1",
    })
    return env
MOCKUP_OUT.mkdir(parents=True, exist_ok=True)


# ============================================================================
# 工具函数：处理 Lovart 去重记录
# ============================================================================
def _compute_sha256(path: str) -> str:
    """计算文件 SHA256（与 Lovart run_official_v53.py 一致）"""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _load_lovart_track() -> list:
    """读取 Lovart 处理记录 track 文件"""
    if not LOVART_TRACK_FILE.exists():
        return []
    try:
        return json.loads(LOVART_TRACK_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []


def _save_lovart_track(track: list):
    """写入 Lovart 处理记录 track 文件"""
    try:
        LOVART_TRACK_FILE.write_text(json.dumps(track, indent=2, ensure_ascii=False), encoding='utf-8')
    except Exception as e:
        print(f"[WARN] 保存 Lovart track 失败: {e}", flush=True)


def _remove_from_lovart_track(img_path: Path) -> int:
    """从 Lovart track 中移除指定图片的 hash / name+size 记录，强制下次重新处理。
    返回移除的条目数。"""
    if not img_path.exists():
        return 0
    try:
        img_hash = _compute_sha256(str(img_path))
        img_size = img_path.stat().st_size
        img_name = img_path.name
    except Exception:
        return 0
    track = _load_lovart_track()
    orig_len = len(track)

    def _matches(e):
        # hash 唯一标识（优先）
        if img_hash and e.get("hash") == img_hash:
            return True
        # 兼容旧记录：同名同尺寸也清除，避免换批次后仍被误判为已处理
        if e.get("name") == img_name and e.get("size") == img_size:
            return True
        return False

    track = [e for e in track if not _matches(e)]
    removed = orig_len - len(track)
    if removed:
        _save_lovart_track(track)
    return removed


# ============================================================================
# 工具函数：Windows 下最小化启动子进程（不抢焦点）
# ============================================================================
def run_minimized(cmd, cwd=None, wait=False, no_console=False, env=None):
    """以最小化/不激活窗口启动子进程，用于 check_rem / PS 贴图等任务。

    参数:
      no_console: True 时使用 CREATE_NO_WINDOW，不弹控制台黑窗，同时把 stdout/stderr 重定向到 DEVNULL。
                  适用于 wb_listing.py / check_online_listed.py 这种自己有日志文件的后台任务。
      env: 额外的环境变量 dict，会合并到当前进程环境（不覆盖未指定的变量）。
            用于按品类向 wb_listing.py 注入 WB_LISTING_CAT 等标识。
    """
    import subprocess
    import ctypes
    from ctypes import wintypes

    STARTUPINFO = subprocess.STARTUPINFO
    SW_SHOWMINNOACTIVE = 7
    STARTF_USESHOWWINDOW = 1

    si = STARTUPINFO()
    si.dwFlags |= STARTF_USESHOWWINDOW
    si.wShowWindow = SW_SHOWMINNOACTIVE

    kwargs = {
        "startupinfo": si,
    }
    if no_console:
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        kwargs["stdout"] = subprocess.DEVNULL
        kwargs["stderr"] = subprocess.DEVNULL
    else:
        kwargs["creationflags"] = subprocess.CREATE_NEW_CONSOLE
    if cwd:
        kwargs["cwd"] = str(cwd)
    if env:
        _merged = dict(os.environ)
        _merged.update(env)
        kwargs["env"] = _merged

    proc = subprocess.Popen(cmd, **kwargs)
    if wait:
        proc.wait()
        return proc
    return proc


def _port_ready(host, port, timeout=2):
    """检查指定端口是否已监听。"""
    start = time.time()
    while time.time() - start < timeout:
        try:
            with socket.create_connection((host, port), timeout=1):
                return True
        except Exception:
            time.sleep(0.2)
    return False


# 去背预览服务：每个品类一个独立实例（单进程服务单一品类，互不串类）
_CHECK_REM_INSTANCES = [("wb", 8766), ("hoodie", 8767)]
_CHECK_REM_PORT_FOR_CAT = {cat: port for cat, port in _CHECK_REM_INSTANCES}


def _check_rem_ensure(cat, port):
    """拉起（若未运行）指定品类/端口的 check_rem.py。"""
    import subprocess
    script = CHECK_REM_SCRIPT
    if not script.exists():
        print(f"  [check_rem daemon] 脚本不存在，跳过 {cat}@{port}", flush=True)
        return
    if _port_ready("127.0.0.1", port, timeout=3):
        return
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    log_path = script.parent / f"check_rem_daemon_{port}.log"
    try:
        env = dict(os.environ)
        env["OPEN_BROWSER"] = "0"
        logf = open(log_path, "a", encoding="utf-8")
        logf.write(f"[{ts}] 端口 {port}({cat}) 未就绪，启动 check_rem.py ...\n")
        logf.flush()
        proc = subprocess.Popen(
            [sys.executable, str(script), "--cat", cat, "--port", str(port)],
            cwd=str(script.parent),
            env=env,
            creationflags=subprocess.CREATE_NO_WINDOW,
            stdout=logf,
            stderr=subprocess.STDOUT,
        )
        print(f"  [check_rem daemon] 已启动 check_rem.py cat={cat} port={port} (PID={proc.pid})", flush=True)
    except Exception as e:
        print(f"  [check_rem daemon] 启动失败 cat={cat} port={port}: {e}", flush=True)


def _check_rem_daemon():
    """后台守护线程：保持各品类 check_rem.py 常驻（wb@8766 / hoodie@8767）。

    使用 CREATE_NO_WINDOW 启动，避免依赖桌面窗口（无头/后台环境下也能拉起）；
    输出重定向到 check_rem_daemon_{port}.log 便于排查。每 5 秒检测一次端口，
    若 check_rem 崩溃退出会自动重拉，实现自愈。
    """
    while True:
        try:
            for cat, port in _CHECK_REM_INSTANCES:
                _check_rem_ensure(cat, port)
            time.sleep(5)
        except Exception:
            time.sleep(5)


# ============================================================================
# Flask App
# ============================================================================
app = Flask(__name__, static_folder=None)

# ============================================================================
# 全局任务状态（持久化到磁盘，重启桥接后仍可见）
# ============================================================================
STATE_FILE = BASE_DIR / ".last_task_state.json"

task_state = {
    "status": "idle",            # idle | running | completed | error | cancelled
    "progress": "",
    "started_at": None,
    "completed_at": None,
    "log": [],
    "selected_files": [],
    "groups_processed": 0,
    "groups_total": 0,
    "task_id": None,
}


def _save_state():
    """将当前 task_state 持久化到磁盘"""
    try:
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(task_state, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _load_state():
    """从磁盘恢复上次的 task_state"""
    global task_state
    if STATE_FILE.exists():
        try:
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                saved = json.load(f)
            status = saved.get("status", "")
            if status in ("completed", "error", "idle"):
                task_state = saved
            elif status == "running":
                # 进程已不在，标记为中断
                saved["status"] = "error"
                saved["progress"] = "⚠️ 上次任务未完成（服务重启中断）"
                saved["completed_at"] = datetime.now().isoformat()
                saved["log"].append(f"[{datetime.now().strftime('%H:%M:%S')}] ⚠️ 服务重启，任务中断")
                task_state = saved
        except Exception:
            pass


# ============================================================================
# 核价任务状态（Hermes / Temu 核价）
# ============================================================================
pricing_task = {
    "status": "idle",          # idle | running | completed | error | stopped
    "mode": None,              # full | no-submit | continue | retry | export
    "task_label": "",
    "started_at": None,
    "completed_at": None,
    "proc": None,
    "log": [],
    "log_index": 0,            # 前端已读取到的位置
    "processed_pages": 0,
    "elapsed_sec": 0,
    "page_records": [],
}
pricing_lock = threading.Lock()


# ============================================================================
# 报活动任务状态
# ============================================================================
activity_task = {
    "status": "idle",          # idle | running | completed | error | stopped
    "started_at": None,
    "completed_at": None,
    "proc": None,
    "log": [],
    "log_index": 0,            # 前端已读取到的位置
}
activity_lock = threading.Lock()


# ============================================================================
# 建议零售价填写任务状态
# ============================================================================
retail_price_task = {
    "status": "idle",          # idle | running | completed | error | stopped
    "task_label": "",
    "started_at": None,
    "completed_at": None,
    "proc": None,
    "log": [],
    "log_index": 0,            # 前端已读取到的位置
    "elapsed_sec": 0,
}
retail_price_lock = threading.Lock()


_lock = threading.Lock()


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def log(msg: str):
    """向任务日志追加一条带时间戳的消息"""
    ts = datetime.now().strftime("%H:%M:%S")
    task_state["log"].append(f"[{ts}] {msg}")


def get_python() -> str:
    """返回可用的 Python 可执行路径"""
    if os.path.exists(PYTHON_EXE):
        return PYTHON_EXE
    return "python"


# ---------------------------------------------------------------------------
# Registry 操作
# ---------------------------------------------------------------------------

def load_registry(path=None) -> dict:
    """加载 .image_registry.json，不存在则返回空骨架。path 缺省为 wb 全局 REGISTRY_FILE（行为不变）。"""
    path = path or REGISTRY_FILE
    if path.exists():
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {"version": 3, "images": {}, "groups": {}, "uid_index": {}, "name_index": {}}
    return {"version": 3, "images": {}, "groups": {}, "uid_index": {}, "name_index": {}}


def save_registry(reg: dict, path=None):
    """原子写入 registry（先写 .tmp 再 rename，防止写半截）。path 缺省为 wb 全局 REGISTRY_FILE（行为不变）。"""
    path = path or REGISTRY_FILE
    tmp = path.with_suffix(".json.tmp")
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(reg, f, indent=2, ensure_ascii=False)
    tmp.replace(path)


def ensure_registry_v4(reg: dict) -> dict:
    """确保 registry 包含 v4 字段（含溯源信息）"""
    v = reg.get("version", 1)
    # v3 fields
    reg.setdefault("groups", {})
    reg.setdefault("uid_index", {})
    reg.setdefault("name_index", {})
    # v4 provenance fields
    reg.setdefault("provenance", {"tree": {}, "broken": []})
    if v < 4:
        reg["version"] = 4
        # 为所有现有图片添加溯源字段
        for md5, entry in reg.get("images", {}).items():
            _add_provenance_fields(entry)
    return reg


# ── 溯源字段 ────────────────────────────────────────

PROVENANCE_FIELDS = {
    "source_md5": "",       # 来源图片的 MD5
    "source_type": "",      # inbox | ai_gen | rembg | upload
    "root_md5": "",         # 最原始 INBOX 图片的 MD5
    "root_name": "",        # 最原始 INBOX 文件名
    "derived_md5s": [],     # 由此图片衍生出的 MD5 列表
    "lineage_status": "",   # confirmed | inferred | missing
}


def _add_provenance_fields(entry: dict):
    """为单条图片记录添加溯源字段"""
    for field, default in PROVENANCE_FIELDS.items():
        entry.setdefault(field, default)


def _register_provenance(reg: dict, child_md5: str, parent_md5: str, source_type: str,
                          lineage_status: str = "inferred"):
    """记录 child_md5 由 parent_md5 通过 source_type 方式生成。
    
    lineage_status:
      confirmed - Hook 实时记录（可信）
      inferred  - Scanner 推断（需验证）
      missing   - 断链
    """
    child = reg["images"].get(child_md5)
    parent = reg["images"].get(parent_md5)
    if not child or not parent:
        return

    _add_provenance_fields(child)
    _add_provenance_fields(parent)

    child["source_md5"] = parent_md5
    child["source_type"] = source_type
    child["lineage_status"] = lineage_status
    # root_md5 继承：如果 parent 有 root_md5 则继承，否则 parent 自己就是 root
    child["root_md5"] = parent.get("root_md5") or parent_md5
    child["root_name"] = parent.get("root_name") or parent.get("inbox_original_name") or parent.get("original_name", "")

    # 在 parent 的 derived 列表中添加 child
    if parent_md5 not in parent["derived_md5s"]:
        parent["derived_md5s"].append(child_md5)

    # 更新 provenance tree 索引
    tree = reg.setdefault("provenance", {}).setdefault("tree", {})
    tree.setdefault(parent_md5, [])
    if child_md5 not in tree[parent_md5]:
        tree[parent_md5].append(child_md5)


# ── 批量扫描：建立现有文件的溯源链 ─────────────────

def scan_provenance():
    """扫描所有 DX 文件夹，通过文件 stem 精确匹配建立血缘关系。
    
    规则（不改文件名，用现有命名语义）：
      AI:       DX{N}_{role}.png                   → 父级 = INBOX 原图（source_map）
      去背:     DX{N}_{role}_cut.png                → 父级 = AI 图（去掉 _cut）
      去背变体:  DX{N}_{Chinese}{role}_cut.png      → 父级 = AI 图（去掉中文+_cut）
      贴图:     DX{N}_{role}_XXX.jpg                → 父级 = 去背图（如存在），否则 = AI 图
    """
    reg = load_registry()
    reg = ensure_registry_v4(reg)
    count = 0

    # 加载 Lovart registry（用于 AI → 原图）
    lovart_reg = {}
    lr_path = WB_REGISTRY_FILE
    if lr_path.exists():
        try:
            with open(lr_path, 'r', encoding='utf-8') as f:
                lovart_reg = json.load(f)
        except Exception:
            pass

    # 预计算：同一 DX 内所有文件的 md5 索引（dx_dir内的相对文件名 → md5）
    # 避免重复 compute_md5
    def _index_dir(dirpath):
        idx = {}
        if dirpath.exists():
            for f in os.listdir(dirpath):
                fp = dirpath / f
                if fp.is_file():
                    idx[f] = compute_md5(str(fp))
        return idx

    for d in sorted(os.listdir(PROJECTS_DIR)):
        if not d.startswith('DX'):
            continue
        ai_dir = PROJECTS_DIR / d / "01_AI"
        rem_dir = PROJECTS_DIR / d / "02_REM_BG"
        up_dir  = PROJECTS_DIR / d / "03_UPLOAD"

        ai_idx = _index_dir(ai_dir)
        rem_idx = _index_dir(rem_dir)
        up_idx = _index_dir(up_dir)

        # ── 1) AI 图 → INBOX 原图 ──
        sm_path = PROJECTS_DIR / d / "source_map.json"
        src_id_map = {}
        if sm_path.exists():
            try:
                with open(sm_path, 'r', encoding='utf-8') as f:
                    sm = json.load(f)
                for src in sm.get("sources", []):
                    src_id_map[src.get("file", "")] = src.get("src_id", "")
            except Exception:
                pass

        for fname, md5 in ai_idx.items():
            if md5 not in reg.get("images", {}):
                continue
            entry = reg["images"].get(md5, {})
            _add_provenance_fields(entry)
            if entry.get("source_md5"):
                continue

            # source_map → Lovart registry → original name
            src_id = src_id_map.get(fname, "")
            if src_id and src_id in lovart_reg:
                orig_name = lovart_reg[src_id].get("original_name", "")
                orig_md5 = reg.get("name_index", {}).get(orig_name, "")
                if orig_md5 and orig_md5 in reg.get("images", {}):
                    _register_provenance(reg, md5, orig_md5, "ai_gen")
                    count += 1
                    continue

            # 后备：inbox_original_name 匹配
            for img_md5, img_info in reg.get("images", {}).items():
                if img_info.get("inbox_original_name") and img_info["inbox_original_name"] in fname:
                    _register_provenance(reg, md5, img_md5, "ai_gen")
                    count += 1
                    break

        # ── 2) 去背图 → AI 图 ──
        # 规则: DX{N}_{role}_cut.png → 去掉 _cut → DX{N}_{role}.png
        #       DX{N}_黑{role}_cut.png → 去掉中文再去掉 _cut → DX{N}_{role}.png
        for fname, md5 in rem_idx.items():
            if md5 not in reg.get("images", {}):
                continue
            entry = reg["images"].get(md5, {})
            _add_provenance_fields(entry)
            if entry.get("source_md5"):
                continue

            stem = fname[:-len("_cut.png")] if fname.endswith("_cut.png") else fname.rsplit('.', 1)[0]
            # 尝试直接匹配: stem → AI 图
            ai_stem = re.sub(r'[\u4e00-\u9fff]+', '', stem)
            ai_candidate = f"{ai_stem}.png"
            if ai_candidate in ai_idx:
                ai_md5 = ai_idx[ai_candidate]
                if ai_md5 in reg.get("images", {}):
                    _register_provenance(reg, md5, ai_md5, "rembg")
                    count += 1
                    continue

        # ── 3) 贴图图 → 去背图 / AI 图 ──
        # 规则: DX{N}_{role}_XXX.jpg → 去掉中文后缀 → DX{N}_{role}
        #       先找 DX{N}_{role}_cut.png（去背），再找 DX{N}_{role}.png（AI）
        for fname, md5 in up_idx.items():
            if md5 not in reg.get("images", {}):
                continue
            entry = reg["images"].get(md5, {})
            _add_provenance_fields(entry)
            if entry.get("source_md5"):
                continue

            # 提取基础 stem：去掉文件后缀和中文部分
            stem = fname.rsplit('.', 1)[0]
            base_stem = re.sub(r'[\u4e00-\u9fff].*$', '', stem)

            # 优先找去背图
            cut_candidate = f"{base_stem}_cut.png"
            if cut_candidate in rem_idx:
                cut_md5 = rem_idx[cut_candidate]
                if cut_md5 in reg.get("images", {}):
                    _register_provenance(reg, md5, cut_md5, "upload")
                    count += 1
                    continue

            # 其次找 AI 图
            ai_candidate = f"{base_stem}.png"
            if ai_candidate in ai_idx:
                ai_md5 = ai_idx[ai_candidate]
                if ai_md5 in reg.get("images", {}):
                    _register_provenance(reg, md5, ai_md5, "upload")
                    count += 1
                    continue

    save_registry(reg)
    return count


def compute_md5(filepath: str) -> str:
    """计算文件的 MD5 哈希"""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(64 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def get_next_uid(reg: dict, prefix: str = "DX") -> str:
    """生成下一个 UID: {prefix}{YYYYMMDD}_{NNNN}，每日从 0001 开始。prefix 缺省 DX（wb 行为不变）"""
    today = datetime.now().strftime("%Y%m%d")
    prefix = f"{prefix}{today}_"
    max_seq = 0
    for uid in reg.get("uid_index", {}):
        if uid.startswith(prefix):
            try:
                seq = int(uid.rsplit("_", 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
    return f"{prefix}{max_seq + 1:04d}"


def get_next_group_id(reg: dict) -> str:
    """生成下一个 group_id: G_{NNNNN}"""
    max_num = 0
    for gid in reg.get("groups", {}):
        if gid.startswith("G_"):
            try:
                num = int(gid.split("_", 1)[-1])
                max_num = max(max_num, num)
            except (ValueError, IndexError):
                pass
    return f"G_{max_num + 1:05d}"


def find_reusable_group(reg: dict, group_number: int, roles_md5: dict, projects_dir: Path = None):
    """同组链接（生图点击时即建立）：编号 group_number 在注册表已有组时，判断本次文件能否并入旧组。

    解决：同一设计的两半分批生图（先生 15W、再生 15B）旧逻辑每次都开新 group →
    新 DX 文件夹 → 同设计被拆到两个文件夹（HX0137BW / HX0144B 事故）。

    roles_md5: {角色: md5}（本组本次选中待生成的文件，角色=B/W/BW/WB）
    组级判定（防 DX0455 式交叉污染——不同设计复用同编号绝不能合并）：
      - 任一角色在旧组已存在且 MD5 不同 → 新设计复用编号 → 返回 None（开新组）；
      - BW/WB 是双面完整图，自带两面，不可能是"后到的另一半"：仅当旧组角色集合
        完全相同且 MD5 全同（同图重生成）才并入，其余一律开新组
        （HX0065W 事故：13bw 新设计复用编号被并进 13W 的组）；
      - 旧组缺本次某角色（后到的另一半）或全部角色 MD5 相同（同图重生）→ 并入旧组；
      - 旧组未落过 dx_folder 或该文件夹已不在磁盘 → None（开了新组才能落新文件夹）。
    多个候选旧组取最新创建的。返回 (gid, dx_folder) 或 (None, None)。
    """
    projects_dir = projects_dir or PROJECTS_DIR
    FULL_SIDES = {"BW", "WB"}  # 双面完整图角色
    candidates = {}  # gid -> {"roles": {role: md5}, "created": str, "dx_folder": str}
    for md5_key, info in reg.get("images", {}).items():
        if info.get("design_number") != group_number:
            continue
        gid = info.get("group_id") or ""
        if not gid:
            continue
        c = candidates.setdefault(gid, {"roles": {}, "created": "", "dx_folder": ""})
        c["roles"][info.get("role", "")] = md5_key
        ginfo = reg.get("groups", {}).get(gid) or {}
        c["created"] = ginfo.get("created", "")
        c["dx_folder"] = ginfo.get("dx_folder", "")
    if not candidates:
        return None, None
    # 最新创建的优先
    for gid in sorted(candidates, key=lambda g: candidates[g]["created"], reverse=True):
        c = candidates[gid]
        dx = c["dx_folder"]
        if not dx or not (projects_dir / dx).is_dir():
            continue
        if any(r in FULL_SIDES for r in roles_md5) or any(r in FULL_SIDES for r in c["roles"]):
            # 涉及双面图：只有"同角色集合 + 同 MD5"的同图重生成才允许并入
            if c["roles"] == roles_md5:
                return gid, dx
            continue
        conflict = any(role in c["roles"] and c["roles"][role] != md5
                       for role, md5 in roles_md5.items())
        if conflict:
            continue   # 同角色不同内容 = 新设计复用编号，绝不能并入
        return gid, dx
    return None, None


# ---------------------------------------------------------------------------
# 文件名自动大写（b→B, w→W）
# ---------------------------------------------------------------------------

def auto_uppercase_inbox():
    """将 INBOX 中后缀为小写 b/w 的文件名改为大写，如 17b.png → 17B.png。
    Windows NTFS 保留大小写但查找不区分大小写，直接 rename 即可。
    """
    if not INBOX_DIR.exists():
        return 0
    count = 0
    # 匹配: 数字 + B/W/BW/WB + .png (不区分大小写)
    pattern = re.compile(r'^(\d+)([bw]+)(\.(png|jpg|jpeg|webp))$', re.IGNORECASE)
    for fname in list(os.listdir(INBOX_DIR)):
        if fname.startswith('_'):
            continue
        m = pattern.match(fname)
        if not m:
            continue
        num = m.group(1)
        suffix = m.group(2)
        ext = m.group(3)
        upper = suffix.upper()
        if suffix == upper:
            continue  # 已是大写
        new_name = f"{num}{upper}{ext}"
        src = INBOX_DIR / fname
        dst = INBOX_DIR / new_name
        # Windows: 同一文件，直接 rename 改变显示大小写
        src.rename(dst)
        count += 1
        print(f"  [AutoUppercase] {fname} → {new_name}")
    return count


# ---------------------------------------------------------------------------
# 悬停预览缩略图（500px 缓存）
# ---------------------------------------------------------------------------

def get_hover_thumb(filename: str, inbox_dir: Path = None, cache_dir: Path = None) -> Path:
    """生成或返回 500px 宽度的预览缓存图；inbox_dir/cache_dir 缺省为 wb 常量（行为不变）"""
    from PIL import Image
    inbox_dir = inbox_dir or INBOX_DIR
    cache_dir = cache_dir or HOVER_CACHE
    safe = os.path.basename(filename)
    src = inbox_dir / safe
    if not src.exists():
        return None
    cache_dir.mkdir(parents=True, exist_ok=True)
    thumb_file = cache_dir / f"{safe}_500.jpg"
    # 缓存有效：源文件未修改
    if thumb_file.exists() and thumb_file.stat().st_mtime >= src.stat().st_mtime:
        return thumb_file
    try:
        img = Image.open(src).convert("RGBA")
        # 白底合成（透明背景看不清楚）
        bg = Image.new("RGBA", img.size, (255, 255, 255, 255))
        bg.alpha_composite(img)
        rgb = bg.convert("RGB")
        # 缩放到最长边 500px
        w, h = rgb.size
        if w > h:
            new_w = 500
            new_h = int(h * 500 / w)
        else:
            new_h = 500
            new_w = int(w * 500 / h)
        rgb = rgb.resize((new_w, new_h), Image.LANCZOS)
        rgb.save(str(thumb_file), "JPEG", quality=90)
        return thumb_file
    except Exception as e:
        print(f"  [HoverThumbError] {filename}: {e}")
        return None


# ---------------------------------------------------------------------------
# 本地回收站：删除到 01_INBOX/回收站/，清空时才送入系统回收站
# ---------------------------------------------------------------------------

TRASH_DIR = INBOX_DIR / "回收站"

FO_DELETE = 3
FOF_ALLOWUNDO = 0x40
FOF_NOCONFIRMATION = 0x10

class SHFILEOPSTRUCTW(ctypes.Structure):
    _fields_ = [
        ("hwnd", wintypes.HWND),
        ("wFunc", wintypes.UINT),
        ("pFrom", wintypes.LPCWSTR),
        ("pTo", wintypes.LPCWSTR),
        ("fFlags", ctypes.c_int),
        ("fAnyOperationsAborted", wintypes.BOOL),
        ("hNameMappings", wintypes.LPVOID),
        ("lpszProgressTitle", wintypes.LPCWSTR),
    ]


def move_to_trash(filename: str, inbox_dir: Path = None, trash_dir: Path = None) -> bool:
    """将 INBOX 中的文件移到本地 回收站 目录；目录缺省为 wb 常量（行为不变）"""
    inbox_dir = inbox_dir or INBOX_DIR
    trash_dir = trash_dir or TRASH_DIR
    safe = os.path.basename(filename)
    src = inbox_dir / safe
    if not src.exists():
        return False
    trash_dir.mkdir(parents=True, exist_ok=True)
    dst = trash_dir / safe
    # 防重名
    if dst.exists():
        stem, ext = os.path.splitext(safe)
        dst = trash_dir / f"{stem}_{int(time.time())}{ext}"
    shutil.move(str(src), str(dst))
    # 清理 hover 缓存
    for f in (inbox_dir / "_hover_cache").glob(f"{safe}*"):
        try: f.unlink()
        except: pass
    return True


def empty_trash_to_system_recycle() -> int:
    """将本地回收站里的所有文件送入系统回收站。返回处理数量。"""
    if not TRASH_DIR.exists():
        return 0
    count = 0
    for f in list(TRASH_DIR.iterdir()):
        if not f.is_file():
            continue
        try:
            # 使用 Windows Shell API 送系统回收站
            fileop = SHFILEOPSTRUCTW()
            fileop.hwnd = 0
            fileop.wFunc = FO_DELETE
            fileop.pFrom = str(f) + "\0"
            fileop.pTo = None
            fileop.fFlags = FOF_ALLOWUNDO | FOF_NOCONFIRMATION
            if ctypes.windll.shell32.SHFileOperationW(ctypes.byref(fileop)) == 0:
                count += 1
            else:
                # 回退：直接删除
                f.unlink()
                count += 1
        except Exception:
            try: f.unlink()
            except: pass
            count += 1
    return count


def send_to_recycle_bin(path: str) -> bool:
    """将指定文件直接送入系统回收站（可手动还原）"""
    try:
        fileop = SHFILEOPSTRUCTW()
        fileop.hwnd = 0
        fileop.wFunc = FO_DELETE
        fileop.pFrom = str(path) + "\0"
        fileop.pTo = None
        fileop.fFlags = FOF_ALLOWUNDO | FOF_NOCONFIRMATION
        return ctypes.windll.shell32.SHFileOperationW(ctypes.byref(fileop)) == 0
    except Exception:
        return False


# ---------------------------------------------------------------------------
# AI 图回收站（用于 ai-review 页面临时删除/还原）
# ---------------------------------------------------------------------------

def _ai_trash_meta_path(dx: str) -> Path:
    """返回某 DX 的回收站元数据文件路径"""
    return AI_TRASH_DIR / dx / ".trash_meta.json"


def move_ai_to_trash(dx: str, filename: str) -> tuple:
    """将 AI 图从 01_AI 移到回收站。返回 (ok, msg)"""
    if not re.match(r"^DX\d+(?:BW|B|W)?$", dx):
        return False, "无效的 DX 编号"
    safe = os.path.basename(filename)
    src = PROJECTS_DIR / dx / "01_AI" / safe
    if not src.exists():
        return False, f"文件不存在: {dx}/01_AI/{safe}"

    AI_TRASH_DIR.mkdir(parents=True, exist_ok=True)
    trash_dx = AI_TRASH_DIR / dx
    trash_dx.mkdir(parents=True, exist_ok=True)
    dst = trash_dx / safe

    # 防重名
    if dst.exists():
        stem, ext = os.path.splitext(safe)
        dst = trash_dx / f"{stem}_{int(time.time())}{ext}"

    try:
        shutil.move(str(src), str(dst))
    except Exception as e:
        return False, f"移动失败: {e}"

    # 记录元数据
    meta = {}
    meta_path = _ai_trash_meta_path(dx)
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    meta[dst.name] = {
        "original_path": f"02_PROJECTS/{dx}/01_AI/{safe}",
        "deleted_at": datetime.now().isoformat(),
        "dx": dx,
    }
    try:
        meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass

    # 清理 AI 对比缩略图缓存（使用与 _get_ai_thumb 一致的 safe_name）
    thumb_safe = re.sub(r'[\\/*?:"<>|]', '_', safe)
    for tf in AI_THUMB_DIR.glob(f"{dx}__{thumb_safe}.*"):
        try:
            tf.unlink()
        except Exception:
            pass

    return True, f"{safe} 已移入 AI 回收站"


def restore_ai_from_trash(dx: str, filename: str) -> tuple:
    """从回收站还原 AI 图到 01_AI。返回 (ok, msg)"""
    if not re.match(r"^DX\d+(?:BW|B|W)?$", dx):
        return False, "无效的 DX 编号"
    safe = os.path.basename(filename)
    src = AI_TRASH_DIR / dx / safe
    if not src.exists():
        return False, f"回收站中不存在: {safe}"

    ai_dir = PROJECTS_DIR / dx / "01_AI"
    ai_dir.mkdir(parents=True, exist_ok=True)
    dst = ai_dir / safe

    # 防重名
    if dst.exists():
        stem, ext = os.path.splitext(safe)
        dst = ai_dir / f"{stem}_restored{ext}"

    try:
        shutil.move(str(src), str(dst))
    except Exception as e:
        return False, f"还原失败: {e}"

    # 清理元数据
    meta_path = _ai_trash_meta_path(dx)
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            meta.pop(safe, None)
            meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    return True, f"{safe} 已还原到 {dx}/01_AI"


def list_ai_trash() -> list:
    """列出 AI 回收站中的所有文件"""
    items = []
    if not AI_TRASH_DIR.exists():
        return items
    for dx_dir in sorted(AI_TRASH_DIR.iterdir()):
        if not dx_dir.is_dir() or not re.match(r"^DX\d+(?:BW|B|W)?$", dx_dir.name):
            continue
        dx = dx_dir.name
        meta = {}
        meta_path = _ai_trash_meta_path(dx)
        if meta_path.exists():
            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
            except Exception:
                meta = {}
        for f in sorted(dx_dir.iterdir()):
            if not f.is_file() or f.name.startswith("."):
                continue
            info = meta.get(f.name, {})
            items.append({
                "dx": dx,
                "filename": f.name,
                "deleted_at": info.get("deleted_at", ""),
                "preview_url": f"/api/ai-review/trash-thumb?dx={dx}&file={f.name}",
            })
    return items


# ---------------------------------------------------------------------------
# INBOX 重命名：B/W → BW（如 2B.png → 2BW.png）
# ---------------------------------------------------------------------------

def rename_to_bw(filename: str, inbox_dir: Path = None) -> tuple:
    """将 B 或 W 文件改名为 BW。inbox_dir 缺省为 wb 常量（行为不变）
    返回 (ok, new_name, msg)
    """
    inbox_dir = inbox_dir or INBOX_DIR
    safe = os.path.basename(filename)
    m = re.match(r'^(\d+)([BW])(\.png)$', safe, re.IGNORECASE)
    if not m:
        return False, "", f"{safe} 不符合格式（需为 数字+B/W+.png）"
    num = m.group(1)
    suffix = m.group(2).upper()
    if suffix not in ("B", "W"):
        return False, "", f"{safe} 已是 BW 或 WB 格式"
    new_name = f"{num}BW.png"
    src = inbox_dir / safe
    dst = inbox_dir / new_name
    if dst.exists():
        return False, "", f"{new_name} 已存在"
    src.rename(dst)
    return True, new_name, f"{safe} → {new_name}"


# ---------------------------------------------------------------------------
# 上款（Upload）：扫描 03_UPLOAD 并提供缩略图
# ---------------------------------------------------------------------------

UPLOAD_THUMB_DIR = BASE_DIR / "_upload_thumbs"
UPLOAD_PROGRESS_FILE = BASE_DIR / ".wb_upload_progress.json"
UPLOAD_RECORD_MD = BASE_DIR / "已上款货号_wb.md"


def _upload_progress_file(cat=None):
    """上款进度文件：按品类落在各自数据根（wb 缺省=D:\\Semems WB，与改造前一致）。

    卫衣等品类上款时进度写入各自根下的 .wb_upload_progress.json，互不串扰。
    """
    return _cat_root(cat or _DEFAULT_CAT) / ".wb_upload_progress.json"


def _cat_suffix(cat=None):
    """记录文件名后缀：wb→'wb'（与历史 已上款货号_wb.md 一致），其余品类用品类名（hoodie→hoodie）。"""
    cat = cat or _DEFAULT_CAT
    return "wb" if cat == _DEFAULT_CAT else cat


def _completed_md_for(cat=None):
    """已上款记录文件：按品类落各自根（wb 缺省=D:\\Semems WB\\已上款货号_wb.md，与改造前一致）。"""
    return _cat_root(cat or _DEFAULT_CAT) / f"已上款货号_{_cat_suffix(cat)}.md"


def _online_listed_file(cat=None):
    """店小秘在线已上款数据文件：按品类落各自根（wb 缺省=D:\\Semems WB\\.wb_online_listed.json，与改造前一致）。"""
    return _cat_root(cat or _DEFAULT_CAT) / ".wb_online_listed.json"


def _title_cache_for(cat=None):
    """标题缓存文件：按品类落各自根（wb 缺省=D:\\Semems WB\\标题缓存_wb.md，与改造前一致）。"""
    return _cat_root(cat or _DEFAULT_CAT) / f"标题缓存_{_cat_suffix(cat)}.md"


def _dx_dir_date(d: Path) -> str:
    """返回 DX 文件夹建立日期（YYMMDD），所有日期分类统一用建立时间。"""
    try:
        return time.strftime("%y%m%d", time.localtime(d.stat().st_ctime))
    except Exception:
        return ""


def _scan_upload_projects(projects_dir: Path = None, prefix: str = "DX", thumb_dir: Path = None):
    """扫描所有 DX 的 03_UPLOAD，返回 [{dx, date, files:[{name,mtime}]}]
    date 统一按 DX 文件夹建立日期分类，不再按文件最后更新时间。
    projects_dir/prefix 缺省为 wb 常量（行为不变）。
    """
    projects_dir = projects_dir or PROJECTS_DIR
    dx_re = _dx_re(prefix)
    projects = []
    if not projects_dir.exists():
        return projects
    for d in sorted(projects_dir.iterdir()):
        if not d.is_dir() or not dx_re.match(d.name):
            continue
        up_dir = d / "03_UPLOAD"
        if not up_dir.is_dir():
            continue
        dx = d.name
        files = []
        for f in sorted(up_dir.iterdir()):
            if not f.is_file():
                continue
            ext = f.suffix.lower()
            if ext not in ('.png', '.jpg', '.jpeg', '.webp'):
                continue
            src_mtime = int(f.stat().st_mtime)
            thumb = _upload_thumb_path(dx, f.name, thumb_dir=thumb_dir)
            thumb_mtime = int(thumb.stat().st_mtime) if thumb.exists() else src_mtime
            files.append({"name": f.name, "mtime": src_mtime, "thumb_mtime": thumb_mtime})
        if not files:
            continue
        dx_date = _dx_dir_date(d)
        projects.append({"dx": dx, "date": dx_date, "files": files})
    return projects


def _scan_ai_review_projects(projects_dir: Path = None, prefix: str = "DX"):
    """扫描所有 DX 项目的 01_AI 目录，直接在同一目录内配对原图与 AI 生成图。
    projects_dir/prefix 缺省为 wb 常量（行为不变）。

    约定：每个 DX/01_AI 中同时存放原图（如 1BW.png）和生成图（如 DX0283_BW.png）。
    配对方式：
      1. 优先读取 source_map.json，并用 Lovart registry 根据 src_id 找到 original_name。
      2. 回退到 uid_map / sidecar 元数据。
      3. 再回退按 role 后缀从 01_AI 中找同 role 的原图。

    返回结构：
    [
      {
        "dx": "DX0287",
        "date": "260703",
        "groups": [
          {
            "group_id": "G_00123",
            "design_number": 1,
            "role": "BW",
            "source_file": "1BW.png",
            "ai_file": "DX0287_BW.png",
            "ai_exists": True,
            "paired": True
          }
        ]
      }
    ]
    """
    projects_dir = projects_dir or PROJECTS_DIR
    dx_re = _dx_re(prefix)
    projects = []
    if not projects_dir.exists():
        return projects

    INBOX_NAME_RE = re.compile(r'^(\d+)(B|W|BW|WB)\.(png|jpg|jpeg|webp)$', re.IGNORECASE)

    for d in sorted(projects_dir.iterdir()):
        if not d.is_dir() or not dx_re.match(d.name):
            continue
        dx = d.name
        ai_dir = d / "01_AI"
        if not ai_dir.is_dir():
            continue

        # 读取目录内所有图片，同时记录 mtime 用于前端缓存刷新
        all_files = []
        file_mtimes = {}
        for f in sorted(ai_dir.iterdir()):
            if not f.is_file():
                continue
            if not f.name.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')):
                continue
            if '_副本' in f.name or '已归档' in f.name or '原图' in f.name:
                continue
            all_files.append(f.name)
            try:
                file_mtimes[f.name] = int(f.stat().st_mtime)
            except Exception:
                file_mtimes[f.name] = 0

        # 分离原图与 AI 生成图
        source_files = [f for f in all_files if INBOX_NAME_RE.match(f)]
        ai_files = [f for f in all_files if not INBOX_NAME_RE.match(f) and f.startswith(f"{dx}_")]

        if not ai_files:
            continue

        # 加载 source_map + Lovart registry，建立 ai_file -> source_file
        sm_map = {}
        sm_path = d / "source_map.json"
        if sm_path.exists():
            try:
                lovart_reg_path = WB_REGISTRY_FILE
                lovart_reg = {}
                if lovart_reg_path.exists():
                    try:
                        with open(lovart_reg_path, 'r', encoding='utf-8') as lf:
                            lovart_reg = json.load(lf)
                    except Exception:
                        lovart_reg = {}

                with open(sm_path, 'r', encoding='utf-8') as f:
                    sm = json.load(f)
                for src in sm.get("sources", []):
                    ai_file = src.get("file", "")
                    orig = src.get("original_name", "")
                    src_id = src.get("src_id", "")
                    if not orig and src_id and src_id in lovart_reg:
                        orig = lovart_reg[src_id].get("original_name", "")
                    if ai_file:
                        sm_map[ai_file] = orig
            except Exception:
                pass

        groups = []
        paired_ai = set()

        # 1. source_map / Lovart registry 精确配对
        for ai_file in ai_files:
            source_name = sm_map.get(ai_file, "")
            role = _role_from_ai_name(ai_file, dx)
            if source_name and source_name in source_files:
                paired_ai.add(ai_file)
                groups.append({
                    "group_id": "",
                    "design_number": _design_number_from_inbox(source_name),
                    "role": role,
                    "source_file": source_name,
                    "source_mtime": file_mtimes.get(source_name, 0),
                    "ai_file": ai_file,
                    "ai_mtime": file_mtimes.get(ai_file, 0),
                    "ai_exists": True,
                    "paired": True,
                })

        # 2. uid_map / sidecar 配对（覆盖或补充）
        if wb_meta is not None:
            try:
                uid_map_data = wb_meta.read_uid_map(d)
                images = uid_map_data.get("images", {})
                source_entries = {}
                ai_entries = {}
                for uid, info in images.items():
                    stage = info.get("stage", "")
                    file_path = Path(info.get("file", "")).name
                    role = info.get("role", "")
                    group_id = info.get("group_id", "")
                    if stage in ("inbox",) and file_path and INBOX_NAME_RE.match(file_path):
                        source_entries[uid] = {"role": role, "group_id": group_id, "file": file_path}
                    elif stage in ("ai", "ai_gen") and file_path:
                        ai_entries[uid] = {"role": role, "group_id": group_id, "file": file_path}

                for uid, src_info in source_entries.items():
                    ai_info = ai_entries.get(uid)
                    if not ai_info:
                        continue
                    source_name = src_info["file"]
                    ai_file = ai_info["file"]
                    if ai_file not in ai_files or source_name not in source_files:
                        continue
                    if ai_file in paired_ai:
                        # 更新已有条目
                        for g in groups:
                            if g["ai_file"] == ai_file:
                                g["source_file"] = source_name
                                g["group_id"] = src_info.get("group_id", "")
                                g["paired"] = True
                                break
                    else:
                        paired_ai.add(ai_file)
                        groups.append({
                            "group_id": src_info.get("group_id", ""),
                            "design_number": _design_number_from_inbox(source_name),
                            "role": src_info.get("role", ""),
                            "source_file": source_name,
                            "source_mtime": file_mtimes.get(source_name, 0),
                            "ai_file": ai_file,
                            "ai_mtime": file_mtimes.get(ai_file, 0),
                            "ai_exists": True,
                            "paired": True,
                        })
            except Exception as e:
                print(f"[AIReview] {dx} uid_map 配对失败: {e}")

        # 3. 把仍未配对的 AI 图按 role 后缀找同 role 原图（最后的回退）
        for ai_file in ai_files:
            if ai_file in paired_ai:
                continue
            role = _role_from_ai_name(ai_file, dx)
            candidates = [f for f in source_files if _role_from_inbox(f) == role]
            source_name = candidates[0] if candidates else ""
            if source_name:
                paired_ai.add(ai_file)
            groups.append({
                "group_id": "",
                "design_number": _design_number_from_inbox(source_name),
                "role": role,
                "source_file": source_name,
                "source_mtime": file_mtimes.get(source_name, 0),
                "ai_file": ai_file,
                "ai_mtime": file_mtimes.get(ai_file, 0),
                "ai_exists": True,
                "paired": source_name != "",
            })

        if not groups:
            continue

        # 合并同一 source_file + role 的多个 AI 变体
        merged = {}
        for g in groups:
            key = (g.get("source_file", ""), g.get("role", "").upper())
            if key not in merged:
                merged[key] = {
                    "group_id": g.get("group_id", ""),
                    "design_number": g.get("design_number", 0),
                    "role": g.get("role", ""),
                    "source_file": g.get("source_file", ""),
                    "source_mtime": g.get("source_mtime", 0),
                    "ai_files": [],
                    "ai_mtimes": [],
                    "paired": False,
                    "ai_exists": False,
                }
            mg = merged[key]
            if g.get("ai_file"):
                mg["ai_files"].append(g["ai_file"])
                mg["ai_mtimes"].append(g.get("ai_mtime", 0))
            if g.get("paired"):
                mg["paired"] = True
            if g.get("ai_exists"):
                mg["ai_exists"] = True
            if g.get("group_id") and not mg["group_id"]:
                mg["group_id"] = g["group_id"]
        groups = list(merged.values())

        # 排序：已配对在前，再按编号、role
        role_order = {"BW": 0, "B": 1, "W": 2, "WB": 3}
        groups.sort(key=lambda g: (
            0 if g["paired"] else 1,
            g["design_number"],
            role_order.get(g["role"].upper(), 99)
        ))

        # 日期统一取 DX 文件夹建立日期
        dx_date = _dx_dir_date(d)

        projects.append({
            "dx": dx,
            "date": dx_date,
            "groups": groups,
        })

    # 排序：未配对/缺图的排前面，其次按日期降序
    projects.sort(key=lambda p: (
        0 if any(not g["paired"] or not g["ai_exists"] for g in p["groups"]) else 1,
        p["date"]
    ), reverse=True)

    return projects


def _role_from_inbox(filename: str) -> str:
    """从 INBOX 文件名提取 role，如 12BW.png -> BW"""
    if not filename:
        return ""
    m = re.match(r'^(\d+)(B|W|BW|WB)\.(png|jpg|jpeg|webp)$', filename, re.IGNORECASE)
    if m:
        return m.group(2).upper()
    return ""

def _design_number_from_inbox(filename: str) -> int:
    """从 INBOX 文件名提取编号，如 12B.png -> 12"""
    if not filename:
        return 0
    m = re.match(r"^(\d+)(B|W|BW|WB)\.(png|jpg|jpeg|webp)$", filename, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return 0


def _role_from_ai_name(filename: str, dx: str) -> str:
    """从 AI 文件名推断 role，如 DX0287_BW.png -> BW, DX0287_B2.png -> B"""
    if not filename:
        return ""
    stem, _ = os.path.splitext(filename)
    prefix = f"{dx}_"
    if stem.startswith(prefix):
        suffix = stem[len(prefix):]
        m = re.match(r'^([BW]+)\d*$', suffix, re.IGNORECASE)
        if m:
            return m.group(1).upper()
    return ""


def _upload_thumb_path(dx: str, filename: str, thumb_dir: Path = None) -> Path:
    """返回 03_UPLOAD 缩略图缓存文件路径（不检查是否存在、不生成）。thumb_dir 缺省 wb 常量"""
    thumb_dir = thumb_dir or UPLOAD_THUMB_DIR
    safe_name = re.sub(r'[\\/*?:"<>|]', '_', filename)
    return thumb_dir / f"{dx}__{safe_name}.jpg"


def _get_upload_thumb(dx: str, filename: str, projects_dir: Path = None, thumb_dir: Path = None, prefix: str = "DX"):
    """返回 03_UPLOAD 缩略图路径（不存在或源文件已更新则重新生成 220px 高）。
    优先使用已缓存缩略图；透明 PNG 则合成白底。目录/前缀缺省为 wb 常量（行为不变）"""
    projects_dir = projects_dir or PROJECTS_DIR
    thumb_dir = thumb_dir or UPLOAD_THUMB_DIR
    if "/" in filename or "\\" in filename or not _dx_re(prefix).match(dx):
        return None
    src = projects_dir / dx / "03_UPLOAD" / filename
    if not src.exists():
        return None
    thumb_dir.mkdir(parents=True, exist_ok=True)
    thumb_file = _upload_thumb_path(dx, filename, thumb_dir=thumb_dir)
    # 缓存有效：缩略图存在且严格比源文件新（mtime 相等时认为可能已更新，重新生成）
    if thumb_file.exists():
        try:
            if thumb_file.stat().st_mtime > src.stat().st_mtime:
                return thumb_file
        except Exception:
            pass
    try:
        from PIL import Image
        img = Image.open(src)
        # 仅当真正存在透明通道时才合成白底，否则直接转 RGB
        if img.mode == 'RGBA':
            # 检查是否有透明像素，若无则直接转 RGB
            alpha = img.getchannel('A')
            if alpha.getextrema()[0] == 255:
                img = img.convert('RGB')
            else:
                bg = Image.new('RGBA', img.size, (255, 255, 255, 255))
                img = Image.alpha_composite(bg, img).convert('RGB')
        elif img.mode == 'P':
            img = img.convert('RGBA')
            bg = Image.new('RGBA', img.size, (255, 255, 255, 255))
            img = Image.alpha_composite(bg, img).convert('RGB')
        else:
            img = img.convert('RGB')
        w, h = img.size
        target_h = 220
        new_w = max(1, int(w * target_h / h))
        img = img.resize((new_w, target_h), Image.LANCZOS)
        img.save(str(thumb_file), "JPEG", quality=85, optimize=True)
        return thumb_file
    except Exception as e:
        print(f"[UploadThumbError] {dx}/{filename}: {e}")
        return None


def _get_ai_thumb(dx: str, filename: str, source: str = "01_AI",
                  projects_dir: Path = None, ai_trash_dir: Path = None,
                  ai_thumb_dir: Path = None, prefix: str = "DX"):
    """返回 01_AI 或回收站中 AI 图的缩略图路径（不存在则生成 300px 高）。
    source: '01_AI' 或 'trash'；目录/前缀缺省为 wb 常量（行为不变）"""
    projects_dir = projects_dir or PROJECTS_DIR
    ai_trash_dir = ai_trash_dir or AI_TRASH_DIR
    ai_thumb_dir = ai_thumb_dir or AI_THUMB_DIR
    if "/" in filename or "\\" in filename or not _dx_re(prefix).match(dx):
        return None
    if source == "trash":
        src = ai_trash_dir / dx / filename
    else:
        src = projects_dir / dx / "01_AI" / filename
    if not src.exists():
        return None
    ai_thumb_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[\\/*?:"<>|]', '_', filename)
    thumb_file = ai_thumb_dir / f"{dx}__{safe_name}.jpg"
    if thumb_file.exists() and thumb_file.stat().st_mtime > src.stat().st_mtime:
        return thumb_file
    try:
        from PIL import Image
        img = Image.open(src)
        if img.mode == 'RGBA':
            alpha = img.getchannel('A')
            if alpha.getextrema()[0] == 255:
                img = img.convert('RGB')
            else:
                bg = Image.new('RGBA', img.size, (255, 255, 255, 255))
                img = Image.alpha_composite(bg, img).convert('RGB')
        elif img.mode == 'P':
            img = img.convert('RGBA')
            bg = Image.new('RGBA', img.size, (255, 255, 255, 255))
            img = Image.alpha_composite(bg, img).convert('RGB')
        else:
            img = img.convert('RGB')
        w, h = img.size
        target_h = 300
        new_w = max(1, int(w * target_h / h))
        img = img.resize((new_w, target_h), Image.LANCZOS)
        img.save(str(thumb_file), "JPEG", quality=85, optimize=True)
        return thumb_file
    except Exception as e:
        print(f"[AIReviewThumbError] {dx}/{filename}: {e}")
        return None


def _get_ai_original(dx: str, filename: str, projects_dir: Path = None, prefix: str = "DX"):
    """返回 01_AI 中 AI 图的原图路径；目录/前缀缺省为 wb 常量（行为不变）"""
    projects_dir = projects_dir or PROJECTS_DIR
    if "/" in filename or "\\" in filename or not _dx_re(prefix).match(dx):
        return None
    src = projects_dir / dx / "01_AI" / filename
    if not src.exists():
        return None
    return src


# ---------------------------------------------------------------------------
# 修复错放文件：把去背/贴图文件移到正确的 DX 文件夹
# ---------------------------------------------------------------------------

def _find_target_dx(filename: str, current_dx: str) -> str:
    """从文件名中提取目标 DX 编号。如 DX0178_B_副本.png → 0178
    提取失败返回空字符串。
    """
    m = re.search(r'(DX\d+)', filename, re.IGNORECASE)
    if m:
        candidate = m.group(1).upper()
        if candidate != current_dx and (PROJECTS_DIR / candidate).exists():
            return candidate
    return ""


@app.route('/api/fix-mismatch', methods=['POST'])
def api_fix_mismatch():
    """修复指定 DX 中的错放文件：
    - 文件名含正确 DX 编号 → 直接移过去
    - 文件名不含 → 无法自动修复，跳过
    - 在目标文件夹写入修复记录
    """
    data = request.get_json(silent=True) or {}
    dx_id = data.get("dx_id", "")
    if not dx_id.startswith("DX") or not (PROJECTS_DIR / dx_id).exists():
        return jsonify({"ok": False, "error": "无效的 DX 编号"}), 400

    rem_dir = PROJECTS_DIR / dx_id / "02_REM_BG"
    report = {"moved": [], "skipped": [], "errors": []}
    log_entries = []  # 写入修复记录

    if not rem_dir.exists():
        return jsonify({"ok": True, "report": report, "msg": "没有 02_REM_BG 目录"})

    for f in list(rem_dir.iterdir()):
        if not f.is_file():
            continue
        if f.name.startswith(f"{dx_id}_"):
            continue
        target_dx = _find_target_dx(f.name, dx_id)
        if not target_dx:
            report["skipped"].append({"file": f.name, "reason": "无法识别目标 DX"})
            continue
        target_rem = PROJECTS_DIR / target_dx / "02_REM_BG"
        if not target_rem.exists():
            target_rem.mkdir(parents=True, exist_ok=True)
        dst = target_rem / f.name
        if dst.exists():
            stem, ext = os.path.splitext(f.name)
            dst = target_rem / f"{stem}_{int(time.time())}{ext}"
        try:
            shutil.move(str(f), str(dst))
            entry = {
                "file": dst.name,
                "from": f"{dx_id}/02_REM_BG",
                "to": f"{target_dx}/02_REM_BG",
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            report["moved"].append(entry)
            log_entries.append(entry)
        except Exception as e:
            report["errors"].append({"file": f.name, "error": str(e)})

    # 写入修复记录到目标文件夹
    if log_entries:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        for entry in log_entries:
            target_dx = entry["to"].split("/")[0]
            log_path = PROJECTS_DIR / target_dx / "_fix_log.json"
            logs = []
            if log_path.exists():
                try:
                    with open(log_path, 'r', encoding='utf-8') as lf:
                        logs = json.load(lf)
                except Exception:
                    logs = []
            logs.append(entry)
            with open(log_path, 'w', encoding='utf-8') as lf:
                json.dump(logs, lf, indent=2, ensure_ascii=False)

    msg_parts = []
    if report["moved"]:
        msg_parts.append(f"已移动 {len(report['moved'])} 个文件")
    if report["skipped"]:
        msg_parts.append(f"跳过 {len(report['skipped'])} 个")
    if report["errors"]:
        msg_parts.append(f"错误 {len(report['errors'])} 个")
    msg = ", ".join(msg_parts) if msg_parts else "无需修复"

    return jsonify({"ok": True, "report": report, "msg": msg})


# ---------------------------------------------------------------------------
# INBOX 分组逻辑
# ---------------------------------------------------------------------------

def group_inbox_files(inbox_dir: Path = None) -> list:
    """
    将 INBOX 中的 .png 按数字编号分组。inbox_dir 缺省为 wb 常量（行为不变）
    例如: 1B.png + 1W.png → group "1"
          13BW.png        → group "13"
    返回: [{"group_number": int, "images": [...], "count": int, "types": [...]}, ...]
    """
    inbox_dir = inbox_dir or INBOX_DIR
    if not inbox_dir.exists():
        return []

    files = [f for f in os.listdir(inbox_dir)
             if f.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')) and not f.startswith('_')]

    groups = {}
    pattern = re.compile(r'^(\d+)(B|W|BW|WB)\.(png|jpg|jpeg|webp)$', re.IGNORECASE)
    for fname in files:
        m = pattern.match(fname)
        if not m:
            continue
        num = m.group(1)
        suffix = m.group(2).upper()
        fp = inbox_dir / fname
        groups.setdefault(num, []).append({
            "filename": fname,
            "suffix": suffix,
            "size": fp.stat().st_size if fp.exists() else 0,
        })

    sorted_groups = []
    for num in sorted(groups, key=lambda x: int(x)):
        images = groups[num]
        sorted_groups.append({
            "group_number": int(num),
            "images": images,
            "count": len(images),
            "types": [img["suffix"] for img in images],
        })
    return sorted_groups


# ============================================================================
# API 路由
# ============================================================================

@app.route('/')
def index():
    """提供 HTML 控制面板"""
    html_file = Path(__file__).parent / "lovart_control.html"
    if html_file.exists():
        return send_file(str(html_file))
    return "<h1>lovart_control.html not found</h1><p>请确保 lovart_control.html 与 bridge.py 在同一目录</p>", 404


@app.route('/api/inbox')
def api_inbox():
    """返回 INBOX 所有图片及分组信息（?cat= 缺省 wb，行为与改造前一致）"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    inbox_dir = ctx["inbox"]
    if not inbox_dir.exists():
        return jsonify({"images": [], "groups": [], "total": 0})

    all_files = []
    for fname in os.listdir(inbox_dir):
        if not fname.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')) or fname.startswith('_'):
            continue
        fp = inbox_dir / fname
        try:
            st = fp.stat()
            all_files.append({
                "filename": fname,
                "size": st.st_size,
                "preview_url": f"/api/preview/{fname}" + _peiyi_cat_qs(cat),
                "modified": datetime.fromtimestamp(st.st_mtime).isoformat(),
            })
        except OSError:
            continue

    groups = group_inbox_files(inbox_dir)
    payload = {"images": all_files, "groups": groups, "total": len(all_files)}
    if cat != _DEFAULT_CAT:
        payload["cat"] = cat          # wb 响应体与改造前逐字节一致
    return jsonify(payload)


@app.route('/api/preview/<path:filename>')
def api_preview(filename):
    """返回原图（由前端 CSS 控制显示大小）；?cat= 缺省 wb"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    safe_name = os.path.basename(filename)
    filepath = ctx["inbox"] / safe_name
    if not filepath.exists():
        abort(404)
    try:
        ext = safe_name.lower()
        if ext.endswith('.jpg') or ext.endswith('.jpeg'):
            ct = 'image/jpeg'
        elif ext.endswith('.webp'):
            ct = 'image/webp'
        else:
            ct = 'image/png'
        return send_file(str(filepath), mimetype=ct, max_age=3600)
    except Exception:
        abort(404)


@app.route('/api/hover/<path:filename>')
def api_hover(filename):
    """返回 500px 悬停预览图（JPEG 白底）；?cat= 缺省 wb"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    safe_name = os.path.basename(filename)
    thumb = get_hover_thumb(safe_name, inbox_dir=ctx["inbox"], cache_dir=ctx["hover_cache"])
    if not thumb:
        abort(404)
    return send_file(str(thumb), mimetype='image/jpeg', max_age=3600)


@app.route('/api/inbox/group')
def api_inbox_group():
    """仅返回分组信息（前端页面可用来刷新分组）；?cat= 缺省 wb"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    groups = group_inbox_files(ctx["inbox"])
    payload = {"groups": groups, "total_groups": len(groups)}
    if cat != _DEFAULT_CAT:
        payload["cat"] = cat
    return jsonify(payload)


@app.route('/api/generate', methods=['POST'])
def api_generate():
    """启动 Lovart 生图任务（?cat= 缺省 wb；非 wb 品类走品类独立目录/注册表/提示词）"""
    global task_state

    cat, err = _resolve_request_cat()
    if err:
        return err
    gp = _gen_paths(cat)

    with _lock:
        if task_state["status"] == "running":
            return jsonify({"error": "已有生图任务正在运行，请等待完成"}), 409

        data = request.get_json(silent=True) or {}
        selected = data.get("selected", [])

        if not selected:
            return jsonify({"error": "请至少选择一张图片"}), 400

        missing = [f for f in selected if not (gp["inbox"] / f).exists()]
        if missing:
            return jsonify({"error": f"以下文件不存在: {', '.join(missing)}"}), 400

        task_id = f"TASK_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        task_state = {
            "status": "starting",
            "progress": "初始化中...",
            "started_at": datetime.now().isoformat(),
            "completed_at": None,
            "log": [],
            "selected_files": selected,
            "groups_processed": 0,
            "groups_total": 0,
            "task_id": task_id,
        }
        if cat != _DEFAULT_CAT:
            task_state["cat"] = cat      # wb 状态体与改造前逐字节一致
        _save_state()

    # 后台线程执行
    t = threading.Thread(target=_run_generation, args=(selected, task_id), kwargs={"cat": cat}, daemon=True)
    t.start()

    resp = {
        "status": "started",
        "task_id": task_id,
        "message": f"已启动生图任务，处理 {len(selected)} 张图片",
    }
    if cat != _DEFAULT_CAT:
        resp["cat"] = cat              # wb 响应体与改造前逐字节一致
    return jsonify(resp)


@app.route('/api/status')
def api_status():
    """返回当前任务状态（含派生字段，便于前端展示）"""
    data = dict(task_state)

    # 运行时长
    started = data.get("started_at")
    completed = data.get("completed_at")
    now = datetime.now()
    elapsed = 0
    if started:
        try:
            start_dt = datetime.fromisoformat(started)
            end_dt = datetime.fromisoformat(completed) if completed else now
            elapsed = max(0, (end_dt - start_dt).total_seconds())
        except Exception:
            pass
    data["elapsed_seconds"] = int(elapsed)

    # 从日志解析成功/失败数量
    success_count = 0
    fail_count = 0
    current_key = ""
    current_dx = ""
    output_folder = ""
    for line in data.get("log", []):
        m = re.search(r'生成\s*(\d+)\s*张[,，]\s*失败\s*(\d+)\s*张', line)
        if m:
            success_count = int(m.group(1))
            fail_count = int(m.group(2))
        m = re.search(r'Key#(\d+)', line)
        if m:
            current_key = f"Key#{m.group(1)}"
        # 当前 DX：从输出路径或任务进度里找
        m = re.search(r'输出到\s+(DX\d+(?:BW|B|W)?)/01_AI', line)
        if m:
            current_dx = m.group(1)
        # Lovart 输出日志：name -> DXxxx/01_AI/name.png
        m = re.search(r'->\s*(DX\d+(?:BW|B|W)?)/01_AI/', line)
        if m:
            current_dx = m.group(1)

    # 目标 DX：重新生图时 Bridge 已指定，优先使用，避免同名文件猜错
    # reuse_dx 可能是 str（单文件）或 dict（批量）。批量时跳过单 DX 输出文件夹逻辑。
    target_dx = data.get("reuse_dx") or data.get("target_dx") or ""
    if target_dx and isinstance(target_dx, str) and (PROJECTS_DIR / target_dx / "01_AI").exists():
        output_folder = target_dx
        if not current_dx:
            current_dx = target_dx

    # 兜底：根据 selected_files + source_map 找到最终输出 DX。
    # 批量重新生图时跳过此猜测，避免返回无关 DX。
    is_batch = data.get("batch") is True
    if data.get("status") in ("completed", "error") and not output_folder and not is_batch:
        for dx in sorted(os.listdir(PROJECTS_DIR)) if PROJECTS_DIR.exists() else []:
            if not dx.startswith("DX"):
                continue
            ai_dir = PROJECTS_DIR / dx / "01_AI"
            if not ai_dir.exists():
                continue
            for src in data.get("selected_files", []):
                if (ai_dir / src).exists():
                    output_folder = dx
                    break
            if output_folder:
                break

    # 批量任务：若未从日志解析到 current_dx，使用 affected_dx 中第一个
    if is_batch and not current_dx:
        affected = data.get("affected_dx", [])
        if affected:
            current_dx = affected[0]

    # 运行状态细化：completed 但有失败时，前端需要明确感知
    display_status = data.get("status", "idle")
    if display_status == "completed":
        if fail_count > 0 and success_count == 0:
            display_status = "error"
        elif fail_count > 0 and success_count > 0:
            display_status = "partial"

    data["success_count"] = success_count
    data["fail_count"] = fail_count
    data["current_key"] = current_key
    data["current_dx"] = current_dx or output_folder
    data["output_folder"] = output_folder
    data["display_status"] = display_status

    # 最新日志（最多 20 条）
    data["latest_log"] = data.get("log", [])[-20:]

    resp = jsonify(data)
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route('/api/projects')
def api_projects():
    """列出最近 DX 项目及关联的 group 信息"""
    reg = load_registry()
    projects = []
    seen = set()
    dx_pattern = re.compile(r'^DX\d+_')

    if PROJECTS_DIR.exists():
        for d in sorted(os.listdir(PROJECTS_DIR), reverse=True):
            if not d.startswith('DX') or d in seen:
                continue
            seen.add(d)

            ai_dir = PROJECTS_DIR / d / "01_AI"
            rem_dir = PROJECTS_DIR / d / "02_REM_BG"
            up_dir  = PROJECTS_DIR / d / "03_UPLOAD"

            if not ai_dir.exists():
                continue

            # 所有文件
            all_ai = sorted([f for f in os.listdir(ai_dir) if f.endswith('.png')])
            # AI 生成的文件（DX{N}_*.png）
            ai_gen = sorted([f for f in all_ai if dx_pattern.match(f)])
            # 去背文件
            rem_files = sorted([f for f in os.listdir(rem_dir) if f.endswith(('.png','.jpg','.jpeg'))]) if rem_dir.exists() else []
            # 贴图文件（多为 jpg）
            up_files = sorted([f for f in os.listdir(up_dir) if f.endswith(('.png','.jpg','.jpeg'))]) if up_dir.exists() else []

            sm_path = PROJECTS_DIR / d / "source_map.json"
            source_map = {}
            ai_src_map = {}  # AI 文件名 → 原图名
            if sm_path.exists():
                try:
                    with open(sm_path, 'r', encoding='utf-8') as f:
                        source_map = json.load(f)
                    # 从 Lovart 注册表查找原图名
                    lovart_reg_path = WB_REGISTRY_FILE
                    if lovart_reg_path.exists():
                        with open(lovart_reg_path, 'r', encoding='utf-8') as lf:
                            lovart_reg = json.load(lf)
                        for src in source_map.get("sources", []):
                            ai_file = src.get("file", "")
                            src_id = src.get("src_id", "")
                            if ai_file and src_id and src_id in lovart_reg:
                                orig = lovart_reg[src_id].get("original_name", "")
                                if orig:
                                    ai_src_map[ai_file] = orig
                except Exception:
                    pass

            # 款号一致性检查：文件夹名 vs AI vs 去背 vs 贴图
            inconsistent = False
            incons_reason = []
            bad_files = []  # 记录不一致的文件名，用于前端高亮
            # 检查 AI 文件
            for f in ai_gen:
                if not f.startswith(f"{d}_"):
                    inconsistent = True
                    bad_files.append(f)
                    incons_reason.append(f"AI文件 {f}")
            # 检查去背文件
            for f in rem_files:
                stem = f.rsplit('.', 1)[0]
                base = stem.replace('_cut', '')
                if not base.startswith(f"{d}_"):
                    inconsistent = True
                    bad_files.append(f)
                    incons_reason.append(f"去背文件 {f}")
            # 检查贴图文件
            for f in up_files:
                if not f.startswith(f"{d}_"):
                    inconsistent = True
                    bad_files.append(f)
                    incons_reason.append(f"贴图文件 {f}")

            projects.append({
                "dx_id": d,
                "file_count": len(all_ai),
                "files": all_ai,
                "ai_gen": ai_gen,
                "rem_files": rem_files,
                "up_files": up_files,
                "has_rembg": len(rem_files) > 0,
                "has_upload": len(up_files) > 0,
                "inconsistent": inconsistent,
                "bad_files": bad_files,
                "incons_reason": "; ".join(incons_reason[:3]) + (f" 等{len(incons_reason)}处" if len(incons_reason) > 3 else ""),
                "source_map": source_map,
                "ai_src_map": ai_src_map,  # AI文件名 → 原图名
                "modified": datetime.fromtimestamp(ai_dir.stat().st_mtime).isoformat(),
            })

    # 关联 group 信息
    group_info = {}
    for gid, ginfo in reg.get("groups", {}).items():
        dx = ginfo.get("dx_folder", "")
        if dx:
            group_info[dx] = {
                "group_id": gid,
                "created": ginfo.get("created", ""),
                "status": ginfo.get("status", ""),
                "images": ginfo.get("images", []),
            }

    # 不一致的排最前面，其余按时间降序
    projects.sort(key=lambda p: (0 if p["inconsistent"] else 1, p["modified"]), reverse=False)
    # 修正倒序：不一致在前，一致的部分内部再按时间倒序
    incons = [p for p in projects if p["inconsistent"]]
    consist = sorted([p for p in projects if not p["inconsistent"]], key=lambda p: p["modified"], reverse=True)
    projects = incons + consist

    return jsonify({
        "projects": projects[:100],   # 最近 100 个
        "group_info": group_info,
        "total": len(projects),
    })


@app.route('/api/open/<path:folder>')
def api_open_folder(folder):
    """在文件管理器中打开指定文件夹（?cat= 缺省 wb，wb 行为不变）"""
    # 支持: DX0001, DX0001/01_AI, DX0001/02_REM_BG, DX0001/03_UPLOAD, INBOX
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    parts = folder.replace('\\', '/').split('/')
    first = parts[0]
    if first.startswith(ctx["prefix"]):
        target = ctx["projects"] / first
        if len(parts) > 1:
            sub = parts[1]
            if sub in ('01_AI', '02_REM_BG', '03_UPLOAD'):
                target = target / sub
    elif first == 'INBOX':
        target = ctx["inbox"]
        if len(parts) > 1:
            target = target / parts[1]
    else:
        abort(404)
    if not target.exists():
        abort(404)
    try:
        os.startfile(str(target))
        return jsonify({"ok": True, "path": str(target)})
    except Exception:
        return jsonify({"error": "打开失败"}), 500


@app.route('/api/delete', methods=['POST'])
def api_delete():
    """将指定文件移到本地 回收站 目录（?cat= 缺省 wb；非 wb 品类移入其自身 INBOX/回收站，不跨品类）"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    data = request.get_json(silent=True) or {}
    filename = data.get("file", "")
    safe = os.path.basename(filename)
    filepath = ctx["inbox"] / safe
    if not filepath.exists():
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    ok = move_to_trash(safe, inbox_dir=ctx["inbox"], trash_dir=ctx["inbox"] / "回收站")
    if ok:
        return jsonify({"ok": True, "msg": f"{safe} 已移到本地回收站"})
    else:
        return jsonify({"ok": False, "error": "删除失败"}), 500


@app.route('/api/empty-trash', methods=['POST'])
def api_empty_trash():
    """将本地回收站里的文件全部送入系统回收站"""
    count = empty_trash_to_system_recycle()
    return jsonify({"ok": True, "count": count, "msg": f"已清空 {count} 个文件到系统回收站"})


@app.route('/api/trash')
def api_trash():
    """列出本地回收站中的文件，按编号分组（同 INBOX 风格）"""
    files = []
    if TRASH_DIR.exists():
        for f in sorted(TRASH_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if f.is_file() and f.suffix.lower() == '.png':
                files.append({
                    "filename": f.name,
                    "size": f.stat().st_size,
                    "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                    "preview_url": f"/api/preview-trash/{f.name}",
                })
    # 分组（同 INBOX 逻辑）
    groups = {}
    pattern = re.compile(r'^(\d+)(B|W|BW|WB)(\.png)$', re.IGNORECASE)
    others = []
    for img in files:
        m = pattern.match(img["filename"])
        if m:
            num = m.group(1)
            suffix = m.group(2).upper()
            groups.setdefault(num, []).append({
                "filename": img["filename"],
                "suffix": suffix,
                "size": img["size"],
                "preview_url": img["preview_url"],
            })
        else:
            others.append(img)
    sorted_groups = []
    for num in sorted(groups, key=lambda x: int(x)):
        imgs = groups[num]
        sorted_groups.append({
            "group_number": int(num),
            "images": imgs,
            "count": len(imgs),
            "types": list(set(img["suffix"] for img in imgs)),
        })
    return jsonify({
        "files": files,
        "groups": sorted_groups,
        "others": others,
        "count": len(files),
    })


@app.route('/api/preview-trash/<path:filename>')
def api_preview_trash(filename):
    """返回回收站中的文件预览"""
    safe = os.path.basename(filename)
    filepath = TRASH_DIR / safe
    if not filepath.exists():
        abort(404)
    try:
        return send_file(str(filepath), mimetype='image/png', max_age=3600)
    except Exception:
        abort(404)


@app.route('/api/restore', methods=['POST'])
def api_restore():
    """从本地回收站恢复文件到 INBOX"""
    data = request.get_json(silent=True) or {}
    filename = data.get("file", "")
    safe = os.path.basename(filename)
    src = TRASH_DIR / safe
    if not src.exists():
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    dst = INBOX_DIR / safe
    if dst.exists():
        # 重名处理：加时间戳
        stem, ext = os.path.splitext(safe)
        dst = INBOX_DIR / f"{stem}_restored{ext}"
    shutil.move(str(src), str(dst))
    return jsonify({"ok": True, "msg": f"{safe} 已恢复到 INBOX"})


@app.route('/api/open/recycle')
def api_open_recycle():
    """打开本地回收站目录（前台显示）"""
    try:
        TRASH_DIR.mkdir(parents=True, exist_ok=True)
        _open_folder_front(TRASH_DIR)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/rename', methods=['POST'])
def api_rename():
    """将 B/W 图片改名为 BW（如 2B.png → 2BW.png）；?cat= 缺省 wb"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    data = request.get_json(silent=True) or {}
    filename = data.get("file", "")
    ok, new_name, msg = rename_to_bw(filename, inbox_dir=ctx["inbox"])
    if ok:
        return jsonify({"ok": True, "new_name": new_name, "msg": msg})
    else:
        return jsonify({"ok": False, "error": msg}), 400


# ============================================================================
# AI 生图对比页面 API
# ============================================================================

@app.route('/ai-review')
def ai_review_page():
    """AI 生图对比页面"""
    html_file = Path(__file__).parent / "ai_review.html"
    if html_file.exists():
        return send_file(str(html_file))
    return "<h1>ai_review.html not found</h1><p>请确保 ai_review.html 与 bridge.py 在同一目录</p>", 404


@app.route('/api/ai-review/projects')
def api_ai_review_projects():
    """返回所有 DX 的 INBOX 原图与 01_AI 生成图配对列表（?cat= 缺省 wb）"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    try:
        projects = _scan_ai_review_projects(projects_dir=ctx["projects"], prefix=ctx["prefix"])
        dates = sorted({p["date"] for p in projects if p["date"]}, reverse=True)
        payload = {"ok": True, "projects": projects, "dates": dates, "total": len(projects)}
        if cat != _DEFAULT_CAT:
            payload["cat"] = cat
        return jsonify(payload)
    except Exception as e:
        import traceback
        print(f"[AIReview] /api/ai-review/projects 错误: {e}\n{traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/ai-review/thumb')
def api_ai_review_thumb():
    """返回 01_AI 中 AI 图的缩略图（?cat= 缺省 wb）"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    dx = request.args.get("dx", "").strip()
    filename = request.args.get("file", "").strip()
    if not _dx_re(ctx["prefix"]).match(dx) or not filename:
        return "bad params", 400
    thumb = _get_ai_thumb(dx, filename, source="01_AI",
                          projects_dir=ctx["projects"], ai_trash_dir=ctx["ai_trash"],
                          ai_thumb_dir=ctx["ai_thumb"], prefix=ctx["prefix"])
    if not thumb:
        return "no thumb", 404
    r = make_response(send_file(str(thumb), mimetype="image/jpeg"))
    r.headers["Cache-Control"] = "public, max-age=3600"
    return r


@app.route('/api/ai-review/original')
def api_ai_review_original():
    """返回 01_AI 中 AI 图的原图（供悬停放大）；?cat= 缺省 wb"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    dx = request.args.get("dx", "").strip()
    filename = request.args.get("file", "").strip()
    if not _dx_re(ctx["prefix"]).match(dx) or not filename:
        return "bad params", 400
    src = _get_ai_original(dx, filename, projects_dir=ctx["projects"], prefix=ctx["prefix"])
    if not src:
        return "not found", 404
    ct = "image/png" if filename.lower().endswith(".png") else "image/jpeg"
    r = make_response(send_file(str(src), mimetype=ct))
    r.headers["Cache-Control"] = "public, max-age=3600"
    return r


@app.route('/api/ai-review/trash-thumb')
def api_ai_review_trash_thumb():
    """返回回收站中 AI 图的缩略图（?cat= 缺省 wb）"""
    cat, err = _resolve_request_cat()
    if err:
        return err
    ctx = _cat_ctx(cat)
    dx = request.args.get("dx", "").strip()
    filename = request.args.get("file", "").strip()
    if not _dx_re(ctx["prefix"]).match(dx) or not filename:
        return "bad params", 400
    thumb = _get_ai_thumb(dx, filename, source="trash",
                          projects_dir=ctx["projects"], ai_trash_dir=ctx["ai_trash"],
                          ai_thumb_dir=ctx["ai_thumb"], prefix=ctx["prefix"])
    if not thumb:
        return "no thumb", 404
    r = make_response(send_file(str(thumb), mimetype="image/jpeg"))
    r.headers["Cache-Control"] = "public, max-age=3600"
    return r


def _stage_source_for_regen(source_path: Path) -> tuple:
    """把 DX/01_AI 中的原图临时复制到 INBOX，处理同名冲突。

    返回: (inbox_path, inbox_conflict_path, error_message)
    - inbox_path: INBOX 中的目标路径
    - inbox_conflict_path: 被移走的冲突文件路径（无冲突时为 None）
    - error_message: 失败时的错误信息（成功时为 None）
    """
    source_file = source_path.name
    inbox_path = INBOX_DIR / source_file
    inbox_conflict_path = None
    try:
        source_md5 = compute_md5(str(source_path))
        if inbox_path.exists():
            inbox_md5 = compute_md5(str(inbox_path))
            if inbox_md5 and inbox_md5 != source_md5:
                conflict_dir = AI_TRASH_DIR / "_inbox_conflicts"
                conflict_dir.mkdir(parents=True, exist_ok=True)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                conflict_name = f"{Path(source_file).stem}_{ts}{Path(source_file).suffix}"
                inbox_conflict_path = conflict_dir / conflict_name
                shutil.move(str(inbox_path), str(inbox_conflict_path))
                log(f"INBOX 同名冲突已移走: {source_file} -> {inbox_conflict_path.name}")
        shutil.copy2(str(source_path), str(inbox_path))
        return inbox_path, inbox_conflict_path, None
    except Exception as e:
        return inbox_path, inbox_conflict_path, f"复制原图到 INBOX 失败: {e}"


def _restore_inbox_after_regen(inbox_path: Path, inbox_conflict_path: Path = None):
    """重新生图结束后清理临时原图，冲突文件保留在暂存区。"""
    source_file = inbox_path.name
    try:
        if inbox_path.exists():
            try:
                inbox_path.unlink()
                log(f"INBOX 临时原图已清理: {source_file}")
            except Exception as e:
                log(f"WARN: 清理 INBOX 临时原图失败: {e}")
        if inbox_conflict_path and inbox_conflict_path.exists():
            log(f"INBOX 冲突文件保留在回收站: {inbox_conflict_path.name}")
    except Exception as e:
        log(f"WARN: 恢复/清理 INBOX 失败 {source_file}: {e}")


def _cleanup_duplicate_sources(dx: str, source_file: str):
    """删除 Lovart 归档源图时产生的重复副本（如 17bw(2).png）。"""
    try:
        ai_dir = PROJECTS_DIR / dx / "01_AI"
        original = ai_dir / source_file
        if not original.exists():
            return
        orig_md5 = compute_md5(str(original))
        stem = Path(source_file).stem
        for f in ai_dir.iterdir():
            if not f.is_file():
                continue
            if re.match(rf'^{re.escape(stem)}\(\d+\)\.png$', f.name, re.IGNORECASE):
                try:
                    if compute_md5(str(f)) == orig_md5:
                        f.unlink()
                        log(f"删除重复源图副本: {f.name}")
                except Exception:
                    pass
    except Exception as e:
        log(f"WARN: 清理重复源图副本失败: {e}")


@app.route('/api/ai-review/regenerate', methods=['POST'])
def api_ai_review_regenerate():
    """对指定 01_AI 中的原图重新生图（会重新生成其所在整组）。"""
    global task_state
    cat, cat_err = _resolve_request_cat()
    if cat_err:
        return cat_err
    if cat != _DEFAULT_CAT:
        return _cat_not_ready(cat, "重新生图", "生图管线目前为 T 恤专用，卫衣管线尚未标定。")
    data = request.get_json(silent=True) or {}
    dx = data.get("dx", "").strip()
    source_file = data.get("source_file", "").strip()

    if not dx or not source_file:
        return jsonify({"ok": False, "error": "缺少 dx 或 source_file"}), 400
    if not re.match(r"^DX\d+(?:BW|B|W)?$", dx):
        return jsonify({"ok": False, "error": "无效的 DX 编号"}), 400

    source_path = PROJECTS_DIR / dx / "01_AI" / source_file
    if not source_path.exists():
        return jsonify({"ok": False, "error": f"{dx}/01_AI 中不存在 {source_file}"}), 404

    with _lock:
        if task_state["status"] == "running":
            return jsonify({"ok": False, "error": "已有生图任务正在运行，请等待完成"}), 409

    inbox_path, inbox_conflict_path, err = _stage_source_for_regen(source_path)
    if err:
        return jsonify({"ok": False, "error": err}), 500

    # 找到该文件所在 group
    inbox_groups = group_inbox_files()
    target_group = None
    for g in inbox_groups:
        if any(img["filename"] == source_file for img in g["images"]):
            target_group = g
            break
    if not target_group:
        # 复制错了，清理掉；如有冲突文件则移回
        try:
            if inbox_path.exists():
                inbox_path.unlink()
            if inbox_conflict_path and inbox_conflict_path.exists():
                shutil.move(str(inbox_conflict_path), str(inbox_path))
        except Exception:
            pass
        return jsonify({"ok": False, "error": f"无法确定 {source_file} 的分组"}), 400

    # 清除 Lovart 处理记录里该原图的 hash，强制重新生成
    try:
        removed = _remove_from_lovart_track(inbox_path)
        if removed:
            log(f"已清除 {removed} 条 Lovart 处理记录，强制重新生图: {source_file}")
    except Exception as e:
        log(f"WARN: 清除 Lovart 处理记录失败: {e}")

    # 启动后台生图任务
    task_id = f"TASK_REGEN_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    with _lock:
        task_state = {
            "status": "starting",
            "display_status": "starting",
            "progress": "初始化重新生图...",
            "started_at": datetime.now().isoformat(),
            "completed_at": None,
            "log": [],
            "selected_files": [source_file],
            "groups_processed": 0,
            "groups_total": 1,
            "task_id": task_id,
            "reuse_dx": dx,
        }
        _save_state()

    def _regen_wrapper():
        try:
            _run_generation([source_file], task_id, reuse_dx=dx)
        except Exception as e:
            log(f"重新生图任务异常: {e}")
        finally:
            _restore_inbox_after_regen(inbox_path, inbox_conflict_path)
            _cleanup_duplicate_sources(dx, source_file)

    t = threading.Thread(target=_regen_wrapper, daemon=True)
    t.start()

    return jsonify({
        "ok": True,
        "task_id": task_id,
        "msg": f"已启动重新生图：{dx}/{source_file}（整组 {len(target_group['images'])} 张）",
    })


@app.route('/api/ai-review/regenerate-batch', methods=['POST'])
def api_ai_review_regenerate_batch():
    """批量重新生图：支持勾选多个 01_AI 原图，利用 Lovart 并发生成。

    请求体：{items: [{dx, source_file}]}
    限制：同一批次内所有 source_file 必须全局唯一（不允许跨 DX 同名），
          因为 LOVART_REGEN_DX_MAP 以文件名为 key。
    """
    global task_state
    cat, cat_err = _resolve_request_cat()
    if cat_err:
        return cat_err
    if cat != _DEFAULT_CAT:
        return _cat_not_ready(cat, "批量重新生图", "生图管线目前为 T 恤专用，卫衣管线尚未标定。")
    data = request.get_json(silent=True) or {}
    items = data.get("items", [])

    if not items or not isinstance(items, list):
        return jsonify({"ok": False, "error": "缺少 items"}), 400

    # 校验每个条目
    seen_files = set()
    dup_files = set()
    validated = []
    for item in items:
        dx = str(item.get("dx", "")).strip()
        source_file = str(item.get("source_file", "")).strip()
        if not dx or not source_file:
            continue
        if not re.match(r"^DX\d+(?:BW|B|W)?$", dx):
            return jsonify({"ok": False, "error": f"无效的 DX 编号: {dx}"}), 400
        source_path = PROJECTS_DIR / dx / "01_AI" / source_file
        if not source_path.exists():
            return jsonify({"ok": False, "error": f"{dx}/01_AI 中不存在 {source_file}"}), 404
        if source_file in seen_files:
            dup_files.add(source_file)
        seen_files.add(source_file)
        validated.append({"dx": dx, "source_file": source_file, "source_path": source_path})

    if not validated:
        return jsonify({"ok": False, "error": "没有有效的重新生图项"}), 400
    if dup_files:
        return jsonify({
            "ok": False,
            "error": f"同一批次内不允许同名文件（跨 DX）：{', '.join(sorted(dup_files))}"
        }), 409

    with _lock:
        if task_state["status"] == "running":
            return jsonify({"ok": False, "error": "已有生图任务正在运行，请等待完成"}), 409

    # 准备 INBOX：复制所有源文件，处理同名冲突
    staged = []  # [{dx, source_file, inbox_path, conflict_path}]
    try:
        for v in validated:
            inbox_path, inbox_conflict_path, err = _stage_source_for_regen(v["source_path"])
            if err:
                # 回滚已复制的文件
                for s in staged:
                    _restore_inbox_after_regen(s["inbox_path"], s["conflict_path"])
                return jsonify({"ok": False, "error": err}), 500
            staged.append({
                "dx": v["dx"],
                "source_file": v["source_file"],
                "inbox_path": inbox_path,
                "conflict_path": inbox_conflict_path,
            })
    except Exception as e:
        for s in staged:
            _restore_inbox_after_regen(s["inbox_path"], s["conflict_path"])
        return jsonify({"ok": False, "error": f"准备 INBOX 失败: {e}"}), 500

    # 校验 INBOX 分组（至少每个文件都能被识别到）
    inbox_groups = group_inbox_files()
    inbox_files_set = {s["source_file"] for s in staged}
    matched_files = set()
    for g in inbox_groups:
        for img in g["images"]:
            if img["filename"] in inbox_files_set:
                matched_files.add(img["filename"])
    if len(matched_files) != len(inbox_files_set):
        missing = inbox_files_set - matched_files
        for s in staged:
            _restore_inbox_after_regen(s["inbox_path"], s["conflict_path"])
        return jsonify({"ok": False, "error": f"以下文件无法确定分组: {', '.join(sorted(missing))}"}), 400

    # 清除 Lovart 处理记录
    for s in staged:
        try:
            removed = _remove_from_lovart_track(s["inbox_path"])
            if removed:
                log(f"已清除 {removed} 条 Lovart 处理记录，强制重新生图: {s['source_file']}")
        except Exception as e:
            log(f"WARN: 清除 Lovart 处理记录失败 {s['source_file']}: {e}")

    # 启动后台生图任务
    selected_files = [s["source_file"] for s in staged]
    regen_map = {s["source_file"]: s["dx"] for s in staged}
    affected_dx = sorted({s["dx"] for s in staged})
    task_id = f"TASK_REGEN_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    with _lock:
        task_state = {
            "status": "starting",
            "display_status": "starting",
            "progress": f"初始化批量重新生图 {len(selected_files)} 张...",
            "started_at": datetime.now().isoformat(),
            "completed_at": None,
            "log": [],
            "selected_files": selected_files,
            "groups_processed": 0,
            "groups_total": len(selected_files),
            "task_id": task_id,
            "reuse_dx": regen_map,
            "batch": True,
            "affected_dx": affected_dx,
        }
        _save_state()

    def _batch_regen_wrapper():
        try:
            _run_generation(selected_files, task_id, reuse_dx=regen_map)
        except Exception as e:
            log(f"批量重新生图任务异常: {e}")
        finally:
            for s in staged:
                _restore_inbox_after_regen(s["inbox_path"], s["conflict_path"])
                _cleanup_duplicate_sources(s["dx"], s["source_file"])

    t = threading.Thread(target=_batch_regen_wrapper, daemon=True)
    t.start()

    return jsonify({
        "ok": True,
        "task_id": task_id,
        "msg": f"已启动批量重新生图：{len(selected_files)} 张，涉及 {len(affected_dx)} 个 DX（{', '.join(affected_dx)}）",
    })


@app.route('/api/ai-review/delete-ai', methods=['POST'])
def api_ai_review_delete_ai():
    """将 AI 图移入回收站"""
    cat, cat_err = _resolve_request_cat()
    if cat_err:
        return cat_err
    if cat != _DEFAULT_CAT:
        return _cat_not_ready(cat, "AI 图回收", "AI 回收站目前为 T 恤专用，卫衣尚未标定。")
    data = request.get_json(silent=True) or {}
    dx = data.get("dx", "").strip()
    filename = data.get("file", "").strip()
    if not dx or not filename:
        return jsonify({"ok": False, "error": "缺少 dx 或 file"}), 400
    ok, msg = move_ai_to_trash(dx, filename)
    if ok:
        return jsonify({"ok": True, "msg": msg})
    return jsonify({"ok": False, "error": msg}), 500


@app.route('/api/ai-review/restore-ai', methods=['POST'])
def api_ai_review_restore_ai():
    """从回收站还原 AI 图"""
    cat, cat_err = _resolve_request_cat()
    if cat_err:
        return cat_err
    if cat != _DEFAULT_CAT:
        return _cat_not_ready(cat, "AI 图还原", "AI 回收站目前为 T 恤专用，卫衣尚未标定。")
    data = request.get_json(silent=True) or {}
    dx = data.get("dx", "").strip()
    filename = data.get("file", "").strip()
    if not dx or not filename:
        return jsonify({"ok": False, "error": "缺少 dx 或 file"}), 400
    ok, msg = restore_ai_from_trash(dx, filename)
    if ok:
        return jsonify({"ok": True, "msg": msg})
    return jsonify({"ok": False, "error": msg}), 500


@app.route('/api/ai-review/trash')
def api_ai_review_trash():
    """返回 AI 图回收站列表（仅 wb；非 wb 品类明确报错）"""
    cat, cat_err = _resolve_request_cat()
    if cat_err:
        return cat_err
    if cat != _DEFAULT_CAT:
        return _cat_not_ready(cat, "AI 回收站", "AI 回收站目前为 T 恤专用，卫衣尚未标定。")
    try:
        items = list_ai_trash()
        return jsonify({"ok": True, "items": items, "count": len(items)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/ai-review/empty-trash', methods=['POST'])
def api_ai_review_empty_trash():
    """永久清空 AI 图回收站（仅 wb）"""
    cat, cat_err = _resolve_request_cat()
    if cat_err:
        return cat_err
    if cat != _DEFAULT_CAT:
        return _cat_not_ready(cat, "清空 AI 回收站", "AI 回收站目前为 T 恤专用，卫衣尚未标定。")
    if not AI_TRASH_DIR.exists():
        return jsonify({"ok": True, "count": 0, "msg": "回收站为空"})
    count = 0
    errors = []
    for dx_dir in list(AI_TRASH_DIR.iterdir()):
        if not dx_dir.is_dir():
            continue
        for f in list(dx_dir.iterdir()):
            if not f.is_file():
                continue
            try:
                f.unlink()
                count += 1
            except Exception as e:
                errors.append(f"{dx_dir.name}/{f.name}: {e}")
        # 尝试删除空目录
        try:
            dx_dir.rmdir()
        except Exception:
            pass
    # 尝试删除根目录
    try:
        AI_TRASH_DIR.rmdir()
    except Exception:
        pass
    msg = f"已清空 {count} 个文件"
    if errors:
        msg += f"，{len(errors)} 个失败"
    return jsonify({"ok": True, "count": count, "errors": errors, "msg": msg})


@app.route('/api/registry/query')
def api_registry_query():
    """查询注册表中单张图片的信息（支持文件名 / UID / MD5）"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"error": "请提供查询参数 ?q=filename_or_uid"}), 400

    reg = load_registry()
    results = []

    # 按 name_index 查找
    md5_by_name = reg.get("name_index", {}).get(q)
    if md5_by_name and md5_by_name in reg.get("images", {}):
        results.append(reg["images"][md5_by_name])

    # 按 uid_index 查找
    md5_by_uid = reg.get("uid_index", {}).get(q)
    if md5_by_uid and md5_by_uid in reg.get("images", {}) and reg["images"][md5_by_uid] not in results:
        results.append(reg["images"][md5_by_uid])

    # 按 MD5 直接查找
    if q in reg.get("images", {}):
        if reg["images"][q] not in results:
            results.append(reg["images"][q])

        return jsonify({"query": q, "results": results, "count": len(results)})


@app.route('/api/provenance')
def api_provenance():
    """查询单张图片的血缘链（溯源）"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"error": "请提供 ?q=filename_or_md5"}), 400

    reg = load_registry()
    reg = ensure_registry_v4(reg)

    # 按 MD5、文件名或 UID 查找
    target_md5 = q
    if q in reg.get("uid_index", {}):
        target_md5 = reg["uid_index"][q]
    elif q in reg.get("name_index", {}):
        target_md5 = reg["name_index"].get(q, q)
    elif q not in reg.get("images", {}):
        # 尝试通过 inbox_original_name 匹配
        for md5, entry in reg.get("images", {}).items():
            if entry.get("inbox_original_name") == q or entry.get("original_name") == q:
                target_md5 = md5
                break

    target = reg["images"].get(target_md5)
    if not target:
        return jsonify({"query": q, "error": "未找到"}), 404

    # 构建血缘链：从 root 到当前
    chain = []
    md5 = target_md5
    while md5 and md5 in reg.get("images", {}):
        entry = reg["images"][md5]
        chain.append({
            "md5": md5,
            "name": entry.get("current_name", ""),
            "path": entry.get("current_path", ""),
            "type": entry.get("source_type", "inbox" if not entry.get("source_md5") else entry["source_type"]),
            "uid": entry.get("uid", ""),
            "role": entry.get("role", ""),
            "root_name": entry.get("root_name", ""),
        })
        md5 = entry.get("source_md5", "")
        if not md5:
            break

    # 衍生图片
    derived = []
    for d_md5 in target.get("derived_md5s", []):
        if d_md5 in reg.get("images", {}):
            dentry = reg["images"][d_md5]
            derived.append({
                "md5": d_md5,
                "name": dentry.get("current_name", ""),
                "path": dentry.get("current_path", ""),
                "type": dentry.get("source_type", ""),
                "uid": dentry.get("uid", ""),
            })

    return jsonify({
        "query": q,
        "target": {
            "md5": target_md5,
            "name": target.get("current_name", ""),
            "path": target.get("current_path", ""),
            "uid": target.get("uid", ""),
            "role": target.get("role", ""),
            "inbox_original": target.get("inbox_original_name", ""),
            "source_type": target.get("source_type", ""),
            "source_md5": target.get("source_md5", ""),
            "root_name": target.get("root_name", ""),
            "root_md5": target.get("root_md5", ""),
        },
        "chain": chain,
        "derived": derived,
    })


@app.route('/api/scan-provenance', methods=['POST'])
def api_scan_provenance():
    """扫描所有 DX 文件夹，建立/更新溯源关系"""
    try:
        count = scan_provenance()
        return jsonify({"ok": True, "count": count, "msg": f"已建立 {count} 条溯源关系"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── 统一血缘注册入口（供 check_rem / 贴图等外部工具 Hook 调用） ──

@app.route('/api/lineage/register', methods=['POST'])
def api_lineage_register():
    """外部工具调用此接口记录血缘关系。
    
    Payload:
      child_path: str  - 输出文件的全路径
      parent_path: str - 输入文件的全路径  
      stage: str       - rembg | upload | ai_gen
    """
    data = request.get_json(silent=True) or {}
    child_path = data.get("child_path", "")
    parent_path = data.get("parent_path", "")
    stage = data.get("stage", "")
    uid = data.get("uid", "")
    group_id = data.get("group_id", "")
    role = data.get("role", "")

    if not child_path or not parent_path or not stage:
        return jsonify({"ok": False, "error": "需要 child_path, parent_path, stage"}), 400
    if stage not in ("rembg", "upload", "ai_gen"):
        return jsonify({"ok": False, "error": f"不支持的 stage: {stage}"}), 400

    child_path = Path(child_path)
    parent_path = Path(parent_path)
    if not child_path.exists():
        return jsonify({"ok": False, "error": f"child_path 不存在: {child_path}"}), 400
    if not parent_path.exists():
        return jsonify({"ok": False, "error": f"parent_path 不存在: {parent_path}"}), 400

    try:
        child_md5 = compute_md5(str(child_path))
        parent_md5 = compute_md5(str(parent_path))

        reg = load_registry()
        reg = ensure_registry_v4(reg)

        # 如果 registry 中没有这两个文件，先注册
        for md5_val, fpath in [(child_md5, child_path), (parent_md5, parent_path)]:
            if md5_val not in reg.get("images", {}):
                fname = fpath.name
                # 用相对路径
                try:
                    rel = fpath.relative_to(BASE_DIR)
                    rel_str = str(rel).replace('\\', '/')
                except ValueError:
                    rel_str = str(fpath)
                reg["images"][md5_val] = {
                    "md5": md5_val,
                    "current_name": fname,
                    "current_path": rel_str,
                    "events": [{
                        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "event": "lineage_register",
                        "detail": f"来自 Hook: {stage}",
                    }],
                }
                _add_provenance_fields(reg["images"][md5_val])

        _register_provenance(reg, child_md5, parent_md5, stage, lineage_status="confirmed")
        save_registry(reg)

        # 同步更新 uid_map / sidecar（如果提供了 UID 或能从 sidecar 读取到）
        if wb_meta:
            try:
                parent_meta = wb_meta.read_meta(parent_path)
                effective_uid = uid or (parent_meta.get("uid") if parent_meta else "")
                effective_gid = group_id or (parent_meta.get("group_id") if parent_meta else "")
                effective_role = role or (parent_meta.get("role") if parent_meta else "")
                if effective_uid and effective_gid:
                    dx_dir = child_path.parent.parent
                    if dx_dir.name.startswith("DX"):
                        stage_key = stage  # rembg/upload/ai_gen
                        if stage == "upload":
                            stage_key = "sticker"
                        wb_meta.register_image_in_map(
                            dx_dir, effective_uid, effective_gid, stage_key,
                            effective_role, str(child_path),
                            parent_uid=effective_uid,
                            source_file=str(parent_path),
                        )
                        wb_meta.ensure_meta(
                            child_path, uid=effective_uid, group_id=effective_gid,
                            stage=stage_key, role=effective_role,
                            parent_uid=effective_uid, source_file=str(parent_path),
                        )
            except Exception as e:
                # 不阻断原有血缘注册
                print(f"[lineage/register] uid_map 同步失败: {e}")

        return jsonify({
            "ok": True,
            "msg": f"已记录 {stage} 血缘: {child_path.name} ← {parent_path.name}",
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/launch-check-rem', methods=['POST'])
def api_launch_check_rem():
    """确保去背预览服务（按品类分端口）已就绪，并返回状态与端口。

    实际页面打开由前端直接执行 window.open，这里只负责兜底拉起 check_rem.py。
    check_rem.py 通常由 Bridge 启动时的守护线程保持常驻（各品类独立实例）。
    """
    data = request.get_json(silent=True) or {}
    cat = data.get("cat", _DEFAULT_CAT)
    port = _CHECK_REM_PORT_FOR_CAT.get(cat, 8766)
    # 兜底：若守护线程还没把 check_rem 拉起来，主动拉一次并再等一会儿让服务 bind
    if not _port_ready("127.0.0.1", port, timeout=2):
        _check_rem_ensure(cat, port)
        if not _port_ready("127.0.0.1", port, timeout=5):
            return jsonify({"ok": False, "error": "去背预览服务未就绪，请稍后再试"}), 503
    return jsonify({"ok": True, "msg": "去背预览服务已就绪", "port": port, "cat": cat})


@app.route('/upload')
def upload_page():
    """上款页面：展示 03_UPLOAD 成品并批量上传"""
    html_file = Path(__file__).parent / "upload.html"
    if html_file.exists():
        return send_file(str(html_file))
    return "<h1>upload.html not found</h1>", 404


def _upload_listing_profile_ready(cat: str) -> bool:
    """非 wb 品类的上款前置条件（「标定」就绪）。

    wb（T恤）永远视为已就绪。

    非 wb 品类（如 hoodie）采用「店小秘引用模板款号」机制：上款时引用对应模板款号
    （如卫衣 HX0000），类目/卖点/属性/尺码表从模板继承，无需静态 listing_profile.json。
    就绪判定：标题提示词 prompts/title_<cat>.md 已就位（即「标定」完成）。
    仍兼容旧式 listing_profile.json / categories.json 配置。
    """
    if cat == _DEFAULT_CAT:
        return True
    # 新机制：标题提示词就位即视为就绪（卫衣引用 HX0000 模板继承 4 字段，无需静态 profile）
    prompt = (WB_LISTING_DIR / "prompts" / f"title_{cat}.md")
    if prompt.exists():
        return True
    # 兼容旧式 listing_profile.json / categories.json 配置
    try:
        cfg = _cat_all().get(cat, {})
        if cfg.get("listing_profile"):
            return True
        if (_cat_root(cat) / "listing_profile.json").exists():
            return True
    except Exception:
        pass
    return False


def _upload_cat_guard():
    """上款相关端点的品类守卫。

    返回 (cat, ctx, error_response)。非法 cat → 400；非 wb 且缺 listing profile → 明确报错。
    """
    cat, err = _resolve_request_cat()
    if err:
        return None, None, err
    if cat != _DEFAULT_CAT and not _upload_listing_profile_ready(cat):
        return None, None, _cat_not_ready(cat, "上款",
            "卫衣标题提示词尚未标定（缺 prompts/title_hoodie.md），为避免上错款已拦截。")
    return cat, _cat_ctx(cat), None


@app.route('/api/category-ready')
def api_category_ready():
    """返回当前品类上款是否就绪（用于前端「待标定」角标动态显示）。

    就绪判定与 _upload_listing_profile_ready 一致：wb 恒就绪；
    非 wb 品类（如 hoodie）标题提示词 prompts/title_<cat>.md 存在即就绪。
    """
    cat = request.args.get("cat") or _DEFAULT_CAT
    ready = _upload_listing_profile_ready(cat)
    return jsonify({"ok": True, "ready": ready, "cat": cat})


@app.route('/api/upload/projects')
def api_upload_projects():
    """返回所有含 03_UPLOAD 成品的 DX 列表，并标记是否在线已上款（?cat= 缺省 wb）"""
    cat, ctx, err = _upload_cat_guard()
    if err:
        return err
    projects = _scan_upload_projects(projects_dir=ctx["projects"], prefix=ctx["prefix"],
                                     thumb_dir=ctx["upload_thumb"])
    online_set = _read_online_listed(cat)
    for p in projects:
        p["online_listed"] = p.get("dx", "") in online_set
    payload = {"ok": True, "projects": projects, "online_updated_at": _online_listed_updated_at(cat), "online_mode": _online_listed_mode(cat)}
    if cat != _DEFAULT_CAT:
        payload["cat"] = cat
    return jsonify(payload)


@app.route('/api/upload/thumb')
def api_upload_thumb():
    """返回 03_UPLOAD 缩略图（?cat= 缺省 wb）"""
    cat, ctx, err = _upload_cat_guard()
    if err:
        return err
    dx = request.args.get("dx", "")
    filename = request.args.get("file", "")
    if not _dx_re(ctx["prefix"]).match(dx) or not filename:
        return "bad params", 400
    thumb = _get_upload_thumb(dx, filename, projects_dir=ctx["projects"],
                              thumb_dir=ctx["upload_thumb"], prefix=ctx["prefix"])
    if not thumb:
        return "no thumb", 404
    r = make_response(send_file(str(thumb), mimetype="image/jpeg"))
    r.headers["Cache-Control"] = "public, max-age=3600"
    return r


@app.route('/api/upload/original')
def api_upload_original():
    """返回 03_UPLOAD 原图（供悬停放大）；?cat= 缺省 wb"""
    cat, ctx, err = _upload_cat_guard()
    if err:
        return err
    dx = request.args.get("dx", "")
    filename = request.args.get("file", "")
    if not _dx_re(ctx["prefix"]).match(dx) or not filename:
        return "bad params", 400
    src = ctx["projects"] / dx / "03_UPLOAD" / filename
    if not src.exists():
        return "not found", 404
    ct = "image/png" if filename.lower().endswith(".png") else "image/jpeg"
    r = make_response(send_file(str(src), mimetype=ct))
    r.headers["Cache-Control"] = "public, max-age=3600"
    return r


@app.route('/api/upload/delete', methods=['POST'])
def api_upload_delete():
    """将 03_UPLOAD 中的成品图删除到系统回收站（?cat= 缺省 wb）"""
    cat, ctx, err = _upload_cat_guard()
    if err:
        return err
    data = request.get_json(force=True, silent=True) or {}
    dx = (data.get("dx") or request.args.get("dx", "")).strip()
    filename = (data.get("file") or request.args.get("file", "")).strip()
    if not _dx_re(ctx["prefix"]).match(dx) or not filename or "/" in filename or "\\" in filename:
        return jsonify({"ok": False, "error": "参数非法"}), 400
    target = ctx["projects"] / dx / "03_UPLOAD" / filename
    if not target.exists():
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    # 同色联动删除：删一张某颜色 → 同款该颜色全部贴图（平铺/模特/BW合成）一起删
    targets = [target]
    try:
        _engine_dir = str(BASE_DIR / "04_OS" / "engine")
        if _engine_dir not in sys.path:
            sys.path.insert(0, _engine_dir)
        import wb_naming as _wbn
        info = _wbn.classify(dx, filename)
        if info and info.get("color"):
            color = info["color"]
            up_dir = ctx["projects"] / dx / "03_UPLOAD"
            targets = [f for f in up_dir.iterdir()
                       if f.is_file() and f.suffix.lower() in (".jpg", ".jpeg", ".png")
                       and (_wbn.classify(dx, f.name) or {}).get("color") == color]
            if target not in targets:
                targets.append(target)
    except Exception as e:
        print(f"[upload/delete] 同色联动解析失败，退回单张删除: {e}")
        targets = [target]
    deleted = []
    for t in targets:
        if send_to_recycle_bin(str(t)):
            deleted.append(t.name)
            safe_name = re.sub(r'[\\/*?:"<>|]', '_', t.name)
            for tf in ctx["upload_thumb"].glob(f"{dx}__{safe_name}.*"):
                try:
                    tf.unlink()
                except Exception:
                    pass
    ok = len(deleted) == len(targets) and bool(deleted)
    if len(deleted) > 1:
        msg = f"已送回收站 {len(deleted)} 张（同色联动）: {', '.join(deleted[:6])}{'…' if len(deleted) > 6 else ''}"
    else:
        msg = f"已送回收站: {filename}" if deleted else "删除失败"
    return jsonify({"ok": ok, "msg": msg, "deleted": deleted})


def _read_completed_md():
    """读取 已上款货号_wb.md 中的所有 DX 货号（已弃用，仅兼容旧逻辑）"""
    md = BASE_DIR / "已上款货号_wb.md"
    if not md.exists():
        return set()
    try:
        with open(md, "r", encoding="utf-8") as f:
            return set(
                line.strip().lstrip("- *").strip()
                for line in f
                if line.strip().startswith("DX")
                or line.strip().startswith("- DX")
                or line.strip().startswith("* DX")
            )
    except Exception:
        return set()


def _read_online_listed(cat=None):
    """读取店小秘在线产品页抓取的已上款款号集合（唯一权威来源，按品类落各自根）"""
    f = _online_listed_file(cat)
    if not f.exists():
        return set()
    try:
        data = json.loads(f.read_text(encoding="utf-8"))
        dx_set = data.get("dx_set", []) or []
        return set(str(dx).upper() for dx in dx_set)
    except Exception:
        return set()


def _online_listed_updated_at(cat=None):
    """返回在线已上款数据最后更新时间"""
    f = _online_listed_file(cat)
    if not f.exists():
        return None
    try:
        data = json.loads(f.read_text(encoding="utf-8"))
        return data.get("updated_at")
    except Exception:
        return None


def _online_listed_mode(cat=None):
    """返回在线已上款数据上次刷新的模式（quick/deep），供前端显示"""
    f = _online_listed_file(cat)
    if not f.exists():
        return None
    try:
        data = json.loads(f.read_text(encoding="utf-8"))
        return data.get("mode")
    except Exception:
        return None


def _remove_from_completed_md(dx_list, cat=None):
    """强制重新上款时，从已上款记录（已上款货号_<cat>.md）中删除指定款号行。
    返回实际删除了哪些款号（list）；文件不存在返回 []；写入失败返回 None（调用方必须拦截，不得继续上款）。
    cat 缺省 wb（路径与改造前一致）。"""
    md = _completed_md_for(cat)
    if not md.exists():
        return []
    targets = set(dx_list)
    # Permission denied 多为瞬时锁（杀毒/同步盘/前次进程未退），重试 3 次再放弃
    for attempt in range(3):
        removed = []
        try:
            with open(md, "r", encoding="utf-8") as f:
                lines = f.readlines()
            new_lines = []
            for line in lines:
                stripped = line.strip()
                # 匹配 '- SKU' 或 '* SKU'（前缀无关，DX/HX 通用）
                if stripped.startswith("- ") or stripped.startswith("* "):
                    dx = stripped.lstrip("- *").strip()
                    if dx in targets:
                        removed.append(dx)
                        continue
                new_lines.append(line)
            if removed:
                with open(md, "w", encoding="utf-8") as f:
                    f.writelines(new_lines)
            return removed
        except PermissionError as e:
            print(f"[batch-upload] 删除已上款记录被占用(第{attempt+1}次): {e}", flush=True)
            time.sleep(0.5)
        except Exception as e:
            print(f"[batch-upload] 删除已上款记录失败: {e}", flush=True)
            return None
    print(f"[batch-upload] 删除已上款记录连续3次被占用，放弃: {md}", flush=True)
    return None


def _remove_from_title_cache(dx_list, cat=None):
    """强制重新上款时，从标题缓存（标题缓存_<cat>.md）中删除指定款号的标题块，
    让 wb_listing.py 重新走豆包生成新标题（否则命中缓存会跳过豆包）。
    返回实际删除了哪些款号（list）；文件不存在返回 []；写入失败返回 None（调用方必须拦截）。
    cat 缺省 wb（路径与改造前一致）。"""
    md = _title_cache_for(cat)
    if not md.exists():
        return []
    # Permission denied 多为瞬时锁（杀毒/同步盘/前次进程未退），重试 3 次再放弃
    for attempt in range(3):
        removed = []
        try:
            with open(md, "r", encoding="utf-8") as f:
                content = f.read()
            for dx in dx_list:
                # 缓存块格式：## DXxxxx\n- 中文：...\n- 英文：...\n- 时间：...
                pattern = rf"## {re.escape(dx)}\s*\n.*?(?=\n## |\Z)"
                if re.search(pattern, content, re.DOTALL):
                    content = re.sub(pattern, "", content, count=1, flags=re.DOTALL)
                    removed.append(dx)
            if removed:
                content = re.sub(r"\n{3,}", "\n\n", content).rstrip() + "\n"
                with open(md, "w", encoding="utf-8") as f:
                    f.write(content)
            return removed
        except PermissionError as e:
            print(f"[batch-upload] 删除标题缓存被占用(第{attempt+1}次): {e}", flush=True)
            time.sleep(0.5)
        except Exception as e:
            print(f"[batch-upload] 删除标题缓存失败: {e}", flush=True)
            return None
    print(f"[batch-upload] 删除标题缓存连续3次被占用，放弃: {md}", flush=True)
    return None


@app.route('/api/upload/progress')
def api_upload_progress():
    """返回 wb_listing.py 写入的上款进度 JSON，并合并历史已完成记录（?cat= 缺省 wb）"""
    cat, ctx, err = _upload_cat_guard()
    if err:
        return err
    data = {
        "ok": True,
        "running": False,
        "started_at": None,
        "finished_at": None,
        "selected": [],
        "pending": [],
        "completed": [],
        "failed": [],
        "current": None,
        "current_start": None,
        "total_count": 0,
        "done_count": 0,
        "fail_count": 0,
        "per_dx": {},
    }
    if _upload_progress_file(cat).exists():
        try:
            with open(_upload_progress_file(cat), "r", encoding="utf-8") as f:
                data.update(json.load(f))
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # 只统计当前选中款范围内的完成/失败，避免历史记录把 done_count 撑爆 total_count
    selected_set = set(data.get("selected", []))
    failed_set = set(data.get("failed", [])) & selected_set
    online_set = _read_online_listed(cat)
    # 店小秘在线产品页为唯一权威来源；同时保留当前运行中的 completed（wb_listing.py 实时写入）
    completed_set = (set(data.get("completed", [])) | online_set) & selected_set

    data["completed"] = sorted(completed_set)
    data["failed"] = sorted(failed_set)
    data["pending"] = sorted(selected_set - completed_set - failed_set)
    data["done_count"] = len(completed_set)
    data["fail_count"] = len(failed_set)
    data["total_count"] = len(selected_set)

    # 在线已上款信息（权威来源）
    data["online_set"] = sorted(online_set & selected_set)
    data["online_count"] = len(online_set & selected_set)
    data["online_updated_at"] = _online_listed_updated_at(cat)

    return jsonify(data)


@app.route('/api/upload/refresh-online-listed', methods=['POST'])
def api_upload_refresh_online_listed():
    """启动 check_online_listed.py，从店小秘在线产品页刷新真正已上款的 DX 集合

    mode=incremental（默认）：日常增量，翻到上次边界款为止，集合相减自动移除下架款；首次运行全量建库
    mode=deep：深度清理，翻完所有页，全量覆盖（准确移除所有已下架款，并重置边界）
    """
    cat, ctx, cat_err = _upload_cat_guard()
    if cat_err:
        return cat_err
    mode = (request.args.get("mode") or "incremental").lower()
    if mode not in ("incremental", "deep"):
        mode = "incremental"

    lock_file = _cat_root(cat or _DEFAULT_CAT) / ".check_online_listed.lock"
    if lock_file.exists():
        # 用户要求：点击即强制重新刷新。读锁内 PID：
        # - 进程已死（崩溃残留锁）→ 直接删锁
        # - 进程还活着 → 先杀掉再删锁，然后重新启动
        killed = None
        try:
            pid = int(lock_file.read_text(encoding="utf-8").strip())
        except Exception:
            pid = None
        if pid:
            try:
                chk = subprocess.run(
                    ["tasklist", "/FI", f"PID eq {pid}", "/FO", "CSV", "/NH"],
                    capture_output=True, text=True, timeout=10)
                alive = str(pid) in chk.stdout
            except Exception:
                alive = False
            if alive:
                try:
                    subprocess.run(["taskkill", "/PID", str(pid), "/T", "/F"],
                                   capture_output=True, timeout=15)
                    killed = pid
                    import time as _time
                    _time.sleep(1)
                except Exception:
                    pass
        try:
            lock_file.unlink()
        except Exception:
            pass
        if killed:
            print(f"[刷新已上款] 强制结束旧任务 PID={killed}，重新启动")

    default_script = str(WB_LISTING_DIR / "check_online_listed.py")
    script_path = Path(default_script)
    if not script_path.exists():
        return jsonify({
            "ok": False,
            "error": f"刷新脚本不存在: {default_script}"
        }), 404

    try:
        # 非 wb 品类注入 WB_LISTING_CAT，使 check_online_listed.py 按品类抓取/落文件（卫衣→D:\Semems Hoodie）
        extra_env = {"WB_LISTING_CAT": cat} if cat and cat != _DEFAULT_CAT else None
        proc = run_minimized([sys.executable, str(script_path), "--mode", mode], wait=False, no_console=True, env=extra_env)
        mode_label = "深度清理" if mode == "deep" else "增量刷新"
        return jsonify({
            "ok": True,
            "msg": f"已开始刷新在线已上款（{mode_label}），完成后页面自动刷新",
            "pid": proc.pid,
            "mode": mode,
        })
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": f"启动刷新脚本失败: {e}"
        }), 500


def _open_folder_front(folder_path: Path):
    """打开文件夹并尝试强制资源管理器窗口前台显示。"""
    folder = str(folder_path)
    # 允许当前进程创建/激活前台窗口
    try:
        import ctypes
        ctypes.windll.user32.AllowSetForegroundWindow(-1)
    except Exception:
        pass

    # 使用 explorer.exe 打开，避免 os.startfile 复用已最小化窗口时不激活
    try:
        subprocess.Popen(['explorer.exe', folder])
    except Exception:
        try:
            os.startfile(folder)
        except Exception:
            subprocess.Popen(f'explorer.exe "{folder}"', shell=True)

    # 尝试找到新打开的资源管理器窗口并置顶
    try:
        import win32gui
        import win32con
        import time

        folder_name = folder_path.name
        best_hwnd = None

        def _enum(hwnd, _):
            nonlocal best_hwnd
            if not win32gui.IsWindowVisible(hwnd):
                return True
            title = win32gui.GetWindowText(hwnd)
            # 资源管理器窗口标题通常包含文件夹名；也可用类名 CabinetWClass
            if folder_name in title and 'CabinetWClass' in win32gui.GetClassName(hwnd):
                best_hwnd = hwnd
                return False
            return True

        # 轮询最多 1 秒，等待窗口创建
        for _ in range(10):
            time.sleep(0.1)
            win32gui.EnumWindows(_enum, None)
            if best_hwnd:
                break

        if best_hwnd:
            win32gui.ShowWindow(best_hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(best_hwnd)
    except Exception:
        pass


@app.route('/api/open')
def api_open_dx():
    """打开指定款的子文件夹（ai/rem/up）；?cat= 缺省 wb，卫衣走 D:\\Semems Hoodie"""
    cat, ctx, err = _upload_cat_guard()
    if err:
        return err
    dx = request.args.get("dx", "")
    which = request.args.get("which", "")
    if not _dx_re(ctx["prefix"]).match(dx) or which not in ("ai", "rem", "up"):
        return jsonify({"ok": False, "error": "参数非法"}), 400
    sub = {"ai": "01_AI", "rem": "02_REM_BG", "up": "03_UPLOAD"}[which]
    folder = ctx["projects"] / dx / sub
    if folder.exists():
        _open_folder_front(folder)
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "文件夹不存在"}), 404


@app.route('/api/open-file')
def api_open_file():
    """打开指定 DX 子目录中的文件所在文件夹，并选中该文件。

    参数：
      dx: DX 编号
      file: 文件名
      sub: 子目录（默认 01_AI，可选 02_REM_BG / 03_UPLOAD / INBOX）
    """
    dx = request.args.get("dx", "").strip()
    filename = request.args.get("file", "").strip()
    sub = request.args.get("sub", "01_AI").strip()

    cat, cat_err = _resolve_request_cat()
    if cat_err:
        return cat_err
    ctx = _cat_ctx(cat)

    if not dx or not filename:
        return jsonify({"ok": False, "error": "缺少 dx 或 file"}), 400
    if "/" in filename or "\\" in filename or ".." in filename:
        return jsonify({"ok": False, "error": "非法文件名"}), 400

    if sub == "INBOX":
        folder = ctx["inbox"]
    elif sub in ("01_AI", "02_REM_BG", "03_UPLOAD") and _dx_re(ctx["prefix"]).match(dx):
        folder = ctx["projects"] / dx / sub
    else:
        return jsonify({"ok": False, "error": "非法 sub 参数"}), 400

    target = folder / filename
    if not target.exists():
        return jsonify({"ok": False, "error": f"文件不存在: {target}"}), 404

    try:
        # /select 参数会打开文件夹并高亮选中指定文件，保证前台显示
        subprocess.Popen(
            ['explorer.exe', '/select,', str(target)],
            shell=False,
        )
        return jsonify({"ok": True, "path": str(target)})
    except Exception as e:
        # 回退：仅打开文件夹
        try:
            os.startfile(str(folder))
            return jsonify({"ok": True, "path": str(folder), "fallback": True})
        except Exception:
            return jsonify({"ok": False, "error": str(e)}), 500


def _kill_stale_wb_listing():
    """启动新上款任务前，清理仍在运行的旧 wb_listing.py 进程。
    僵尸进程的危害：①占用已上款记录/标题缓存文件 → 强制重新上款删记录 Permission denied；
    ②与新任务并发操作同一浏览器页面。只杀命令行含 wb_listing.py 的 python 进程，不误杀别的。
    返回被杀掉的 PID 列表。"""
    killed = []
    try:
        out = subprocess.check_output([
            "powershell", "-NoProfile", "-Command",
            "Get-CimInstance Win32_Process -Filter \"Name like '%python%'\" | "
            "Where-Object { $_.CommandLine -like '*wb_listing.py*' } | "
            "Select-Object -ExpandProperty ProcessId"
        ], text=True, timeout=15)
        for line in out.splitlines():
            pid = line.strip()
            if pid.isdigit():
                r = subprocess.run(["taskkill", "/F", "/PID", pid],
                                   capture_output=True, text=True, timeout=10)
                if r.returncode == 0:
                    killed.append(int(pid))
    except Exception as e:
        print(f"[batch-upload] 清理旧上款进程失败: {e}", flush=True)
    if killed:
        print(f"[batch-upload] 已清理旧上款进程: {killed}", flush=True)
        time.sleep(0.5)  # 等文件句柄释放，避免紧随其后的删记录仍撞锁
    return killed


@app.route('/api/batch-upload', methods=['POST'])
def api_batch_upload():
    """批量上款：调用 E:\Claude code\wb上款\wb_listing.py 逐个 DX 上款。
    可通过环境变量 LOVART_UPLOAD_SCRIPT 覆盖脚本路径。
    """
    cat, ctx, cat_err = _upload_cat_guard()
    if cat_err:
        return cat_err
    data = request.get_json(silent=True) or {}
    dx_list = data.get("dx_list", [])
    force = data.get("force", False)
    if not dx_list:
        return jsonify({"ok": False, "error": "请指定DX号"}), 400

    # 防并发：本品类已有上款任务进行中（progress running 且 30 分钟内）时拒绝重复启动。
    # 背景：用户连续点两次「批量上传」会开两个 wb_listing 进程抢同一个 Edge，
    # 导致后一个进程 playwright 连接断开崩溃（2026-08-21 实测）。
    _pf = _upload_progress_file(cat)
    if _pf.exists():
        try:
            _pd = json.loads(_pf.read_text(encoding="utf-8"))
            if _pd.get("running") and _pd.get("started_at"):
                from datetime import datetime as _dt
                try:
                    _start = _dt.fromisoformat(str(_pd["started_at"]))
                    _age = (_dt.now() - _start).total_seconds()
                    if 0 <= _age < 30 * 60:
                        return jsonify({
                            "ok": False,
                            "error": f"已有上款任务正在运行（{int(_age)}s 前启动）。请等它跑完再点，不要重复点击（重复点击会开两个进程抢浏览器导致崩溃）。"
                        }), 409
                except Exception:
                    pass
        except Exception:
            pass

    default_script = str(WB_LISTING_DIR / "wb_listing.py")
    upload_script = os.environ.get("LOVART_UPLOAD_SCRIPT", default_script)
    script_path = Path(upload_script)
    if not script_path.exists():
        return jsonify({
            "ok": False,
            "error": f"上款脚本不存在: {upload_script}"
        }), 404

    # 启动前先清理仍在运行的旧 wb_listing 进程（僵尸进程会占用记录文件导致强制上款删记录失败，
    # 且可能与新任务并发操作同一页面）；用户规则：只执行最新的上款任务
    stale_killed = _kill_stale_wb_listing()

    # 强制重新上款：先从已上款记录中删除对应款号，让 wb_listing.py 正常执行
    # 同时清除标题缓存，强制重新走豆包用最新提示词生成标题
    removed = []
    cache_removed = []
    if force:
        removed = _remove_from_completed_md(dx_list, cat=cat)
        cache_removed = _remove_from_title_cache(dx_list, cat=cat)
        # 删除失败（文件被占用等）必须拦截：否则记录还在 → wb_listing 判"已上款"跳过，
        # 表现为"开了 Edge 却不传图"的静默空跑，用户完全看不到原因
        if removed is None or cache_removed is None:
            which = "、".join([n for n, r in (("已上款记录", removed), ("标题缓存", cache_removed)) if r is None])
            return jsonify({
                "ok": False,
                "error": f"强制重新上款失败：{which}文件被占用（Permission denied，重试3次仍失败）。请等当前上款进程结束或关闭占用程序后重试。"
            }), 500

    # wb_listing.py --only 模式：只处理勾选的确切款号，不会继续后续款
    valid_dx = []
    for dx in dx_list:
        dx_folder = ctx["projects"] / dx
        if dx_folder.exists() and (dx_folder / "03_UPLOAD").exists():
            valid_dx.append(dx)

    if not valid_dx:
        return jsonify({"ok": False, "error": "勾选的款均无 03_UPLOAD 成品"}), 400

    args = [sys.executable, str(script_path)]
    for dx in valid_dx:
        args.extend(["--only", dx])

    # 重置进度文件：避免前端读到上次任务的完成状态而立即显示"上款完成"
    try:
        now_iso = datetime.now().isoformat(timespec="seconds")
        _upload_progress_file(cat).write_text(
            json.dumps({
                "running": True,
                "started_at": now_iso,
                "finished_at": None,
                "selected": valid_dx,
                "pending": valid_dx,
                "completed": [],
                "failed": [],
                "current": None,
                "current_start": None,
                "total_count": len(valid_dx),
                "done_count": 0,
                "fail_count": 0,
                "per_dx": {},
                "updated_at": now_iso,
            }, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        print(f"[batch-upload] 重置进度文件失败: {e}", flush=True)

    try:
        # wait=False: wb_listing.py 运行时间较长，API 立即返回，后台执行
        # no_console=True: 不弹控制台黑窗（wb_listing.py 自己写日志到品类根 _debug）
        # 非 wb 品类注入 WB_LISTING_CAT，使 wb_listing.py 解析到对应数据根与引用模板（HX0000）
        extra_env = {"WB_LISTING_CAT": cat} if cat and cat != _DEFAULT_CAT else None
        run_minimized(args, wait=False, no_console=True, env=extra_env)
    except Exception as e:
        print(f"[batch-upload] 启动 {valid_dx} 失败: {e}", flush=True)
        return jsonify({"ok": False, "error": f"启动脚本失败: {e}"}), 500

    msg = f"已启动 wb上款脚本，精确处理 {len(valid_dx)} 个款：{', '.join(valid_dx)}"
    if force:
        msg = f"【强制重新上款】已删除已上款记录中的 {len(removed)} 个款、清除标题缓存 {len(cache_removed)} 个（将重新生成标题），并启动处理：{', '.join(valid_dx)}"
    return jsonify({
        "ok": True,
        "msg": msg,
        "script": str(script_path),
        "selected": valid_dx,
        "force": force,
        "stale_killed": stale_killed,
        "removed": removed,
        "cache_removed": cache_removed,
    })


@app.route('/api/upload/stop', methods=['POST'])
def api_upload_stop():
    """停止当前品类的上款任务：按 progress 里的 pid 终止 wb_listing.py 进程，
    并复位 running 状态，让用户可以重新开始。
    只杀 wb_listing 主进程（不 /T 连 Edge 一起杀——保留 Edge 登录态供下次复用）。
    """
    cat, _ctx, cat_err = _upload_cat_guard()
    if cat_err:
        return cat_err
    pf = _upload_progress_file(cat)
    pid = None
    if pf.exists():
        try:
            pid = (json.loads(pf.read_text(encoding="utf-8")) or {}).get("pid")
        except Exception:
            pid = None

    killed = []
    if pid:
        try:
            r = subprocess.run(["taskkill", "/PID", str(pid), "/F"],
                               capture_output=True, text=True, timeout=15)
            if r.returncode == 0:
                killed.append(f"进程 {pid}")
            else:
                killed.append(f"进程 {pid}（可能已退出: {r.stderr.strip()[:60]}）")
        except Exception as e:
            killed.append(f"进程 {pid} 终止异常: {e}")
    else:
        killed.append("progress 无 pid（任务可能已结束或为旧版本启动）")

    # 复位 running，让用户可以重新开始
    try:
        now_iso = datetime.now().isoformat(timespec="seconds")
        pf.write_text(json.dumps({
            "running": False,
            "pid": None,
            "started_at": None,
            "finished_at": now_iso,
            "selected": [],
            "pending": [],
            "completed": [],
            "failed": [],
            "current": None,
            "current_start": None,
            "total_count": 0,
            "done_count": 0,
            "fail_count": 0,
            "per_dx": {},
            "updated_at": now_iso,
        }, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        return jsonify({"ok": False, "error": f"复位进度文件失败: {e}"}), 500

    return jsonify({"ok": True, "msg": f"已停止上款任务（{', '.join(killed)}），可以重新开始"})


# ============================================================================
# Temu 核价（Hermes）集成
# ============================================================================

def _read_pricing_state():
    """读取 Hermes 核价状态文件，失败返回空字典。"""
    if not PRICING_STATE_FILE.exists():
        return {}
    try:
        return json.loads(PRICING_STATE_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[pricing] 读取状态失败: {e}", flush=True)
        return {}


def _pricing_log_reader(proc, mode):
    """后台线程：读取核价脚本 stdout/stderr 并写入 pricing_task 日志。"""
    def _read_stream(stream, kind):
        try:
            for raw in iter(stream.readline, b""):
                # Hermes 脚本在 PIPE 下受 PYTHONIOENCODING=utf-8 影响输出 UTF-8；优先 UTF-8，失败回退 GBK
                line = None
                for enc in ("utf-8", "gbk", "gb2312"):
                    try:
                        line = raw.decode(enc, errors="strict").rstrip("\r\n")
                        break
                    except Exception:
                        continue
                if line is None:
                    line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
                if not line:
                    continue
                with pricing_lock:
                    pricing_task["log"].append({"line": line, "kind": kind})
        except Exception as e:
            with pricing_lock:
                pricing_task["log"].append({"line": f"日志读取异常: {e}", "kind": "error"})
        finally:
            try:
                stream.close()
            except Exception:
                pass

    threads = []
    if proc.stdout:
        t = threading.Thread(target=_read_stream, args=(proc.stdout, ""), daemon=True)
        t.start()
        threads.append(t)
    if proc.stderr:
        t = threading.Thread(target=_read_stream, args=(proc.stderr, "error"), daemon=True)
        t.start()
        threads.append(t)

    # 等待进程结束
    rc = proc.wait()
    for t in threads:
        t.join(timeout=2)

    elapsed = 0
    if pricing_task.get("started_at"):
        try:
            elapsed = int((datetime.now() - datetime.fromisoformat(pricing_task["started_at"])).total_seconds())
        except Exception:
            pass

    with pricing_lock:
        pricing_task["elapsed_sec"] = elapsed
        state = _read_pricing_state()
        pricing_task["page_records"] = state.get("page_records", [])
        pricing_task["processed_pages"] = len(pricing_task["page_records"])
        if pricing_task["status"] == "running":
            if rc == 0:
                pricing_task["status"] = "completed"
                pricing_task["task_label"] = f"{mode} 完成"
            else:
                pricing_task["status"] = "error"
                pricing_task["task_label"] = f"{mode} 退出码 {rc}"
        pricing_task["completed_at"] = datetime.now().isoformat()
        pricing_task["proc"] = None


def _start_pricing_script(mode, args, label, extra_env=None):
    """通用启动 Hermes 核价子进程。extra_env: 附加环境变量（如底价覆盖文件路径）。"""
    with pricing_lock:
        if pricing_task.get("status") == "running" and pricing_task.get("proc") and pricing_task["proc"].poll() is None:
            return {"error": "已有核价任务在运行，请先停止"}, 409

        pricing_task["status"] = "running"
        pricing_task["mode"] = mode
        pricing_task["task_label"] = label
        pricing_task["started_at"] = datetime.now().isoformat()
        pricing_task["completed_at"] = None
        pricing_task["log"] = [{"line": f"[{datetime.now().strftime('%H:%M:%S')}] 启动: {label}", "kind": ""}]
        pricing_task["log_index"] = 0
        pricing_task["processed_pages"] = 0
        pricing_task["elapsed_sec"] = 0
        pricing_task["page_records"] = []

    if not PRICING_DIR.exists():
        return {"error": f"核价项目目录不存在: {PRICING_DIR}"}, 404

    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env.setdefault("PYTHONIOENCODING", "utf-8")
    if extra_env:
        env.update(extra_env)

    try:
        proc = subprocess.Popen(
            args,
            cwd=str(PRICING_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,
            env=env,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    except Exception as e:
        with pricing_lock:
            pricing_task["status"] = "error"
            pricing_task["task_label"] = f"启动失败: {e}"
            pricing_task["completed_at"] = datetime.now().isoformat()
            pricing_task["proc"] = None
        return {"error": f"启动脚本失败: {e}"}, 500

    with pricing_lock:
        pricing_task["proc"] = proc

    threading.Thread(target=_pricing_log_reader, args=(proc, mode), daemon=True).start()
    return {"ok": True, "msg": f"已启动 {label}"}, 200


@app.route('/pricing')
def pricing_page():
    """Temu 核价页面。"""
    return send_file(str(Path(__file__).parent / 'pricing.html'))


@app.route('/api/pricing/start', methods=['POST'])
def api_pricing_start():
    """启动完整核价或仅核价不提交。

    body: {"mode": "full" | "no-submit"}
    """
    data = request.get_json(silent=True) or {}
    mode = data.get("mode", "full")
    pages = (data.get("pages") or "").strip()
    args = [get_python(), str(PRICING_MAIN)]
    if mode == "no-submit":
        args.append("--no-submit")
        # 仅核价不提交：始终只核 1 页并停在提交前供检查，忽略页码范围
    elif pages:
        import re
        if not re.match(r'^\d+-\d+$', pages):
            return jsonify({"error": "页码范围格式应为 起始-结束，如 2-52"}), 400
        args.append(f"--pages={pages}")
    label = "仅核价不提交" if mode == "no-submit" else "完整自动核价"
    if pages and mode != "no-submit":
        label += f" 页{pages}"
    resp, code = _start_pricing_script(mode, args, label)
    return jsonify(resp), code


@app.route('/api/pricing/continue', methods=['POST'])
def api_pricing_continue():
    """从已填价状态继续提交。"""
    script = PRICING_ENTRYPOINT / "continue_run.py"
    if not script.exists():
        return jsonify({"error": f"脚本不存在: {script}"}), 404
    resp, code = _start_pricing_script("continue", [get_python(), str(script)], "继续提交")
    return jsonify(resp), code


@app.route('/api/pricing/retry', methods=['POST'])
def api_pricing_retry():
    """重试指定页。body: {"pages": "2 5"} 或 {"pages": "2-52"}（支持区间）。"""
    import re
    data = request.get_json(silent=True) or {}
    raw = (data.get("pages") or "").strip()
    if not raw:
        return jsonify({"error": "请输入页码"}), 400
    pages_list = []
    for tok in raw.split():
        m = re.match(r'^(\d+)-(\d+)$', tok)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            if a > b:
                return jsonify({"error": f"页码区间无效: {tok}"}), 400
            pages_list.extend(str(i) for i in range(a, b + 1))
        else:
            if not re.match(r'^\d+$', tok):
                return jsonify({"error": f"页码格式无效: {tok}"}), 400
            pages_list.append(tok)
    if not pages_list:
        return jsonify({"error": "请输入页码"}), 400
    script = PRICING_ENTRYPOINT / "retry_pages.py"
    if not script.exists():
        return jsonify({"error": f"脚本不存在: {script}"}), 404
    args = [get_python(), str(script)] + pages_list
    resp, code = _start_pricing_script("retry", args, f"重试页 {raw}")
    return jsonify(resp), code


@app.route('/api/pricing/reprice', methods=['POST'])
def api_pricing_reprice():
    """按模拟档位底价重新核价（临时覆盖，不改原始底价表）。

    body: {"overrides": {"德国站": 61.14, ...}, "pages": "2-52"(可选)}
    覆盖通过 HENGJIA_FLOOR_OVERRIDE 环境变量传给核价子进程，
    常规「完整自动核价」不带此变量，永远按原始底价。
    """
    data = request.get_json(silent=True) or {}
    overrides = data.get("overrides") or {}
    if not isinstance(overrides, dict) or not overrides:
        return jsonify({"error": "请提供 overrides（站点→模拟底价）"}), 400
    clean = {}
    for site, v in overrides.items():
        if isinstance(v, (int, float)) and v > 0:
            clean[str(site)] = round(float(v), 2)
    if not clean:
        return jsonify({"error": "overrides 无有效项"}), 400

    override_path = PRICING_DIR / "state" / "floor_override.json"
    try:
        override_path.parent.mkdir(parents=True, exist_ok=True)
        override_path.write_text(json.dumps(clean, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError as e:
        return jsonify({"error": f"写入覆盖文件失败: {e}"}), 500

    pages = (data.get("pages") or "").strip()
    args = [get_python(), str(PRICING_MAIN)]
    if pages:
        import re
        if not re.match(r'^\d+-\d+$', pages):
            return jsonify({"error": "页码范围格式应为 起始-结束，如 2-52"}), 400
        args.append(f"--pages={pages}")
    desc = "，".join(f"{s}→{v}" for s, v in clean.items())
    label = f"按模拟底价重新核价（{desc}）"
    if pages:
        label += f" 页{pages}"
    resp, code = _start_pricing_script(
        "reprice", args, label,
        extra_env={"HENGJIA_FLOOR_OVERRIDE": str(override_path)},
    )
    return jsonify(resp), code


@app.route('/api/pricing/export', methods=['POST'])
def api_pricing_export():
    """导出核价结果到 Excel。"""
    script = PRICING_ENTRYPOINT / "export_prices.py"
    if not script.exists():
        return jsonify({"error": f"脚本不存在: {script}"}), 404
    resp, code = _start_pricing_script("export", [get_python(), str(script)], "导出核价结果")
    return jsonify(resp), code


@app.route('/api/pricing/stop', methods=['POST'])
def api_pricing_stop():
    """停止当前核价任务。"""
    with pricing_lock:
        proc = pricing_task.get("proc")
        if proc and proc.poll() is None:
            try:
                proc.terminate()
                try:
                    proc.wait(timeout=3)
                except subprocess.TimeoutExpired:
                    proc.kill()
            except Exception as e:
                return jsonify({"error": f"停止失败: {e}"}), 500
        pricing_task["status"] = "stopped"
        pricing_task["task_label"] = "已停止"
        pricing_task["completed_at"] = datetime.now().isoformat()
        pricing_task["proc"] = None
    return jsonify({"ok": True, "msg": "已停止核价任务"})


@app.route('/api/pricing/status')
def api_pricing_status():
    """获取核价任务状态、增量日志和分页记录。"""
    with pricing_lock:
        state = _read_pricing_state()
        page_records = state.get("page_records", [])
        processed = len(page_records)

        # 计算运行时长
        elapsed = pricing_task.get("elapsed_sec", 0)
        if pricing_task.get("status") == "running" and pricing_task.get("started_at"):
            try:
                elapsed = int((datetime.now() - datetime.fromisoformat(pricing_task["started_at"])).total_seconds())
            except Exception:
                pass

        # 返回未读取过的日志（使用绝对长度作为下标，避免追加日志时漏读）
        idx = pricing_task.get("log_index", 0)
        all_logs = pricing_task.get("log", [])
        logs = all_logs[idx:]
        pricing_task["log_index"] = len(all_logs)

        return jsonify({
            "status": pricing_task.get("status", "idle"),
            "mode": pricing_task.get("mode"),
            "task_label": pricing_task.get("task_label", ""),
            "task": pricing_task.get("task_label", ""),
            "started_at": pricing_task.get("started_at"),
            "completed_at": pricing_task.get("completed_at"),
            "processed_pages": processed,
            "elapsed_sec": elapsed,
            "page_records": page_records,
            "log": logs,
        })


@app.route('/api/pricing/result-files')
def api_pricing_result_files():
    """列出 OUTPUT_DIR 中的核价 Excel 结果文件。"""
    files = []
    if PRICING_OUTPUT_DIR.exists():
        for p in sorted(PRICING_OUTPUT_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            size = p.stat().st_size
            size_str = f"{size/1024/1024:.2f} MB" if size > 1024*1024 else f"{size/1024:.1f} KB"
            files.append({
                "name": p.name,
                "path": str(p),
                "size": size_str,
                "mtime": datetime.fromtimestamp(p.stat().st_mtime).isoformat(),
            })
    return jsonify({"files": files})


@app.route('/api/pricing/download')
def api_pricing_download():
    """下载核价结果 Excel 文件。"""
    filename = request.args.get("file", "").strip()
    if not filename:
        return jsonify({"error": "请指定文件名"}), 400
    # 安全校验：只取文件名，不允许路径穿越
    filename = os.path.basename(filename)
    if not filename.endswith(".xlsx"):
        return jsonify({"error": "仅支持 .xlsx 文件"}), 400
    path = PRICING_OUTPUT_DIR / filename
    if not path.exists():
        return jsonify({"error": "文件不存在"}), 404
    return send_file(str(path), as_attachment=True, download_name=filename)


@app.route('/api/pricing/fail_analysis')
def api_pricing_fail_analysis():
    """读取 Hermes 核价引擎每页落盘的「不通过核价底价分析」（底价模拟）。"""
    path = PRICING_DIR / "state" / "fail_analysis.json"
    if not path.exists():
        return jsonify({"updated": None, "pages": [], "sites": []})
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return jsonify(data)
    except (json.JSONDecodeError, OSError) as e:
        return jsonify({"error": f"读取分析文件失败: {e}", "pages": [], "sites": []})


@app.route('/api/pricing/signal', methods=['POST'])
def api_pricing_signal():
    """创建 go.signal 文件，通知 Hermes 脚本用户已准备好开始核价。"""
    signal_path = PRICING_DIR / "go.signal"
    try:
        signal_path.write_text("go", encoding="utf-8")
        exists = signal_path.exists()
        return jsonify({"ok": True, "msg": "已发送 '好了' 信号，核价脚本将继续运行", "path": str(signal_path), "exists": exists})
    except Exception as e:
        return jsonify({"error": f"创建 signal 文件失败: {e}"}), 500


# ============================================================================
# Temu 报活动集成
# ============================================================================

def _read_activity_state():
    """读取报活动状态文件，失败或不存在返回空字典。"""
    path = Path(ACTIVITY_STATE_FILE)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[activity] 读取状态失败: {e}", flush=True)
        return {}


def _activity_log_reader(proc):
    """后台线程：读取报活动脚本 stdout/stderr 并写入 activity_task 日志。"""
    def _read_stream(stream, kind):
        try:
            for raw in iter(stream.readline, b""):
                line = None
                for enc in ("utf-8", "gbk", "gb2312"):
                    try:
                        line = raw.decode(enc, errors="strict").rstrip("\r\n")
                        break
                    except Exception:
                        continue
                if line is None:
                    line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
                if not line:
                    continue
                with activity_lock:
                    if kind == "error":
                        activity_task["log"].append(f"[ERR] {line}")
                    else:
                        activity_task["log"].append(line)
        except Exception as e:
            with activity_lock:
                activity_task["log"].append(f"[ERR] 日志读取异常: {e}")
        finally:
            try:
                stream.close()
            except Exception:
                pass

    threads = []
    if proc.stdout:
        t = threading.Thread(target=_read_stream, args=(proc.stdout, ""), daemon=True)
        t.start()
        threads.append(t)
    if proc.stderr:
        t = threading.Thread(target=_read_stream, args=(proc.stderr, "error"), daemon=True)
        t.start()
        threads.append(t)

    rc = proc.wait()
    for t in threads:
        t.join(timeout=2)

    with activity_lock:
        if activity_task["status"] == "running":
            if rc == 0:
                activity_task["status"] = "completed"
            else:
                activity_task["status"] = "error"
        activity_task["completed_at"] = datetime.now().isoformat()
        activity_task["proc"] = None


def _start_activity_script(label, extra_env=None):
    """通用启动 Temu 报活动子进程。extra_env: 透传给引擎的环境变量（如活动类型/折扣门槛）。"""
    with activity_lock:
        if activity_task.get("status") == "running" and activity_task.get("proc") and activity_task["proc"].poll() is None:
            return {"success": False, "message": "已有报活动任务在运行，请先停止"}, 409

        activity_task["status"] = "running"
        activity_task["started_at"] = datetime.now().isoformat()
        activity_task["completed_at"] = None
        activity_task["log"] = [f"[{datetime.now().strftime('%H:%M:%S')}] 启动: {label}"]
        activity_task["log_index"] = 0

    if not os.path.exists(ACTIVITY_DIR):
        with activity_lock:
            activity_task["status"] = "error"
            activity_task["completed_at"] = datetime.now().isoformat()
        return {"success": False, "message": f"报活动项目目录不存在: {ACTIVITY_DIR}"}, 404

    if not os.path.exists(ACTIVITY_ENTRYPOINT):
        with activity_lock:
            activity_task["status"] = "error"
            activity_task["completed_at"] = datetime.now().isoformat()
        return {"success": False, "message": f"报活动入口脚本不存在: {ACTIVITY_ENTRYPOINT}"}, 404

    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env.setdefault("PYTHONIOENCODING", "utf-8")
    if extra_env:
        env.update(extra_env)

    # 清除上一轮残留的勾选文件，避免新任务误消费旧勾选
    try:
        if os.path.exists(ACTIVITY_SELECTION_FILE):
            os.remove(ACTIVITY_SELECTION_FILE)
    except Exception as e:
        print(f"[activity] 清除残留勾选文件失败: {e}", flush=True)

    try:
        proc = subprocess.Popen(
            [get_python(), ACTIVITY_ENTRYPOINT],
            cwd=ACTIVITY_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,
            env=env,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    except Exception as e:
        with activity_lock:
            activity_task["status"] = "error"
            activity_task["completed_at"] = datetime.now().isoformat()
            activity_task["proc"] = None
        return {"success": False, "message": f"启动脚本失败: {e}"}, 500

    with activity_lock:
        activity_task["proc"] = proc

    threading.Thread(target=_activity_log_reader, args=(proc,), daemon=True).start()
    return {"success": True, "message": f"已启动 {label}"}, 200


def _retail_price_log_reader(proc):
    """后台线程：读取建议零售价脚本 stdout/stderr 并写入 retail_price_task 日志。"""
    def _read_stream(stream, kind):
        try:
            for raw in iter(stream.readline, b""):
                line = None
                for enc in ("utf-8", "gbk", "gb2312"):
                    try:
                        line = raw.decode(enc, errors="strict").rstrip("\r\n")
                        break
                    except Exception:
                        continue
                if line is None:
                    line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
                if not line:
                    continue
                with retail_price_lock:
                    retail_price_task["log"].append({"line": line, "kind": kind})
        except Exception as e:
            with retail_price_lock:
                retail_price_task["log"].append({"line": f"日志读取异常: {e}", "kind": "error"})
        finally:
            try:
                stream.close()
            except Exception:
                pass

    threads = []
    if proc.stdout:
        t = threading.Thread(target=_read_stream, args=(proc.stdout, ""), daemon=True)
        t.start()
        threads.append(t)
    if proc.stderr:
        t = threading.Thread(target=_read_stream, args=(proc.stderr, "error"), daemon=True)
        t.start()
        threads.append(t)

    rc = proc.wait()
    for t in threads:
        t.join(timeout=2)

    with retail_price_lock:
        elapsed = 0
        if retail_price_task.get("started_at"):
            try:
                elapsed = int((datetime.now() - datetime.fromisoformat(retail_price_task["started_at"])).total_seconds())
            except Exception:
                pass

        retail_price_task["elapsed_sec"] = elapsed
        if retail_price_task["status"] == "running":
            if rc == 0:
                retail_price_task["status"] = "completed"
                retail_price_task["task_label"] = "填写完成"
            else:
                retail_price_task["status"] = "error"
                retail_price_task["task_label"] = f"填写失败 (退出码 {rc})"
        retail_price_task["completed_at"] = datetime.now().isoformat()
        retail_price_task["proc"] = None


def _start_retail_price_script(label, diagnose=False):
    """通用启动 Temu 建议零售价填写子进程。diagnose=True 时附加 --diagnose 参数（仅 dump 抽屉结构，不填写/不提交）。"""
    with retail_price_lock:
        if retail_price_task.get("status") == "running" and retail_price_task.get("proc") and retail_price_task["proc"].poll() is None:
            return {"error": "已有建议零售价任务在运行，请先停止"}, 409

        retail_price_task["status"] = "running"
        retail_price_task["task_label"] = label
        retail_price_task["started_at"] = datetime.now().isoformat()
        retail_price_task["completed_at"] = None
        retail_price_task["log"] = [{"line": f"[{datetime.now().strftime('%H:%M:%S')}] 启动: {label}", "kind": ""}]
        retail_price_task["log_index"] = 0
        retail_price_task["elapsed_sec"] = 0

    if not RETAIL_PRICE_DIR.exists():
        with retail_price_lock:
            retail_price_task["status"] = "error"
            retail_price_task["completed_at"] = datetime.now().isoformat()
            retail_price_task["proc"] = None
        return {"error": f"建议零售价项目目录不存在: {RETAIL_PRICE_DIR}"}, 404

    if not RETAIL_PRICE_SCRIPT.exists():
        with retail_price_lock:
            retail_price_task["status"] = "error"
            retail_price_task["completed_at"] = datetime.now().isoformat()
            retail_price_task["proc"] = None
        return {"error": f"建议零售价脚本不存在: {RETAIL_PRICE_SCRIPT}"}, 404

    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env.setdefault("PYTHONIOENCODING", "utf-8")

    try:
        node_args = ["node", str(RETAIL_PRICE_SCRIPT), "--no-close-browser"]
        if diagnose:
            # --diagnose 插在脚本名之后、--no-close-browser 之前，保持参数顺序清晰
            node_args.insert(2, "--diagnose")
        proc = subprocess.Popen(
            node_args,
            cwd=str(RETAIL_PRICE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,
            env=env,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    except Exception as e:
        with retail_price_lock:
            retail_price_task["status"] = "error"
            retail_price_task["completed_at"] = datetime.now().isoformat()
            retail_price_task["proc"] = None
        return {"error": f"启动脚本失败: {e}"}, 500

    with retail_price_lock:
        retail_price_task["proc"] = proc

    threading.Thread(target=_retail_price_log_reader, args=(proc,), daemon=True).start()
    return {"ok": True, "msg": f"已启动 {label}"}, 200


@app.route('/activity')
def activity_page():
    """Temu 报活动页面。"""
    return send_file(str(Path(__file__).parent / 'activity.html'))


@app.route('/order-price')
def order_price_page():
    """Temu 价格申报视角 —— 单独页面，一键进入并自动点「待卖家确认」+ 设每页200条。"""
    html_file = Path(__file__).parent / 'order_price.html'
    if html_file.exists():
        return send_file(str(html_file))
    return "<h1>order_price.html not found</h1><p>请确保 order_price.html 与 bridge.py 在同一目录</p>", 404


def _edge_running():
    """是否已有 msedge 进程在运行（避免又开一个 Edge）。"""
    try:
        out = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq msedge.exe'],
                             capture_output=True, text=True, timeout=10).stdout
        return 'msedge.exe' in out.lower()
    except Exception:
        return False


def _kill_edge():
    """关闭所有 Edge 进程（用于把普通 Edge 换成带 9222 的共用 Edge）。"""
    try:
        subprocess.run(['taskkill', '/IM', 'msedge.exe', '/F'],
                       capture_output=True, text=True, timeout=15)
    except Exception:
        pass


def _ensure_edge_cdp(p, log, timeout=40):
    """连接本机 Edge CDP(9222)——这是大家共用的调试端口。

    规则：
    - 9222 可达 → 直接连（使用已开的、带调试端口的 Edge，不新开）。
    - 9222 不可达但已有其它 Edge 在跑（没带 9222）→ 先关掉它，再用标准
      端口/配置（9222 + C:\\edge-cdp-profile）重开那「一个」共用的 Edge
      （关掉旧的再开新的，绝不出现两个 Edge）。
    - 9222 不可达且无任何 Edge → 同样用标准配置启动那一个共用 Edge。

    返回 (browser, edge_auto_launched)。
    """
    import time as _t
    import urllib.request
    try:
        return p.chromium.connect_over_cdp("http://127.0.0.1:9222"), False
    except Exception:
        pass
    if _edge_running():
        log.append("⚠️ 检测到已有 Edge 在运行但未带 9222 端口，将关闭它并用共用端口(9222)重新打开（只开这一个，不会多开）…")
        _kill_edge()
        for _ in range(20):
            if not _edge_running():
                break
            _t.sleep(0.5)
    log.append("ℹ️ 启动共用的 Edge（端口 9222）…")
    edge = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    if not os.path.exists(edge):
        edge = r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
    if not os.path.exists(edge):
        raise RuntimeError("未找到 Edge 安装路径，请先安装 Microsoft Edge。")
    profile = r"C:\edge-cdp-profile"
    args = [edge, "--remote-debugging-port=9222", f"--user-data-dir={profile}",
            "--no-first-run", "--no-default-browser-check",
            # Temu 页面 JS 堆随运行时间增长，默认 ~2GB 易撞顶报 Out of Memory；
            # 4GB 是 V8 指针压缩模式的上限
            "--js-flags=--max-old-space-size=4096"]
    try:
        # DETACHED_PROCESS(0x8) + CREATE_NEW_PROCESS_GROUP(0x200)：脱离 bridge 独立存活
        subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                         creationflags=0x00000008 | 0x00000200)
    except Exception as e:
        raise RuntimeError(f"启动 Edge 失败：{e}")
    deadline = _t.time() + timeout
    while _t.time() < deadline:
        try:
            with urllib.request.urlopen("http://127.0.0.1:9222/json/version", timeout=2) as r:
                if r.status == 200:
                    break
        except Exception:
            pass
        _t.sleep(0.5)
    else:
        raise RuntimeError("Edge 已启动，但 CDP 端口 9222 在限定时间内未就绪，请检查是否弹出 Edge 窗口。")
    log.append("✅ Edge 已启动（端口 9222），CDP 就绪")
    return p.chromium.connect_over_cdp("http://127.0.0.1:9222"), True


@app.route('/api/order_price/enter', methods=['POST'])
def api_order_price_enter():
    """连接本机 Edge（CDP 9222），确保打开发价格申报视角标签页，
    自动点击「待卖家确认」标签并将每页条数设为 200。

    注意：只用 p.stop() 断开 CDP 连接，绝不调用 browser.close()，
    否则会把用户正在使用的 Edge 整体关掉。
    """
    log = []
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return jsonify({"ok": False, "error": "Bridge 环境未安装 playwright，无法控制 Edge。", "log": log})

    TARGET = "agentseller.temu.com/main/adjust-price-manage/order-price"
    p = sync_playwright().start()
    try:
        try:
            browser, edge_auto = _ensure_edge_cdp(p, log)
        except Exception as e:
            return jsonify({"ok": False,
                            "error": str(e),
                            "detail": "若 Edge 已开但端口 9222 未启用，请先关闭所有 Edge 再用「start-edge-cdp.bat」启动。",
                            "log": log})

        # 找已有标签页，没有就新开一个
        page = None
        for ctx in browser.contexts:
            for pg in ctx.pages:
                if TARGET in pg.url:
                    page = pg
                    break
            if page:
                break
        opened_new = False
        if not page:
            try:
                ctx0 = browser.contexts[0] if browser.contexts else browser.new_context()
                page = ctx0.new_page()
                page.goto("https://" + TARGET, timeout=30000)
                opened_new = True
                log.append("🌐 已新开「价格申报视角」标签页并导航")
            except Exception as e:
                return jsonify({"ok": False, "error": f"无法打开价格申报视角页面：{e}", "log": log})

        # 等标签栏/TAB 容器就绪
        try:
            page.wait_for_selector("div[class*='TAB_outerWrapper']", timeout=20000)
            log.append("✅ 页面已加载（识别到标签栏）")
        except Exception:
            log.append("⚠️ 未等到标签栏，仍尝试继续操作")
        page.wait_for_timeout(1500)

        # ① 点击「待卖家确认」
        try:
            tab = page.locator("div[class*='TAB_tabItem']", has_text="待卖家确认").first
            if tab.count() > 0:
                cls = tab.get_attribute("class") or ""
                if "TAB_active" in cls:
                    log.append("ℹ️ 「待卖家确认」已处于选中状态，跳过点击")
                else:
                    tab.click(timeout=6000)
                    log.append("✅ 已点击「待卖家确认」")
                page.wait_for_timeout(1500)
            else:
                log.append("⚠️ 未找到「待卖家确认」标签（页面结构可能变化）")
        except Exception as e:
            log.append(f"⚠️ 点击「待卖家确认」失败：{e}")

        # ② 设置每页 200 条
        try:
            size_select = page.locator("li[class*='PGT_sizeChanger'] div[class*='PGT_sizeSelect']").first
            if size_select.count() > 0:
                cur = (size_select.inner_text() or "").replace("\n", " ").strip()
                if "200" in cur:
                    log.append("ℹ️ 每页已是 200 条，跳过设置")
                else:
                    size_select.click(timeout=6000)
                    page.wait_for_timeout(800)
                    opt = page.locator("li[class*='cIL_item']", has_text="200").first
                    if opt.count() > 0:
                        opt.click(timeout=6000)
                        log.append("✅ 已将每页条数设置为 200")
                    else:
                        log.append("⚠️ 未找到「200」选项，请手动选择")
                    page.wait_for_timeout(1200)
            else:
                log.append("⚠️ 未找到每页条数选择器（页面结构可能变化）")
        except Exception as e:
            log.append(f"⚠️ 设置每页条数失败：{e}")

        try:
            url = page.url
        except Exception:
            url = ""
        return jsonify({"ok": True, "log": log, "url": url, "opened_new": opened_new})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "log": log})
    finally:
        try:
            p.stop()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────
# Temu 价格申报视角：按核价底价自动批量确认/拒绝
# ─────────────────────────────────────────────────────────────
# 核价底价（站点 -> 核价下限）。权威来源：Temu 核价仓 config/prices.py 的 PRICE_MAP（17站）。
ORDER_PRICE_FLOOR = {
    "波兰": 52, "匈牙利": 56, "立陶宛": 56, "德国": 63, "捷克": 65,
    "斯洛伐克": 67, "葡萄牙": 76, "西班牙": 80, "比利时": 80, "法国": 70,
    "丹麦": 76, "斯洛文尼亚": 84, "奥地利": 75, "荷兰": 73, "罗马尼亚": 93,
    "瑞典": 134, "芬兰": 142,
}

# 注入到页面上下文的 JS 助手（window._op_*）。表格用嵌套结构，真实数据行靠
# 「同时含 调整/不调整 链接」来识别；列索引按表头文本动态定位，扛结构微调。
ORDER_PRICE_JS = r"""
window._op_main=function(){
  const tables=[...document.querySelectorAll('table')];
  return tables.find(t=>t.querySelector('thead') && /单号/.test(t.querySelector('thead').innerText));
};
window._op_cols=function(){
  const main=window._op_main(); if(!main) return null;
  const ths=[...main.querySelectorAll('thead th')].map(th=>th.innerText.replace(/\n/g,' ').trim());
  const find=(kw)=>ths.findIndex(t=>t.includes(kw));
  return {site:find('站点'), adj:find('调整后申报价格'), order:find('单号')};
};
window._op_rows=function(){
  const main=window._op_main(); if(!main) return [];
  const cols=window._op_cols(); if(!cols) return [];
  const rows=[];
  for(const tr of main.querySelectorAll('tr')){
    const links=[...tr.querySelectorAll('a')].map(a=>a.innerText.trim());
    if(links.indexOf('调整')>=0 && links.indexOf('不调整')>=0){
      const tds=[...tr.querySelectorAll('td')].map(td=>td.innerText.replace(/\n/g,' ').trim());
      rows.push({order: cols.order>=0?tds[cols.order]:'', site: cols.site>=0?tds[cols.site]:'', adj: cols.adj>=0?tds[cols.adj]:''});
    }
  }
  return rows;
};
window._op_pager=function(){
  const sizeEl=document.querySelector("li[class*='PGT_sizeChanger']");
  const sizeText=sizeEl?sizeEl.innerText.replace(/\n/g,' ').trim():'';
  const pageItems=[...document.querySelectorAll('li[class*="PGT_pagerItem"]')].map(li=>li.innerText.trim()).filter(t=>/^\d+$/.test(t));
  const act=document.querySelector("li[class*='PGT_pagerItemActive']");
  const activePage=act?act.innerText.trim():'';
  return {sizeText, pageItems, activePage};
};
window._op_click=function(order, action){
  const main=window._op_main(); if(!main) return false;
  const cols=window._op_cols(); if(!cols) return false;
  for(const tr of main.querySelectorAll('tr')){
    const tds=[...tr.querySelectorAll('td')].map(td=>td.innerText.replace(/\n/g,' ').trim());
    const o = cols.order>=0 ? tds[cols.order] : '';
    if(o && o.indexOf(order)>=0){
      for(const a of tr.querySelectorAll('a')){
        if(a.innerText.trim()===action){ a.click(); return true; }
      }
    }
  }
  return false;
};
window._op_confirm_modal=function(){
  // 只找「可见」且含「确认」按钮的 弹窗容器(modal/dialog)，不断覆盖取最后一个(=最顶层)。
  // 避免点到被堆叠的多个同名弹窗里的隐藏/错误那一个。
  const modals=[...document.querySelectorAll("div[class*='modal'], div[class*='Modal'], div[class*='dialog'], div[class*='Dialog']")];
  let target=null;
  for(const m of modals){
    const cs=getComputedStyle(m);
    const r=m.getBoundingClientRect();
    if(cs.display==='none'||cs.visibility==='hidden'||r.width===0||r.height===0) continue;
    const hasConfirm=[...m.querySelectorAll('button,a')].some(x=>{const t=x.innerText.trim();return t==='确认'||t==='确认调整';});
    if(hasConfirm) target=m;   // 最后一个可见且含确认的 → 最顶层
  }
  if(!target) return false;
  const btns=[...target.querySelectorAll('button,a')].filter(x=>{const t=x.innerText.trim();return t==='确认'||t==='确认调整';});
  if(btns.length){ btns[btns.length-1].click(); return true; }
  return false;
};
window._op_modal_count=function(){
  const modals=[...document.querySelectorAll("div[class*='modal'], div[class*='Modal'], div[class*='dialog'], div[class*='Dialog']")];
  return modals.filter(m=>{const cs=getComputedStyle(m);const r=m.getBoundingClientRect();return cs.display!=='none'&&cs.visibility!=='hidden'&&r.width>0&&r.height>0;}).length;
};
window._op_dismiss_cancel=function(){
  // 点最顶层可见弹窗的「取消」，兜底清理关不掉的确认框（避免叠加）。
  const modals=[...document.querySelectorAll("div[class*='modal'], div[class*='Modal'], div[class*='dialog'], div[class*='Dialog']")];
  let target=null;
  for(const m of modals){
    const cs=getComputedStyle(m); const r=m.getBoundingClientRect();
    if(cs.display==='none'||cs.visibility==='hidden'||r.width===0||r.height===0) continue;
    const hasCancel=[...m.querySelectorAll('button,a')].some(x=>x.innerText.trim()==='取消');
    if(hasCancel) target=m;
  }
  if(!target) return false;
  const cb=[...target.querySelectorAll('button,a')].filter(x=>x.innerText.trim()==='取消');
  if(cb.length){ cb[cb.length-1].click(); return true; }
  return false;
};
// ── 批量拒绝相关助手 ──
window._op_check=function(order, want){
  // 按单号找到行，勾选/取消其勾选框（只动这一行，绝不全选）。
  const main=window._op_main(); if(!main) return false;
  const cols=window._op_cols(); if(!cols) return false;
  for(const tr of main.querySelectorAll('tr')){
    const tds=[...tr.querySelectorAll('td')].map(td=>td.innerText.replace(/\n/g,' ').trim());
    const o = cols.order>=0 ? tds[cols.order] : '';
    if(o && o.indexOf(order)>=0){
      const cell = tr.querySelector("td[class*='TB_checkCell']");
      if(!cell) return false;
      const inp = cell.querySelector("input");
      const lab = cell.querySelector("label");
      const cur = inp ? inp.checked : false;
      if(cur !== want){
        if(lab){ lab.click(); }            // 点可见 label，触发 React onChange
        else if(inp){ inp.click(); }
      }
      return true;
    }
  }
  return false;
};
window._op_checked_count=function(){
  const main=window._op_main(); if(!main) return 0;
  let n=0;
  for(const tr of main.querySelectorAll('tr')){
    const cell = tr.querySelector("td[class*='TB_checkCell']");
    if(cell){ const inp=cell.querySelector('input'); if(inp && inp.checked) n++; }
  }
  return n;
};
window._op_click_batch_reject=function(){
  // 点操作栏「批量拒绝」按钮（排除已禁用/灰的，以及面板内的「拒绝」）。
  const els=[...document.querySelectorAll('button,a,div,span')].filter(x=>x.innerText.trim()==='批量拒绝');
  for(const e of els){
    const cs=getComputedStyle(e); const r=e.getBoundingClientRect();
    if(cs.display==='none'||cs.visibility==='hidden'||r.width===0||r.height===0) continue;
    if(e.disabled) continue;
    e.click(); return true;
  }
  return false;
};
window._op_panel_visible=function(){
  return !!document.querySelector("div[class*='TB_innerRight']");
};
window._op_reason_fields=function(){
  // 只收集「批量拒绝」右侧面板 TB_innerRight 内的真正原因框（textarea）。
  // 不跨 MDL_ 弹窗收集：最终确认弹窗的 DOM 多层嵌套同名 class，
  // querySelectorAll 会在每层父级重复匹配同一批 textarea，导致计数虚高、误判「未填满」。
  const panel=document.querySelector("div[class*='TB_innerRight']");
  if(!panel) return [];
  const tas=[...panel.querySelectorAll('textarea')].filter(el=>el.offsetParent!==null);
  const inp=[...panel.querySelectorAll('input')].filter(el=>{
    const t=(el.type||'').toLowerCase();
    const ph=(el.placeholder||'');
    return t!=='checkbox'&&t!=='radio'&&t!=='hidden'&&t!=='button'&&t!=='submit' && (ph.includes('原因')||ph.includes('请输入'));
  }).filter(el=>el.offsetParent!==null);
  return [...tas, ...inp];
};
window._op_fill_first_reason=function(text){
  const f=window._op_reason_fields();
  if(!f.length) return false;
  const el=f[0];
  const proto = el.tagName==='TEXTAREA' ? window.HTMLTextAreaElement.prototype : window.HTMLInputElement.prototype;
  const setter=Object.getOwnPropertyDescriptor(proto,'value').set;
  // React 受控组件兼容：先重置 valueTracker，使合成 onChange 真正触发（否则值写不进 state）。
  try { if(el._valueTracker) el._valueTracker.setValue(''); } catch(e){}
  setter.call(el, text);
  el.dispatchEvent(new Event('input',{bubbles:true}));
  el.dispatchEvent(new Event('change',{bubbles:true}));
  return true;
};
window._op_all_reason_filled=function(){
  const f=window._op_reason_fields();
  const empty=f.filter(el=>!(el.value||'').trim()).length;
  return {total:f.length, empty:empty, all: empty===0};
};
window._op_click_reject_final=function(){
  // 点最终确认弹窗（含「拒绝调价」字样，如「已选 N 个调价单，拒绝调价？」）里的「拒绝」提交按钮。
  // 注意：该弹窗的提交按钮文本是「拒绝」而非「确认」。
  const modals=[...document.querySelectorAll("div[class*='MDL_'], div[class*='modal'], div[class*='Modal']")];
  let target=null;
  for(const m of modals){
    const cs=getComputedStyle(m); const r=m.getBoundingClientRect();
    if(cs.display==='none'||cs.visibility==='hidden'||r.width===0||r.height===0) continue;
    if((m.innerText||'').includes('拒绝调价')) target=m;
  }
  if(!target) return false;
  const btns=[...target.querySelectorAll('button,a')].filter(b=>{
    const t=b.innerText.trim(); const cs=getComputedStyle(b); const r=b.getBoundingClientRect();
    return t==='拒绝' && cs.display!=='none' && cs.visibility!=='hidden' && r.width>0 && r.height>0 && !b.disabled;
  });
  if(btns.length){ btns[btns.length-1].click(); return true; }
  return false;
};
window._op_final_modal_present=function(){
  // 检测最终确认弹窗（含「拒绝调价」字样）是否可见。该弹窗为 MDL_ 类，_op_modal_count 不识别，故单独检测。
  const modals=[...document.querySelectorAll("div[class*='MDL_'], div[class*='modal'], div[class*='Modal']")];
  for(const m of modals){
    const cs=getComputedStyle(m); const r=m.getBoundingClientRect();
    if(cs.display==='none'||cs.visibility==='hidden'||r.width===0||r.height===0) continue;
    if((m.innerText||'').includes('拒绝调价')) return true;
  }
  return false;
};
window._op_copy_ready_in_first_block=function(){
  const fields=window._op_reason_fields(); if(!fields.length) return false;
  let node=fields[0].parentElement;
  while(node){
    const cp=[...node.querySelectorAll('a,span,button,div')].find(x=>x.innerText.trim()==='一键复制');
    if(cp){ const cs=getComputedStyle(cp); const r=cp.getBoundingClientRect();
      return cs.display!=='none'&&cs.visibility!=='hidden'&&r.width>0&&r.height>0 && !cp.disabled; }
    node=node.parentElement;
  }
  return false;
};
window._op_click_copy_in_first_block=function(){
  const fields=window._op_reason_fields(); if(!fields.length) return false;
  let node=fields[0].parentElement;
  while(node){
    const cp=[...node.querySelectorAll('a,span,button,div')].find(x=>x.innerText.trim()==='一键复制');
    if(cp){ const cs=getComputedStyle(cp); const r=cp.getBoundingClientRect();
      if(cs.display!=='none'&&cs.visibility!=='hidden'&&r.width>0&&r.height>0 && !cp.disabled){ cp.click(); return true; } }
    node=node.parentElement;
  }
  return false;
};
window._op_click_reject=function(){
  // 点可见的「拒绝」提交按钮（排除「批量拒绝」）。该按钮位于面板 TB_innerRight 之外，
  // 故在 document 范围查找；此时最终确认弹窗尚未出现，不会误点。
  const els=[...document.querySelectorAll('button,a')].filter(x=>{
    const t=x.innerText.trim();
    return (t==='拒绝'||t.startsWith('拒绝')) && !t.startsWith('批量拒绝');
  });
  for(const e of els){
    const cs=getComputedStyle(e); const r=e.getBoundingClientRect();
    if(cs.display!=='none'&&cs.visibility!=='hidden'&&r.width>0&&r.height>0){ e.click(); return true; }
  }
  return false;
};
"""

TARGET_ORDER_PRICE = "agentseller.temu.com/main/adjust-price-manage/order-price"


def _op_open_tab(browser, log):
    """找/开「价格申报视角」标签页，返回 page。"""
    page = None
    for ctx in browser.contexts:
        for pg in ctx.pages:
            if TARGET_ORDER_PRICE in pg.url:
                page = pg
                break
        if page:
            break
    if not page:
        try:
            ctx0 = browser.contexts[0] if browser.contexts else browser.new_context()
            page = ctx0.new_page()
            page.goto("https://" + TARGET_ORDER_PRICE, timeout=30000)
            log.append("🌐 已新开「价格申报视角」标签页并导航")
        except Exception as e:
            raise RuntimeError(f"无法打开价格申报视角页面：{e}")
    return page


def _op_setup(page, log):
    """① 点「待卖家确认」② 每页 200 条 ③ 注入 JS 助手。"""
    try:
        page.wait_for_selector("div[class*='TAB_outerWrapper']", timeout=20000)
        log.append("✅ 页面已加载（识别到标签栏）")
    except Exception:
        log.append("⚠️ 未等到标签栏，仍尝试继续操作")
    page.wait_for_timeout(1500)
    # ① 待卖家确认
    try:
        tab = page.locator("div[class*='TAB_tabItem']", has_text="待卖家确认").first
        if tab.count() > 0:
            cls = tab.get_attribute("class") or ""
            if "TAB_active" in cls:
                log.append("ℹ️ 「待卖家确认」已选中，跳过点击")
            else:
                tab.click(timeout=6000)
                log.append("✅ 已点击「待卖家确认」")
            page.wait_for_timeout(1200)
        else:
            log.append("⚠️ 未找到「待卖家确认」标签")
    except Exception as e:
        log.append(f"⚠️ 点击「待卖家确认」失败：{e}")
    # ② 每页 200
    try:
        size_select = page.locator("li[class*='PGT_sizeChanger'] div[class*='PGT_sizeSelect']").first
        if size_select.count() > 0:
            cur = (size_select.inner_text() or "").replace("\n", " ").strip()
            if "200" in cur:
                log.append("ℹ️ 每页已是 200 条，跳过")
            else:
                size_select.click(timeout=6000)
                page.wait_for_timeout(800)
                opt = page.locator("li[class*='cIL_item']", has_text="200").first
                if opt.count() > 0:
                    opt.click(timeout=6000)
                    log.append("✅ 每页条数设为 200")
                else:
                    log.append("⚠️ 未找到 200 选项")
                page.wait_for_timeout(1200)
        else:
            log.append("⚠️ 未找到每页条数选择器")
    except Exception as e:
        log.append(f"⚠️ 设置每页条数失败：{e}")
    # ③ 注入 JS 助手
    page.evaluate(ORDER_PRICE_JS)


def _parse_price(s):
    """解析价格文本为 float，正确处理欧洲格式：
    - 逗号小数（如 '60,50' → 60.5）、点小数（'60.50' → 60.5）
    - 千分位（'1.234,56' / '1,234.56' / '1,234'）
    """
    import re
    s = re.sub(r'[^0-9.,]', '', s or '')
    if not s:
        return None
    if '.' in s and ',' in s:
        if s.rfind('.') > s.rfind(','):
            s = s.replace(',', '')          # 点是小数，逗号是千分位
        else:
            s = s.replace('.', '').replace(',', '.')  # 逗号是小数，点是千分位
    elif ',' in s:
        if re.fullmatch(r'\d{1,3}(,\d{3})+', s):
            s = s.replace(',', '')          # 纯千分位分组
        else:
            s = s.replace(',', '.')         # 逗号作小数位
    try:
        return float(s)
    except Exception:
        return None


def _op_decide(row):
    """按核价底价决定 accept/reject/skip。返回 (决策, 原因, 底价, 解析后的价格)。"""
    site = row.get('site', '') or ''
    adj_s = row.get('adj', '') or ''
    adj = _parse_price(adj_s)
    if adj is None:
        return ('skip', '价格无法解析', None, None)
    matched = None
    for k, v in ORDER_PRICE_FLOOR.items():
        if k in site:
            matched = v
            break
    if matched is None:
        return ('skip', '站点未配置(留人工)', None, adj)
    if adj >= matched:
        return ('accept', '建议价≥底价', matched, adj)
    return ('reject', '建议价<底价', matched, adj)


def _op_click_next_page(page, seen):
    """点击比当前激活页更大的下一页，返回是否成功点击。"""
    info = page.evaluate("()=>({pager:window._op_pager()})")
    pager = info['pager']
    active = pager['activePage']
    nums = sorted(int(x) for x in pager['pageItems'] if str(x).isdigit())
    nexts = [n for n in nums if (not str(active).isdigit() or n > int(active))]
    if not nexts:
        return False
    target = min(nexts)
    if str(target) in seen:
        return False
    clicked = page.evaluate(
        """(n)=>{const lis=[...document.querySelectorAll('li[class*="PGT_pagerItem"]')];"""
        """const li=lis.find(l=>l.innerText.trim()===String(n));if(li){li.click();return true;}return false;}""",
        target)
    if clicked:
        seen.add(str(target))
        page.wait_for_timeout(1800)
        return True
    return False


def _op_scan(page, log):
    """扫描所有页，返回去重后的行列表。"""
    rows = []
    seen_pages = set()
    for _ in range(60):
        info = page.evaluate("()=>({rows:window._op_rows(), pager:window._op_pager()})")
        rows.extend(info['rows'])
        if not _op_click_next_page(page, seen_pages):
            break
    # 按 order 去重（翻页时可能重复读到同一行）
    seen_o, uniq = set(), []
    for r in rows:
        if r.get('order') and r['order'] in seen_o:
            continue
        seen_o.add(r.get('order'))
        uniq.append(r)
    return uniq


def _op_dismiss_cancel(page, baseline, max_tries=12):
    """兜底清理：若仍有未关闭的确认弹窗（弹窗数>基线），点最顶层「取消」直到回到基线。"""
    for _ in range(max_tries):
        if page.evaluate("()=>window._op_modal_count()") <= baseline:
            return True
        if not page.evaluate("()=>window._op_dismiss_cancel()"):
            return False
        page.wait_for_timeout(400)
    return page.evaluate("()=>window._op_modal_count()") <= baseline


def _op_auto(page, log, passed):
    """自动处理：只对「建议价≥底价」(accept) 的行，按正确顺序逐条处理：
    点「调整」→ 弹出确认框 → 点「确认」(关掉这一个) → 该条完成 → 再处理下一条。
    「建议价<底价」(reject) 的行不点击，保持原样（留待人工）。
    已点过「调整」的订单记入 attempted，下一轮扫描不再重复点击，避免确认弹窗叠加。
    返回 (accept, skipped, err)。"""
    def modal_count():
        return page.evaluate("()=>window._op_modal_count()")
    # 基线：开工前页面已挂着的确认弹窗数。若已堆着卡住的弹窗(如上次未关闭)，
    # 遮罩会挡住表格点击，必须先在 Edge 刷新本页面清除，否则无法正常工作。
    baseline = modal_count()
    if baseline > 3:
        raise RuntimeError(
            f"检测到页面已挂着 {baseline} 个确认弹窗（疑似上次未关闭而堆积）。"
            f"请先在 Edge 里刷新本页面（按 F5）清除这些卡住的弹窗，再执行自动处理。")

    accept = skipped = err = 0
    attempted = set()  # 已点过「调整」的订单，避免重复点击导致确认弹窗叠加
    rejected_seen = set()  # 已计入 skipped 的拒绝订单，避免跨轮扫描重复计数
    seen_pages = set()
    for _ in range(80):
        info = page.evaluate("()=>({rows:window._op_rows(), pager:window._op_pager()})")
        rows = info['rows']
        # 本页去重，避免重复点击同一行
        seen_o, uniq = set(), []
        for r in rows:
            if r.get('order') and r['order'] in seen_o:
                continue
            seen_o.add(r.get('order'))
            uniq.append(r)
        decisions = [(r, _op_decide(r)) for r in uniq]
        # 本页「拒绝(不调整)」不点击，仅累计计数（按订单去重，避免跨轮扫描重复计数）
        for r, d in decisions:
            if d[0] == 'reject' and r.get('order') not in rejected_seen:
                rejected_seen.add(r.get('order'))
                skipped += 1
        # 只处理 accept（≥底价）→ 点「调整」并确认；已尝试过的跳过（防弹窗叠加）
        config = [(r, d) for r, d in decisions
                  if d[0] == 'accept' and (r.get('order') not in attempted)]
        if not config:
            if not _op_click_next_page(page, seen_pages):
                break
            continue
        for r, d in config:
            attempted.add(r.get('order'))  # 标记已尝试，下一轮不再点，杜绝反复点击
            ok = page.evaluate("""(a)=>window._op_click(a.order, a.action)""",
                               {"order": r.get('order'), "action": "调整"})
            if not ok:
                err += 1
                log.append(f"⚠️ 未找到 {r.get('order')} 的「调整」按钮")
                continue
            # 点「调整」后必弹出确认框；轮询：等确认框出现(弹窗数>基线) → 点最顶层「确认」→ 等其关闭。
            confirmed = False
            for _k in range(20):
                mc = modal_count()
                if mc > baseline:
                    if page.evaluate("()=>window._op_confirm_modal()"):
                        confirmed = True
                    page.wait_for_timeout(500)
                    if modal_count() <= baseline:  # 已关闭 → 这一条完成
                        break
                else:
                    page.wait_for_timeout(300)
            if not confirmed:
                log.append(f"⚠️ {r.get('order')}({r.get('site')}) 点击调整后未弹出确认框，可能需人工处理")
                err += 1
                # 兜底：若仍有弹窗未关，点「取消」清理，避免叠加
                _op_dismiss_cancel(page, baseline)
                continue
            page.wait_for_timeout(600)
            still = page.evaluate("""(o)=>window._op_rows().some(r=>r.order===o)""", r.get('order'))
            if still:
                err += 1
                log.append(f"⚠️ {r.get('order')}({r.get('site')}) 确认后行未消失，可能需人工确认")
            else:
                accept += 1
                passed.append({"order": r.get('order'), "site": r.get('site'),
                               "price": r.get('adj'), "floor": d[2]})
                # 实时：每通过一条立刻记录价格（前端会逐条显示，不等待整体结束）
                log.append(f"✅ 通过(已调整) 订单{r.get('order')} ｜ {r.get('site')} ｜ 通过价格 {r.get('adj')}（≥核价底价 {d[2]}）")
    return accept, skipped, err


def _op_reject(page, log, rejected):
    """批量拒绝所有「建议价<底价」(reject) 的行：
    逐个勾选这些行(绝不点全选) → 点「批量拒绝」→ 右侧面板填原因「价格过低」
    → 点首行「一键复制」→ 确认复制弹窗 → 点面板内「拒绝」(真实改数据)。
    只勾选 reject 行，accept/skip 行保持未勾选，绝不误拒。
    返回 (rejected_count, err)。"""
    baseline = page.evaluate("()=>window._op_modal_count()")
    if baseline > 3:
        raise RuntimeError(
            f"检测到页面已挂着 {baseline} 个确认弹窗（疑似上次未关闭而堆积）。"
            f"请先在 Edge 里刷新本页面（按 F5）清除这些卡住的弹窗，再执行批量拒绝。")

    # 1) 扫描本页所有行，找出 reject 行，逐个勾选（按单号精确匹配，不点全选）
    rows = page.evaluate("()=>window._op_rows()")
    seen_o, uniq = set(), []
    for r in rows:
        if r.get('order') and r['order'] in seen_o:
            continue
        seen_o.add(r.get('order'))
        uniq.append(r)
    targets = [r for r in uniq if _op_decide(r)[0] == 'reject']
    if not targets:
        log.append("ℹ️ 当前页没有需要拒绝（低于底价）的订单")
        return 0, 0
    log.append(f"📋 计划逐个勾选并拒绝 {len(targets)} 个低于底价的订单（不点全选）")

    err = 0
    for r in targets:
        ok = page.evaluate("""(a)=>window._op_check(a.order, true)""", {"order": r.get('order')})
        if not ok:
            log.append(f"⚠️ 未找到 {r.get('order')} 的勾选框，跳过")
            err += 1
            continue
        d = _op_decide(r)
        rejected.append({"order": r.get('order'), "site": r.get('site'),
                         "price": r.get('adj'), "floor": d[2]})
        log.append(f"☑️ 已勾选 订单{r.get('order')} ｜ {r.get('site')} ｜ 价格 {r.get('adj')}（<核价底价 {d[2]}）")
        page.wait_for_timeout(120)

    n = page.evaluate("()=>window._op_checked_count()")
    log.append(f"🔢 实际已勾选 {n} 行（预期 {len(targets)}）")
    if n == 0:
        log.append("⚠️ 没有任何行被勾选，中止（避免误点批量拒绝按钮）")
        return 0, err

    # 2) 点「批量拒绝」→ 右侧面板
    if not page.evaluate("()=>window._op_click_batch_reject()"):
        log.append("⚠️ 未找到/无法点击「批量拒绝」按钮（可能未勾选或按钮不可点）")
        return 0, err
    log.append("✅ 已点击「批量拒绝」")
    try:
        page.wait_for_selector("div[class*='TB_innerRight']", timeout=8000)
        log.append("✅ 批量拒绝面板已打开")
    except Exception:
        log.append("⚠️ 未检测到批量拒绝面板，可能批量拒绝按钮未生效")
        return len(targets), err

    # 3)+4) 用 Playwright 真实填充每个原因框（.fill 模拟真实输入，React 受控组件必更新），
    #        再轮询验证全部填满（不依赖面板「一键复制」，因其对面板 textarea 不生效）。
    all_ok = False
    for attempt in range(6):
        try:
            tas = page.locator("div[class*='TB_innerRight'] textarea")
            n = tas.count()
            for i in range(n):
                tas.nth(i).fill("价格过低")
                page.wait_for_timeout(120)
        except Exception as e:
            log.append(f"⚠️ 填充原因框异常: {e}")
        page.wait_for_timeout(400)
        res = page.evaluate("()=>window._op_all_reason_filled()")
        if res and res.get('all'):
            all_ok = True
            log.append(f"✅ 所有 {res.get('total')} 个原因框已填「价格过低」")
            break
        else:
            tot = (res or {}).get('total'); empt = (res or {}).get('empty')
            log.append(f"⚠️ 原因框未全填满（空 {empt}/{tot}），重试（第 {attempt+1} 次）")
            page.wait_for_timeout(400)
    if not all_ok:
        log.append("⚠️ 多次重试仍未能填满所有原因框，继续尝试点拒绝（可能失败）")

    # 5) 点面板外「拒绝」提交按钮 → 弹出最终确认弹窗（真实改数据）
    clicked = False
    for _k in range(15):
        if page.evaluate("()=>window._op_click_reject()"):
            page.wait_for_timeout(900)
            if page.evaluate("()=>window._op_final_modal_present()"):
                log.append("✅ 已点击「拒绝」，弹出最终确认弹窗")
                clicked = True
                break
            else:
                log.append("⚠️ 点了「拒绝」但未弹出最终弹窗，重试")
        page.wait_for_timeout(300)
    if not clicked:
        log.append("⚠️ 未成功点击「拒绝」或弹出最终弹窗，可能需人工点击")
        return len(targets), err

    # 6) 点最终确认弹窗（「已选 N 个调价单，拒绝调价？」）里的「拒绝」提交按钮
    page.wait_for_timeout(800)
    for _k in range(25):
        if page.evaluate("()=>window._op_final_modal_present()"):
            if page.evaluate("()=>window._op_click_reject_final()"):
                log.append("✅ 已点击最终弹窗「拒绝」")
            else:
                page.evaluate("()=>window._op_confirm_modal()")
                log.append("✅ 已确认最终拒绝弹窗（确认按钮）")
            page.wait_for_timeout(800)
            if not page.evaluate("()=>window._op_final_modal_present()"):
                break
        else:
            page.wait_for_timeout(300)

    # 7) 验证：被拒订单是否从列表消失
    page.wait_for_timeout(1500)
    after = page.evaluate("()=>window._op_rows()")
    after_orders = {r.get('order') for r in after}
    still = [r for r in targets if r.get('order') in after_orders]
    if still:
        log.append(f"⚠️ 仍有 {len(still)} 个计划拒绝的订单未从列表消失，可能需人工确认（如最终确认弹窗被遮挡）")
    else:
        log.append(f"✅ 全部 {len(targets)} 个订单已成功拒绝并从列表消失")
    return len(targets), err


@app.route('/api/order_price/scan', methods=['POST'])
def api_order_price_scan():
    """只读扫描：按核价底价给出 接受/拒绝/跳过 预览（不点任何按钮）。"""
    log = []
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return jsonify({"ok": False, "error": "Bridge 环境未安装 playwright，无法控制 Edge。", "log": log})
    p = sync_playwright().start()
    try:
        try:
            browser, _ = _ensure_edge_cdp(p, log)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e), "log": log})
        try:
            page = _op_open_tab(browser, log)
            _op_setup(page, log)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e), "log": log})
        if not page.evaluate("()=>!!window._op_main()"):
            return jsonify({"ok": False,
                            "error": "未找到价格申报表格，可能未登录 Temu 或页面未加载。请先在 Edge 登录 Temu 卖家后台，再执行。",
                            "log": log})
        rows = _op_scan(page, log)
        accept = reject = skip = 0
        by_site = {}
        for r in rows:
            d = _op_decide(r)
            decision, floor, adj_val = d[0], d[2], d[3]
            if decision == 'accept':
                accept += 1
            elif decision == 'reject':
                reject += 1
            else:
                skip += 1
            s = r.get('site') or '未知'
            bs = by_site.setdefault(
                s, {"accept": 0, "reject": 0, "skip": 0,
                    "floor": None, "accept_min": None, "reject_max": None})
            bs[decision] = bs.get(decision, 0) + 1
            # 记录该站点的核价底价（非 None 才覆盖，避免被未配置行清空）
            if floor is not None:
                bs["floor"] = floor
            if decision == 'accept' and adj_val is not None:
                if bs["accept_min"] is None or adj_val < bs["accept_min"]:
                    bs["accept_min"] = adj_val
            if decision == 'reject' and adj_val is not None:
                if bs["reject_max"] is None or adj_val > bs["reject_max"]:
                    bs["reject_max"] = adj_val
        return jsonify({"ok": True, "log": log, "total": len(rows),
                        "accept": accept, "reject": reject, "skip": skip,
                        "by_site": by_site, "floor": ORDER_PRICE_FLOOR, "rows": rows})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "log": log})
    finally:
        try:
            p.stop()
        except Exception:
            pass


# ===== 价格申报：自动处理（后台执行 + 前端轮询实时进度） =====
OP_TASKS = {}
OP_TASKS_LOCK = threading.Lock()


def _op_auto_run(task_id):
    """后台线程：连接 Edge → 进入页面 → 逐条自动确认/拒绝，进度实时写入 OP_TASKS[task_id]。"""
    task = OP_TASKS.get(task_id)
    if not task:
        return
    log = task["log"]
    passed = task["passed"]
    lock = task["lock"]
    try:
        from playwright.sync_api import sync_playwright
        p = sync_playwright().start()
        try:
            try:
                browser, _ = _ensure_edge_cdp(p, log)
            except Exception as e:
                with lock:
                    task["error"] = str(e)
                    task["done"] = True
                return
            try:
                page = _op_open_tab(browser, log)
                _op_setup(page, log)
            except Exception as e:
                with lock:
                    task["error"] = str(e)
                    task["done"] = True
                return
            if not page.evaluate("()=>!!window._op_main()"):
                with lock:
                    task["error"] = ("未找到价格申报表格，可能未登录 Temu 或页面未加载。"
                                     "请先在 Edge 登录 Temu 卖家后台，再执行。")
                    task["done"] = True
                return
            accept, skipped, err = _op_auto(page, log, passed)
            with lock:
                task["result"] = {"ok": True, "accept": accept, "skipped": skipped,
                                  "error_count": err,
                                  "note": "所有站点(含意大利，底价115)均按核价底价自动判定；未匹配到站点名的行留人工处理。"}
                task["done"] = True
        finally:
            try:
                p.stop()
            except Exception:
                pass
    except Exception as e:
        with lock:
            task["error"] = str(e)
            task["done"] = True


@app.route('/api/order_price/auto', methods=['POST'])
def api_order_price_auto():
    """启动后台自动批量处理；立即返回 task_id，前端轮询 /api/order_price/status 获取实时进度与每条价格。"""
    import uuid
    task_id = uuid.uuid4().hex
    with OP_TASKS_LOCK:
        OP_TASKS[task_id] = {"lock": threading.Lock(), "log": [], "passed": [], "rejected": [],
                             "done": False, "result": None, "error": None, "kind": "auto"}
    t = threading.Thread(target=_op_auto_run, args=(task_id,), daemon=True)
    t.start()
    return jsonify({"ok": True, "task_id": task_id})


@app.route('/api/order_price/status', methods=['GET'])
def api_order_price_status():
    """轮询任务进度：返回已产生的日志(实时)、通过价格清单、是否完成、最终结果/错误。"""
    task_id = request.args.get('task_id')
    if not task_id or task_id not in OP_TASKS:
        return jsonify({"ok": False, "error": "无效或已过期的任务ID"})
    task = OP_TASKS[task_id]
    with task["lock"]:
        return jsonify({"ok": True, "done": task["done"], "error": task["error"],
                        "result": task["result"], "kind": task.get("kind"),
                        "log": list(task["log"]),
                        "passed": list(task["passed"]),
                        "rejected": list(task["rejected"])})


def _op_reject_run(task_id):
    """后台线程：连接 Edge → 进入页面 → 逐个勾选低于底价的订单并批量拒绝，
    进度实时写入 OP_TASKS[task_id]。"""
    task = OP_TASKS.get(task_id)
    if not task:
        return
    log = task["log"]
    rejected = task["rejected"]
    lock = task["lock"]
    try:
        from playwright.sync_api import sync_playwright
        p = sync_playwright().start()
        try:
            try:
                browser, _ = _ensure_edge_cdp(p, log)
            except Exception as e:
                with lock:
                    task["error"] = str(e)
                    task["done"] = True
                return
            try:
                page = _op_open_tab(browser, log)
                _op_setup(page, log)
            except Exception as e:
                with lock:
                    task["error"] = str(e)
                    task["done"] = True
                return
            if not page.evaluate("()=>!!window._op_main()"):
                with lock:
                    task["error"] = ("未找到价格申报表格，可能未登录 Temu 或页面未加载。"
                                     "请先在 Edge 登录 Temu 卖家后台，再执行。")
                    task["done"] = True
                return
            cnt, err = _op_reject(page, log, rejected)
            with lock:
                task["result"] = {"ok": True, "rejected_count": cnt,
                                  "error_count": err,
                                  "note": "已拒绝所有低于核价底价的订单(含意大利，底价115)；未匹配到站点名的行留人工处理。"}
                task["done"] = True
        finally:
            try:
                p.stop()
            except Exception:
                pass
    except Exception as e:
        with lock:
            task["error"] = str(e)
            task["done"] = True


@app.route('/api/order_price/reject', methods=['POST'])
def api_order_price_reject():
    """启动后台批量拒绝（低于底价）任务；立即返回 task_id，前端轮询 /api/order_price/status。"""
    import uuid
    task_id = uuid.uuid4().hex
    with OP_TASKS_LOCK:
        OP_TASKS[task_id] = {"lock": threading.Lock(), "log": [], "passed": [], "rejected": [],
                             "done": False, "result": None, "error": None, "kind": "reject"}
    t = threading.Thread(target=_op_reject_run, args=(task_id,), daemon=True)
    t.start()
    return jsonify({"ok": True, "task_id": task_id})


# ===== Temu 流量加速器：开启流量加速（用户登录后点「好了」，脚本接管标签页） =====
TRAFFIC_TASKS = {}
TRAFFIC_TASKS_LOCK = threading.Lock()
TRAFFIC_SKU_MAP = {}  # skuId → goodsId(SPU)，响应拦截器从列表接口数据里建（submit higher custom price 报错只给 sku）
TRAFFIC_STOP = threading.Event()   # 全局停止信号：前端点「停止」即置位，循环安全退出
# 本轮（含历史记录）被 Temu 提交校验拒绝（submit higher custom price，连最少让价档都不行）的 SPU。
# 这些 SPU 会一直留在「待开启」列表里，每轮重扫都会再次被拒、白跑一整圈剔除流程，
# 所以分析阶段直接当不通过处理，不再进提交抽屉。任务启动时从记录文件预载。
TRAFFIC_REJECTED_SPUS = set()
# 会话级记录去重：垃圾商品（价格不通过/读不出价/未提交成功）永远留在「待开启」列表，
# 每轮重扫都会重复写 xlsx（实测 35 轮写了 374 条不通过记录，其实只有十几个 SPU）。
# 每个 (SPU,站点)/(SPU) 每次任务只写一条。任务启动时清空。
TRAFFIC_RECORDED_ONCE = set()

# 已知垃圾 (SPU,站点) 黑名单：价格不通过（低于底价10元+）/未读到申报价或档位/Temu拒绝。
# 这些行永远不离开「待开启」列表，每次扫到都重新分析一遍（实测 18 个垃圾 pair 产生了
# 1843 条重复记录）。落盘持久化 + 时间戳，7 天过期（防用户改申报价后误杀）。
# 必须按 (SPU,站点) 对拉黑：Temu 勾选是 SPU 粒度，多数垃圾 SPU 同时有可通过的站点行，
# 按 SPU 拉黑会误杀（2026-08-08 用户确认方案）。
TRAFFIC_BLACKLIST = {}  # {(spu, site): {"ts": iso字符串, "status": str}}
TRAFFIC_BLACKLIST_FILE = TEMU_ANALYSIS_DIR / "traffic_blacklist.json"
TRAFFIC_BLACKLIST_TTL_DAYS = 7


def _traffic_blacklist_alive(entry):
    try:
        ts = datetime.fromisoformat(entry.get("ts", ""))
        return datetime.now() - ts < timedelta(days=TRAFFIC_BLACKLIST_TTL_DAYS)
    except Exception:
        return False


def _traffic_blacklist_add(spu, site, status):
    """把 (SPU,站点) 加入黑名单并落盘（含过期清理）。写失败不中断主流程。"""
    TRAFFIC_BLACKLIST[(str(spu), str(site))] = {
        "ts": datetime.now().isoformat(timespec="seconds"), "status": status}
    try:
        data = {f"{k[0]}|{k[1]}": v for k, v in TRAFFIC_BLACKLIST.items()
                if _traffic_blacklist_alive(v)}
        TRAFFIC_BLACKLIST_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    except Exception:
        pass


def _traffic_load_blacklist():
    """预载黑名单：落盘文件（未过期）+ 记录文件里最终状态为垃圾且未过期的 (SPU,站点) 对。"""
    n = 0
    try:
        if TRAFFIC_BLACKLIST_FILE.exists():
            data = json.loads(TRAFFIC_BLACKLIST_FILE.read_text(encoding="utf-8"))
            for k, v in data.items():
                if _traffic_blacklist_alive(v):
                    spu, _, site = k.partition("|")
                    TRAFFIC_BLACKLIST[(spu, site)] = v
                    n += 1
    except Exception:
        pass
    try:
        from openpyxl import load_workbook
        if TRAFFIC_RECORD_FILE.exists():
            wb = load_workbook(TRAFFIC_RECORD_FILE, read_only=True)
            final = {}  # 同一 (SPU,站点) 最后一条记录为准（避免拉黑后来改价通过的）
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 9 and row[1]:
                    final[(str(row[1]).strip(), str(row[2] or "").strip())] = \
                        (str(row[0] or ""), str(row[8] or ""))
            wb.close()
            for (spu, site), (t, status) in final.items():
                if ("价格不通过" not in status and "Temu拒绝" not in status) \
                        or (spu, site) in TRAFFIC_BLACKLIST:
                    continue
                try:  # 记录时间超过 7 天的不再预载（价格可能已改）
                    if datetime.now() - datetime.strptime(t, "%Y-%m-%d %H:%M:%S") \
                            >= timedelta(days=TRAFFIC_BLACKLIST_TTL_DAYS):
                        continue
                except Exception:
                    pass
                TRAFFIC_BLACKLIST[(spu, site)] = {"ts": t.replace(" ", "T"), "status": status}
                n += 1
    except Exception:
        pass
    return n


def _traffic_record_once(rec, key):
    """同一次任务里同一个 key 只写一条记录（垃圾商品每轮重扫不重复写）。"""
    if key in TRAFFIC_RECORDED_ONCE:
        return
    TRAFFIC_RECORDED_ONCE.add(key)
    _traffic_record(rec)


def _traffic_load_rejected_spus():
    """从记录文件预载历史上被 Temu 拒绝（要求更高价格，已剔除）的 SPU。"""
    try:
        from openpyxl import load_workbook
        if not TRAFFIC_RECORD_FILE.exists():
            return
        wb = load_workbook(TRAFFIC_RECORD_FILE, read_only=True)
        ws = wb.active
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 9 and row[1] and "Temu拒绝" in str(row[8] or ""):
                TRAFFIC_REJECTED_SPUS.add(str(row[1]).strip())
                n += 1
        wb.close()
        return n
    except Exception:
        return None


class _TrafficStopped(Exception):
    """用户点了停止按钮，安全中断当前流量加速任务。"""


def _traffic_check_stop():
    if TRAFFIC_STOP.is_set():
        raise _TrafficStopped()

# 流量加速器页面 URL 特征（小写子串匹配）。用户手动打开页面后按特征找标签页。
# 真实页面：https://agentseller-eu.temu.com/main/flux-analysis（2026-08-06 用户提供）
TRAFFIC_URL_HINTS = ["flux-analysis", "flux"]
TRAFFIC_HOME = "https://agentseller-eu.temu.com/main/flux-analysis"
# 处理记录文件（xlsx，避免 CSV 编码/分列问题）
TRAFFIC_RECORD_FILE = TEMU_ANALYSIS_DIR / "流量加速器记录.xlsx"


def _traffic_record(rec):
    """把一行处理结果追加到 xlsx 记录文件。写失败不中断主流程。"""
    try:
        from openpyxl import Workbook, load_workbook
        if TRAFFIC_RECORD_FILE.exists():
            wb = load_workbook(TRAFFIC_RECORD_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["时间", "SPU", "站点", "申报价(CNY)", "核价底价",
                       "选择档位", "让价", "最终价格", "结果"])
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   rec.get("spu", ""), rec.get("site", ""), rec.get("price", ""),
                   rec.get("floor", ""), rec.get("level", ""), rec.get("discount", ""),
                   rec.get("final", ""), rec.get("status", "")])
        wb.save(TRAFFIC_RECORD_FILE)
    except Exception:
        pass   # 记录失败不中断主流程


@app.route('/traffic')
def traffic_page():
    """Temu 流量加速器页面 —— 打开后台 → 用户登录进页面 → 点「好了」→ 脚本接管。"""
    html_file = Path(__file__).parent / 'traffic.html'
    if html_file.exists():
        return send_file(str(html_file))
    return "<h1>traffic.html not found</h1><p>请确保 traffic.html 与 lovart_bridge.py 在同一目录</p>", 404


def _traffic_find_tab(browser):
    """在已打开的标签页里按 URL 特征找「流量加速器」页面，返回 (page, 候选URL列表)。"""
    pages = [pg for ctx in browser.contexts for pg in ctx.pages]
    for pg in pages:
        u = (pg.url or "").lower()
        if any(h in u for h in TRAFFIC_URL_HINTS) and "temu" in u:
            return pg, [p.url for p in pages]
    return None, [p.url for p in pages]


@app.route('/api/traffic/open', methods=['POST'])
def api_traffic_open():
    """① 连接共用 Edge(CDP 9222) 并打开 Temu 卖家后台首页，供用户登录。

    只用 p.stop() 断开 CDP 连接，绝不 browser.close()（会关掉用户的 Edge）。
    """
    log = []
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return jsonify({"ok": False, "error": "Bridge 环境未安装 playwright，无法控制 Edge。", "log": log})

    p = sync_playwright().start()
    try:
        try:
            browser, _ = _ensure_edge_cdp(p, log)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e),
                            "detail": "若 Edge 已开但端口 9222 未启用，请先关闭所有 Edge 再用「start-edge-cdp.bat」启动。",
                            "log": log})
        # 已有流量加速器标签页就复用，其次任意 Temu 后台标签页，都没有才新开
        page = None
        eu_page = None
        for ctx in browser.contexts:
            for pg in ctx.pages:
                u = (pg.url or "").lower()
                if any(h in u for h in TRAFFIC_URL_HINTS) and "temu" in u:
                    page = pg
                    break
                if not eu_page and "agentseller" in u and "temu" in u:
                    eu_page = pg
            if page:
                break
        if not page:
            page = eu_page
        if page:
            try:
                page.goto(TRAFFIC_HOME, timeout=30000)
                page.bring_to_front()
            except Exception:
                pass
            log.append("ℹ️ 复用已打开的 Temu 后台标签页，已跳转到「流量加速器」")
        else:
            try:
                ctx0 = browser.contexts[0] if browser.contexts else browser.new_context()
                page = ctx0.new_page()
                page.goto(TRAFFIC_HOME, timeout=30000)
                log.append("🌐 已新开「流量加速器」标签页")
            except Exception as e:
                return jsonify({"ok": False, "error": f"无法打开 Temu 卖家后台：{e}", "log": log})
        return jsonify({"ok": True, "log": log, "url": page.url})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "log": log})
    finally:
        try:
            p.stop()
        except Exception:
            pass


def _traffic_click_batch_button(page, log):
    """点「批量开启流量加速器」按钮并等抽屉弹出。返回 True=抽屉已弹出 / False=本页跳过。
    弹窗处理（轮询检测，2026-08-07 改）：
    - 「部分商品不可开启，要过滤并继续吗」→ 点「过滤并继续」，继续等抽屉；
    - 「商品已开启流量加速器，正在调价中」→ 点确认，本页无可开启商品，返回 False。"""
    clicked = False
    for _ in range(3):
        clicked = page.evaluate("""() => {
      const b = [...document.querySelectorAll("button")]
        .find(b => (b.innerText||'').includes('批量开启流量加速器'));
      if (!b || b.disabled) return false;
      b.click();
      return true;
    }""")
        if clicked:
            break
        page.wait_for_timeout(1500)
    if not clicked:
        raise RuntimeError("「批量开启流量加速器」按钮不可点击（可能未勾选到商品）。")
    log.append("✅ 已点击「批量开启流量加速器」，等待抽屉弹出…")
    page.wait_for_timeout(800)
    for _ in range(20):  # 轮询最多 ~16s：等抽屉 / 处理弹窗
        r = page.evaluate("""() => {
      const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
      if (drawer) return 'drawer';
      const dlgs = [...document.querySelectorAll("[class*='modal'], [class*='Modal'], [class*='MDL'], [class*='dialog'], [class*='Dialog'], [role='dialog']")]
        .filter(el => el.getBoundingClientRect().width > 50);
      for (const d of dlgs) {
        const t = d.innerText || '';
        const btn = (re) => [...d.querySelectorAll('button')].find(b => re.test((b.innerText||'').trim()));
        if (t.includes('不可开启') || t.includes('过滤并继续')) {
          const b = btn(/过滤并继续|确定|确认/);
          if (b) { b.click(); return 'filtered'; }
        } else if (t.includes('调价中') || t.includes('已开启流量加速器')) {
          const b = btn(/确定|确认|知道了|^好$/);
          if (b) { b.click(); return 'busy'; }
          return 'busy-no-btn';
        }
      }
      return 'waiting';
    }""")
        if r == 'drawer':
            page.wait_for_timeout(800)
            log.append("✅ 抽屉已弹出")
            return True
        if r == 'filtered':
            log.append("ℹ️ 检测到「部分商品不可开启」弹窗，已点「过滤并继续」")
            page.wait_for_timeout(1500)
        elif r in ('busy', 'busy-no-btn'):
            log.append("ℹ️ 本页商品已开启流量加速器（正在调价中），跳过本页")
            page.wait_for_timeout(1000)
            return False
        page.wait_for_timeout(800)
    _traffic_dump_dialogs(page, log)
    raise RuntimeError("等待批量开启抽屉弹出超时（页面结构可能变化或有未处理弹窗）。")


# 抽屉在原地关闭后，旧元素会带着 transform 滑出屏幕右侧但仍留在 DOM（僵尸抽屉），
# querySelector 取第一个匹配会拿到它。所有抽屉操作必须选「屏幕上可见」的那个。
_TRAFFIC_VISIBLE_DRAWER_FIND_JS = """[...document.querySelectorAll("div[class*='Drawer_content']")].find(d => {
  const r = d.getBoundingClientRect();
  return (Math.min(r.right, innerWidth) - Math.max(r.left, 0)) > 50 &&
         (Math.min(r.bottom, innerHeight) - Math.max(r.top, 0)) > 50;
})"""

_TRAFFIC_HAS_VISIBLE_DRAWER_JS = ("() => !! " + _TRAFFIC_VISIBLE_DRAWER_FIND_JS)


def _traffic_open_batch_drawer(page, log):
    """确保「批量开启流量加速器」抽屉已打开；没开就 全选 → 点批量开启按钮。
    返回 True=抽屉已弹出 / False=本页商品都在调价中，跳过。"""
    if page.evaluate(_TRAFFIC_HAS_VISIBLE_DRAWER_JS):
        log.append("ℹ️ 检测到批量开启抽屉已打开，直接处理当前抽屉内容")
        return True
    # 列表可能还在加载，重试几次找全选框（优先点表头的全选，没有才点最上方的）
    n = 0
    for _ in range(4):
        n = page.evaluate("""() => {
      const head = document.querySelector("thead [class*='CBX_outerWrapper'], tr[class*='TB_header'] [class*='CBX_outerWrapper']");
      const boxes = [...document.querySelectorAll("[class*='CBX_outerWrapper']")]
        .filter(el => { const r = el.getBoundingClientRect(); return r.width > 0; })
        .sort((a,b) => a.getBoundingClientRect().y - b.getBoundingClientRect().y);
      if (head) { head.click(); return Math.max(boxes.length, 1); }
      if (!boxes.length) return 0;
      boxes[0].click();
      return boxes.length;
    }""")
        if n:
            break
        page.wait_for_timeout(2000)
    if not n:
        raise RuntimeError("未找到商品列表的全选复选框（页面可能未加载完，或未登录）。")
    log.append(f"✅ 已点击全选（页面共 {n} 个复选框）")
    page.wait_for_timeout(600)
    return _traffic_click_batch_button(page, log)


def _traffic_read_list_rows(page):
    """读主列表行的 (SPU, 站点)（不开抽屉——行文本里直接有），用于黑名单预筛。"""
    return page.evaluate("""() => {
      const rows = [...document.querySelectorAll("tr")].filter(tr =>
        (tr.innerText||'').includes('SPU ID') && !tr.closest("div[class*='Drawer_content']"));
      return rows.map(tr => {
        const t = tr.innerText || '';
        const mSpu = t.match(/SPU ID[：:]\\s*(\\d+)/);
        const mSite = t.match(/经营站点[：:]\\s*([^\\n]+)/);
        return {spu: mSpu ? mSpu[1] : '', site: mSite ? mSite[1].trim() : ''};
      }).filter(x => x.spu);
    }""")


def _traffic_read_rows(page):
    """读取抽屉内所有数据行：SPU、站点、申报价、档位选项（含让价）。"""
    return page.evaluate("""() => {
      const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
      if (!drawer) return null;
      const rows = [...drawer.querySelectorAll("tr")].filter(tr => (tr.innerText||'').includes('SPU ID'));
      return rows.map((tr, ri) => {
        const tds = [...tr.querySelectorAll('td')];
        // SPU/站点从整行文本提取（商品信息不一定在第一个 td）
        const info = tr.innerText || '';
        const mSpu = info.match(/SPU ID[：:]\\s*(\\d+)/);
        const mSite = info.match(/经营站点[：:]\\s*([^\\n]+)/);
        // 申报价单元格 = 含 ¥ 且无单选框；档位单元格 = 含单选框且含 ¥（时效单元格有单选框但无 ¥）
        let priceTxt = '';
        let optCell = null;
        tds.forEach(td => {
          const t = td.innerText || '';
          const hasRadio = !!td.querySelector("label[class*='RD_outerWrapper']");
          if (hasRadio && t.includes('¥')) { if (!optCell) optCell = td; }
          else if (!hasRadio && !priceTxt) {
            const mP = t.match(/¥\\s*([\\d.,]+)/);
            if (mP) priceTxt = mP[1];
          }
        });
        const opts = optCell ? [...optCell.querySelectorAll("label[class*='RD_outerWrapper']")].map((lb, oi) => {
          const t = (lb.innerText||'');
          const name = (t.split('\\n')[0]||'').trim();
          const mPrice = t.match(/¥\\s*([\\d.,]+)/);
          return {oi, name, discount: mPrice ? mPrice[1] : ''};
        }) : [];
        return {ri, spu: mSpu ? mSpu[1] : '', site: mSite ? mSite[1].trim() : '', priceTxt, opts};
      });
    }""")


def _traffic_click_option(page, ri, oi):
    """点击第 ri 行第 oi 个档位选项，返回 'ok' / 'clicked-unverified' / 错误串。"""
    return page.evaluate("""(args) => {
      const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
      if (!drawer) return 'no-drawer';
      const rows = [...drawer.querySelectorAll("tr")].filter(tr => (tr.innerText||'').includes('SPU ID'));
      const tr = rows[args.ri];
      if (!tr) return 'row-missing';
      const tds = [...tr.querySelectorAll('td')];
      const optCell = tds.find(td => {
        const t = td.innerText || '';
        return !!td.querySelector("label[class*='RD_outerWrapper']") && t.includes('¥');
      });
      if (!optCell) return 'cell-missing';
      const lbs = [...optCell.querySelectorAll("label[class*='RD_outerWrapper']")];
      const lb = lbs[args.oi];
      if (!lb) return 'opt-missing';
      lb.click();
      const inp = lb.querySelector('input');
      return (inp && inp.checked) ? 'ok' : 'clicked-unverified';
    }""", {"ri": ri, "oi": oi})


def _traffic_analyze_rows(rows, log):
    """只分析不点击：对抽屉每行按核价底价规则算决策（规则 2026-08-06 用户确认）：
    - P = 日常申报价(CNY)，floor = 站点核价底价（沿用 ORDER_PRICE_FLOOR），L = 档位让价
    - 有选项 P-L ≥ floor            → 选让价最大的（有效价最接近底价）
    - 否则最小让价选项 P-L ≥ floor-10 → 选让价最少的
    - 否则（P-L 低于底价 10 元及以上）→ 价格不通过
    返回 (decisions, failed)：decisions = {spu: {name,L,eff,floor,price,site}}；不通过的已写记录。
    """
    decisions, failed = {}, []
    bl_skipped = []  # 命中黑名单跳过的 SPU（记日志用）
    for row in rows:
        _traffic_check_stop()
        spu = row.get("spu") or f"第{row['ri']+1}行"
        site = (row.get("site") or "").rstrip("站").strip()
        if (spu, site) in TRAFFIC_BLACKLIST:
            # 已知垃圾（落盘 7 天过期）：不重复分析/写记录，按不通过处理（不进决策、会被取消勾选）
            bl_skipped.append(spu)
            failed.append(spu)
            continue
        price = _parse_price(row.get("priceTxt") or "")
        opts = []
        for o in row.get("opts") or []:
            L = _parse_price(o.get("discount") or "")
            if L is not None:
                opts.append({"oi": o["oi"], "name": o.get("name") or "", "L": L})
        floor = ORDER_PRICE_FLOOR.get(site)
        tag = f"[{spu} {site}站]"
        rec = {"spu": spu, "site": site, "price": price if price is not None else (row.get("priceTxt") or ""),
               "floor": floor if floor is not None else "", "level": "", "discount": "", "final": ""}
        if floor is None:
            log.append(f"⚠️ {tag} 站点未配置核价底价，按不通过处理")
            failed.append(spu)
            rec["status"] = "价格不通过（站点未配置底价）"
            _traffic_record_once(rec, f"F:{spu}:{site}")
            continue
        if price is None or not opts:
            log.append(f"⚠️ {tag} 未读到申报价或档位选项，按不通过处理")
            failed.append(spu)
            rec["status"] = "价格不通过（未读到申报价或档位）"
            _traffic_record_once(rec, f"F:{spu}:{site}")
            _traffic_blacklist_add(spu, site, rec["status"])
            continue
        above = [o for o in opts if round(price - o["L"], 2) >= floor]
        if above:
            pick = max(above, key=lambda o: o["L"])
        else:
            cheap = min(opts, key=lambda o: o["L"])
            pick = cheap if round(price - cheap["L"], 2) >= floor - 10 else None
        if pick is None:
            log.append(f"❌ {tag} 申报价{price}，最少让价后仍低于底价{floor}超10元，价格不通过")
            failed.append(spu)
            rec["status"] = "价格不通过（低于底价10元及以上）"
            _traffic_record_once(rec, f"F:{spu}:{site}")
            _traffic_blacklist_add(spu, site, rec["status"])
            continue
        decisions[spu] = {"name": pick["name"], "L": pick["L"], "eff": round(price - pick["L"], 2),
                          "floor": floor, "price": price, "site": site}
    if bl_skipped:
        log.append(f"⏭ 跳过 {len(set(bl_skipped))} 个黑名单已知垃圾 SPU（不再重复分析）：{','.join(sorted(set(bl_skipped)))}")
    return decisions, failed


def _traffic_apply_decisions(page, decisions, log, record=True):
    """在当前抽屉里按 SPU 点选决策的档位（同名优先、让价兜底）。成功的写入记录。
    返回 (clicked, missing)：clicked=点中的 SPU 列表；missing=行始终没出现在抽屉里的 SPU
    （2026-08-08 实测：重开的抽屉偶发只含部分勾选商品，等 8s 也不再加载，只能关抽屉重开）。
    注意：Temu 列表/抽屉里同一 SPU 可能出现多行（重复行），每行都要点选，但日志和记录只写首次。
    record=False 用于剔除被拒 SPU 后的重开抽屉重选（记录已在首次写过，不重复写）。"""
    rows = _traffic_read_rows(page) or []
    # 重开抽屉后行有短暂异步渲染（实测真实行 ~2s 内渲染完），轮询等决策 SPU 的行出现。
    # 注意：等再久也不会出现的行 = 分析抽屉的过期快照/幻影数据或商品已流出本页
    # （2026-08-08 查实，见 _traffic_batch_enable 提交条件注释），所以最多等 ~3s 就交回上层判断。
    missing = set(decisions) - {r.get("spu") for r in rows}
    for _ in range(6):
        if not missing:
            break
        page.wait_for_timeout(500)
        page.evaluate("""() => {
          const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
          if (!drawer) return;
          [...drawer.querySelectorAll('*')]
            .filter(el => el.scrollHeight > el.clientHeight + 50)
            .forEach(el => { el.scrollTop = el.scrollHeight; });
        }""")
        rows = _traffic_read_rows(page) or []
        missing = set(decisions) - {r.get("spu") for r in rows}
    if missing:
        log.append(f"🔍 抽屉等 3s 后仍缺 {len(missing)} 个 SPU 的行：{','.join(sorted(missing))}（实际读到 {len(rows)} 行）")
    missing_spus = sorted(missing)
    clicked = []
    seen = set()
    for row in rows:
        _traffic_check_stop()
        spu = row.get("spu")
        d = decisions.get(spu)
        if not d:
            continue
        oi = None
        for o in row.get("opts") or []:
            if (o.get("name") or "") == d["name"]:
                oi = o["oi"]
                break
        if oi is None:
            for o in row.get("opts") or []:
                if _parse_price(o.get("discount") or "") == d["L"]:
                    oi = o["oi"]
                    break
        tag = f"[{spu} {d['site']}站]"
        if oi is None:
            if spu not in seen:
                log.append(f"⚠️ {tag} 抽屉里未找到「{d['name']}」选项，未选；请人工核对")
            continue
        r = _traffic_click_option(page, row["ri"], oi)
        if r != "ok":
            page.wait_for_timeout(400)
            r = _traffic_click_option(page, row["ri"], oi)
        if spu in seen:
            # 重复行：点选即可，日志/记录已在首次写过
            if r == "ok":
                clicked.append(spu)
            page.wait_for_timeout(50)
            continue
        seen.add(spu)
        rec = {"spu": spu, "site": d["site"], "price": d["price"], "floor": d["floor"],
               "level": d["name"], "discount": d["L"], "final": d["eff"]}
        if r == "ok":
            log.append(f"✅ {tag} 选「{d['name']}」让价{d['L']} → 有效价{d['eff']}（底价{d['floor']}）")
            clicked.append(spu)
            rec["status"] = "通过"
        else:
            log.append(f"⚠️ {tag} 点击「{d['name']}」未确认选中（{r}），请人工核对该行")
            rec["status"] = f"待人工核对（点击未确认：{r}）"
        if record:
            _traffic_record_once(rec, f"P:{spu}:{d['site']}")
        page.wait_for_timeout(50)
    return clicked, missing_spus


def _traffic_submit_drawer(page, log):
    """点抽屉底部「立即加速」提交（含二次确认弹窗）。返回 True/False。
    档位选完后按钮变可用可能有延迟，轮询等待几秒。
    提交后必须等抽屉自动关闭才算成功（2026-08-07：JS 点击偶发被吞/提交未生效时
    抽屉会一直开着，不关就会污染下一组）。
    若 Temu 报「submit higher custom price, sku:xxx」（该 SKU 要求更高价格），
    把对应行改选「让价最少」的档位重试（2026-08-07 用户确认）。
    返回 (ok, rejected_spu)：ok=提交成功；rejected_spu=连最少让价档都被拒的 SPU
    （调用方应把它从本组剔除后重试），无法定位时为 None。"""
    def drawer_open():
        return page.evaluate(_TRAFFIC_HAS_VISIBLE_DRAWER_JS)

    def click_confirm_dialog():
        """提交确认框「确认要批量开启流量加速器吗」→ 点「确认」。"""
        page.evaluate("""() => {
      const els = [...document.querySelectorAll("[class*='MDL'], [role='dialog'], [class*='Modal'], [class*='modal']")];
      for (const el of els) {
        const r = el.getBoundingClientRect();
        if (r.width < 10 || r.height < 10) continue;
        if (el.querySelector("div[class*='Drawer_content']") || el.closest("div[class*='Drawer_content']")) continue;
        const t = el.innerText || '';
        if (!t.includes('批量开启流量加速器')) continue;
        const b = [...el.querySelectorAll('button')].find(b => /确定|确认/.test((b.innerText||'').trim()));
        if (b) { b.click(); return true; }
      }
      return false;
    }""")

    def read_price_error():
        """读页面上的「submit higher custom price, sku:xxx」报错，返回 sku 或 None。"""
        return page.evaluate("""() => {
      const m = (document.body.innerText||'').match(/submit higher custom price[，,\\s]*sku[:：]\\s*(\\d+)/i);
      return m ? m[1] : null;
    }""")

    def switch_to_min_discount(sku, spu=None):
        """把报错 SKU 对应的抽屉行改选让价最少的档位（价格最高）。定位行的四级办法：
        ① 报错元素就近 closest('tr')（行内报错）；② 行 innerText 含 sku；③ 行 innerHTML 含 sku；
        ④ 接口映射 sku→SPU 后按 SPU 找行。返回 'ok' / 'already'（已是最少让价）/ 'not-found'。"""
        return page.evaluate("""(args) => {
          const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
          if (!drawer) return 'not-found';
          const rows = [...drawer.querySelectorAll("tr")].filter(tr => (tr.innerText||'').includes('SPU ID'));
          let tr = null;
          // ① 报错元素就近找行
          const errEl = [...drawer.querySelectorAll('*')].find(el =>
            el.children.length === 0 && (el.textContent||'').includes('submit higher custom price'));
          if (errEl) tr = errEl.closest('tr');
          // ② 行可见文本含 sku
          if (!tr) tr = rows.find(r => (r.innerText||'').includes(args.sku));
          // ③ 行 HTML（含隐藏属性/未渲染文本）含 sku
          if (!tr) tr = rows.find(r => (r.innerHTML||'').includes(args.sku));
          // ④ 接口映射出的 SPU 找行
          if (!tr && args.spu) tr = rows.find(r => (r.innerText||'').includes(args.spu));
          if (!tr) return 'not-found';
          const tds = [...tr.querySelectorAll('td')];
          const optCell = tds.find(td => {
            const t = td.innerText || '';
            return !!td.querySelector("label[class*='RD_outerWrapper']") && t.includes('¥');
          });
          if (!optCell) return 'not-found';
          const lbs = [...optCell.querySelectorAll("label[class*='RD_outerWrapper']")];
          let best = null, bestL = null;
          for (const lb of lbs) {
            const m = (lb.innerText||'').match(/¥\\s*([\\d.,]+)/);
            if (!m) continue;
            const L = parseFloat(m[1].replace(/,/g, ''));
            if (bestL === null || L < bestL) { bestL = L; best = lb; }
          }
          if (!best) return 'not-found';
          const inp = best.querySelector('input');
          if (inp && inp.checked) return 'already';
          best.click();
          const inp2 = best.querySelector('input');
          return (inp2 && inp2.checked) ? 'ok' : 'clicked-unverified';
        }""", {"sku": sku, "spu": spu or ""})

    def dump_row_diagnosis():
        """映射不到行时，把报错元素位置和前两行的 HTML 片段写日志，供下一步改进映射。"""
        try:
            infos = page.evaluate("""() => {
              const out = [];
              const errEls = [...document.querySelectorAll('body *')].filter(el =>
                el.children.length === 0 && (el.textContent||'').includes('submit higher custom price'));
              out.push('报错元素数=' + errEls.length + errEls.slice(0,2).map(el =>
                ' tag=' + el.tagName + ' cls=' + (el.className||'').toString().slice(0,60) +
                ' inDrawer=' + !!el.closest("div[class*='Drawer_content']") +
                ' closestTr=' + (el.closest('tr') ? 'Y' : 'N')).join('｜'));
              const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
              if (drawer) {
                const rows = [...drawer.querySelectorAll("tr")].filter(tr => (tr.innerText||'').includes('SPU ID'));
                rows.slice(0, 2).forEach((tr, i) =>
                  out.push('行' + i + ' HTML: ' + (tr.innerHTML||'').replace(/\\s+/g, ' ').slice(0, 500)));
              }
              return out;
            }""")
            for info in infos:
                log.append("🔍 行诊断: " + info)
        except Exception:
            pass

    def click_submit_js():
        return page.evaluate("""() => {
          const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
          if (!drawer) return 'no-drawer';
          const b = [...drawer.querySelectorAll("button")]
            .find(b => (b.innerText||'').trim().includes('立即加速'));
          if (!b) return 'no-btn';
          if (b.disabled) return 'disabled';
          b.click();
          return 'clicked';
        }""")

    def click_submit_real():
        """Playwright 真实点击「立即加速」（isTrusted 校验通过，实测最可靠）。"""
        try:
            page.locator("div[class*='Drawer_content'] button", has_text="立即加速").last.click(timeout=4000)
            return True
        except Exception:
            return False

    # 先等「立即加速」可用并点击（优先真实点击：JS 点击常被 isTrusted 吞掉，2026-08-07 实测几乎全靠真实点击兜底）
    clicked = False
    for _ in range(8):
        if click_submit_real() or click_submit_js() == 'clicked':
            clicked = True
            break
        page.wait_for_timeout(1000)
    if not clicked:
        log.append("⚠️ 未找到可点的「立即加速」按钮，请人工提交")
        return False, None
    log.append("✅ 已自动点击「立即加速」提交")

    retried_real = False
    for _ in range(16):  # 每轮：等 1s → 处理确认框 → 抽屉关了=成功；没关看是不是价格报错
        page.wait_for_timeout(1000)
        click_confirm_dialog()
        if not drawer_open():
            return True, None
        sku = read_price_error()
        if sku:
            spu = TRAFFIC_SKU_MAP.get(sku)
            r = switch_to_min_discount(sku, spu)
            if r in ('ok', 'clicked-unverified'):
                via = f"（映射到 SPU {spu}）" if spu else ""
                log.append(f"⚠️ Temu 要求 sku {sku}{via} 更高价格，已改选让价最少档，重试提交")
                page.wait_for_timeout(800)
                if not click_submit_real():
                    click_submit_js()
            elif r == 'already':
                log.append(f"⚠️ sku {sku} 已是让价最少档仍被 Temu 拒绝，需从本组剔除")
                return False, spu
            else:
                log.append(f"⚠️ Temu 拒绝 sku {sku} 但抽屉里找不到对应行，本组按未提交处理")
                dump_row_diagnosis()
                return False, None
        elif not retried_real:
            # 没有价格报错但抽屉没关：点击可能被吞了，再补一次真实点击
            retried_real = True
            log.append("⚠️ 提交后抽屉未关闭，用真实点击重试「立即加速」…")
            click_submit_real()
    log.append("⚠️ 提交后抽屉始终未关闭，按未提交处理（避免污染下一组）")
    return False, None


def _traffic_batch_enable(page, log):
    """每页流程（2026-08-07 用户确认；2026-08-08 加黑名单预筛）：
    先不开抽屉读列表行做黑名单预筛 → 只勾选非垃圾行 → 开抽屉按核价底价规则分析 →
    全部通过：直接选档提交；部分通过：关抽屉 → 只勾选通过的商品 → 重开抽屉选档提交；
    全部不通过：跳过本页。返回 (passed, failed, submitted)。"""
    # 黑名单预筛（2026-08-08 用户确认）：垃圾行永远不离开列表，每次全选开抽屉分析一遍
    # 要烧 20~35s；列表行文本直接带 SPU+站点，不开抽屉就能过滤——整页垃圾 ~2s 跳过，
    # 部分垃圾则只勾选非垃圾行的 SPU（垃圾行不进抽屉，省掉关/重开一整轮）。
    listed = _traffic_read_list_rows(page) or []
    if listed:
        fresh = [x for x in listed
                 if (x["spu"], (x["site"] or "").rstrip("站").strip()) not in TRAFFIC_BLACKLIST]
        if not fresh:
            log.append(f"🗑 本页 {len(listed)} 行全部命中黑名单（已知垃圾），跳过本页")
            return [], [x["spu"] for x in listed], False
        if len(fresh) < len(listed):
            log.append(f"ℹ️ 本页 {len(listed)} 行，{len(listed) - len(fresh)} 行命中黑名单不勾选，只勾选其余 {len(fresh)} 行开抽屉")
        _traffic_select_spus(page, sorted({x["spu"] for x in fresh}), log)
        if not _traffic_click_batch_button(page, log):
            log.append("⏸ 本页无可开启商品（均已开启/调价中），跳过提交")
            return [], [], False
    elif not _traffic_open_batch_drawer(page, log):  # 列表行读不到（结构变化兜底）：走全选流程
        log.append("⏸ 本页无可开启商品（均已开启/调价中），跳过提交")
        return [], [], False
    rows = _traffic_read_rows(page)
    if not rows:
        raise RuntimeError("抽屉里未读到任何商品行。")
    log.append(f"ℹ️ 抽屉内共 {len(rows)} 个商品，开始按核价底价规则分析…")
    decisions, failed = _traffic_analyze_rows(rows, log)
    # 历史上被 Temu 拒绝（要求更高价格）的 SPU 直接跳过：再提交也只会再被拒，白跑剔除流程
    rej_skip = [s for s in decisions if s in TRAFFIC_REJECTED_SPUS]
    if rej_skip:
        for s in rej_skip:
            decisions.pop(s)
        failed += rej_skip
        log.append(f"⏭ 跳过 {len(rej_skip)} 个曾被 Temu 拒绝的 SPU：{','.join(rej_skip)}（不再尝试提交）")

    if decisions and failed:
        log.append(f"ℹ️ 本页 {len(decisions)} 个通过 / {len(failed)} 个不通过：只勾选通过的商品重新打开抽屉…")
        st = _traffic_page_state(page)
        _traffic_close_drawer(page, log, target_page=st["pageNo"])
        _traffic_select_spus(page, list(decisions.keys()), log)
        if not _traffic_click_batch_button(page, log):
            log.append("⚠️ 重开抽屉时提示商品调价中，本页改为人工处理")
            return [], failed, False
    elif not decisions:
        log.append("❌ 价格不通过 " + str(len(failed)) + " 个：" + ",".join(failed))
        log.append("⏸ 本页全部价格不通过，跳过提交")
        st = _traffic_page_state(page)
        _traffic_close_drawer(page, log, target_page=st["pageNo"])
        return [], failed, False

    passed = []
    excluded = []  # 被 Temu 提交校验拒绝（要求更高价格）、从本组剔除的 SPU
    for attempt in range(6):
        passed, missing = _traffic_apply_decisions(page, decisions, log, record=(attempt == 0))
        log.append("✅ 价格通过 " + str(len(passed)) + " 个：" + ",".join(passed))
        log.append("❌ 价格不通过 " + str(len(failed)) + " 个：" + (",".join(failed) if failed else "无"))
        # 提交条件：当前抽屉里真实存在的决策行全部点中（clicked 按行计、decisions 按 SPU 去重，用集合判断）。
        # 缺行的 SPU 不阻塞提交（2026-08-08 查实：深页分析抽屉会拿到过期快照/幻影数据——
        # 同一 SPU+站点行连价格都一字不差地出现在多个页码的分析里，甚至包含已提交成功的商品；
        # 只有按当前真实勾选重开的抽屉可信。幻影/已漂走的 SPU 等后续轮次遇到真身再处理）。
        if missing:
            log.append(f"ℹ️ {len(missing)} 个 SPU 不在当前抽屉（分析快照过期或已流出本页），按实际在抽屉的 {len(passed)} 个处理：{','.join(sorted(missing))}")
        if not (passed and set(passed) >= (set(decisions.keys()) - set(missing))):
            if passed:
                log.append("⏸ 有商品未能选中档位，本页不自动提交，请人工核对后处理")
            return passed, failed, False
        submitted, rej = _traffic_submit_drawer(page, log)
        if submitted:
            return passed, failed, True
        if not rej or rej in excluded or rej not in decisions:
            return passed, failed, False
        # 连最少让价档都被 Temu 拒绝的 SPU：剔除出本组，剩余商品重开抽屉重试提交；
        # 并记入 TRAFFIC_REJECTED_SPUS，后续页/后续轮次分析阶段直接跳过，不再浪费剔除流程
        excluded.append(rej)
        TRAFFIC_REJECTED_SPUS.add(rej)
        failed.append(rej)
        d0 = decisions.get(rej) or {}
        _traffic_blacklist_add(rej, d0.get("site", ""), "Temu拒绝（要求更高价格，已剔除）")
        _traffic_record({"spu": rej, "site": d0.get("site", ""), "price": d0.get("price", ""),
                         "floor": d0.get("floor", ""), "level": "", "discount": "", "final": "",
                         "status": "Temu拒绝（要求更高价格，已剔除）"})
        decisions = {k: v for k, v in decisions.items() if k not in excluded}
        log.append(f"⚠️ SPU {rej} 被 Temu 拒绝（要求更高价格），从本组剔除，剩余 {len(decisions)} 个重试提交")
        if not decisions:
            return passed, failed, False
        st = _traffic_page_state(page)
        _traffic_close_drawer(page, log, target_page=st["pageNo"])
        _traffic_select_spus(page, list(decisions.keys()), log)
        if not _traffic_click_batch_button(page, log):
            log.append("⚠️ 重开抽屉时提示商品调价中，本组剩余改为人工处理")
            return passed, failed, False
    return passed, failed, False


def _traffic_apply_pending_filter(page, log):
    """点顶部「流量加速器待开启」快捷筛选卡片（已选中则跳过——橙色边框代表选中）。"""
    for _ in range(3):
        r = page.evaluate("""() => {
      const card = [...document.querySelectorAll("div[class*='quick-overdue-filter_card']")]
        .find(el => (el.innerText||'').includes('流量加速器待开启'));
      if (!card) return 'missing';
      const active = (card.getAttribute('style')||'').includes('255, 103, 2') ||
                     getComputedStyle(card).borderColor === 'rgb(255, 103, 2)';
      if (active) return 'already';
      card.click();
      return 'clicked';
    }""")
        if r == 'clicked':
            log.append("✅ 已切换到「流量加速器待开启」筛选")
            page.wait_for_timeout(3000)
            return
        if r == 'already':
            log.append("ℹ️ 「流量加速器待开启」筛选已选中")
            return
        page.wait_for_timeout(2000)
    raise RuntimeError("未找到「流量加速器待开启」筛选卡片（页面结构可能变化）。")


def _traffic_page_state(page):
    """主列表状态：总条数、当前页码、上/下一页是否可用、可见复选框数、抽屉是否开着、每页条数。"""
    return page.evaluate("""() => {
      const drawerEl = [...document.querySelectorAll("div[class*='Drawer_content']")].find(d => {
        const r = d.getBoundingClientRect();
        const visW = Math.min(r.right, innerWidth) - Math.max(r.left, 0);
        const visH = Math.min(r.bottom, innerHeight) - Math.max(r.top, 0);
        return visW > 50 && visH > 50;  // 按屏幕实际可见像素判断（关闭后元素滑出屏幕但仍留在DOM）
      });
      const drawerOpen = !!drawerEl;
      const bodyText = document.body.innerText || '';
      const t = bodyText.match(/共有\\s*([\\d,]+)\\s*条/);
      const act = document.querySelector("li[class*='PGT_pagerItemActive']");
      const next = document.querySelector("li[class*='PGT_next']");
      const nextDisabled = !next || (next.className||'').toLowerCase().includes('disab');
      const prev = document.querySelector("li[class*='PGT_prev']");
      const prevDisabled = !prev || (prev.className||'').toLowerCase().includes('disab');
      const boxes = [...document.querySelectorAll("[class*='CBX_outerWrapper']")]
        .filter(el => el.getBoundingClientRect().width > 0).length;
      const sizeIpt = document.querySelector("[class*='PGT_sizeSelect'] input");
      const size = sizeIpt ? (parseInt(sizeIpt.value, 10) || 30) : 30;
      return {total: t ? parseInt(t[1].replace(/,/g,'')) : -1,
              pageNo: act ? parseInt((act.innerText||'').trim(), 10) : -1,
              nextDisabled, prevDisabled, boxes, drawerOpen, size};
    }""")


def _traffic_goto_page(page, target, log):
    """刷新后页码会回到 1，逐页点「下一页」回到 target 页。"""
    for _ in range(80):
        st = _traffic_page_state(page)
        if st["pageNo"] < 1 or st["pageNo"] >= target or st["nextDisabled"]:
            break
        page.evaluate("() => { const n = document.querySelector(\"li[class*='PGT_next']\"); if (n) n.click(); }")
        page.wait_for_timeout(1800)
    cur = _traffic_page_state(page)["pageNo"]
    log.append(f"ℹ️ 已回到第 {cur} 页" + ("" if cur == target else f"（目标第 {target} 页，未能到达，从当前页继续）"))


def _traffic_list_signature(page):
    """主列表前 5 行的 SPU 签名：用于判断跳页后行内容是否真的刷新了。
    实测页码先变、旧页的行（含勾选状态）会残留 ~1-3s 才被替换，只看页码会在旧行上误操作。"""
    try:
        return page.evaluate("""() => {
      const rows = [...document.querySelectorAll("tr")].filter(tr =>
        (tr.innerText||'').includes('SPU ID') && !tr.closest("div[class*='Drawer_content']"));
      return rows.slice(0, 5).map(tr => {
        const m = (tr.innerText||'').match(/SPU ID[：:]\\s*(\\d+)/);
        return m ? m[1] : '';
      }).join(',');
    }""")
    except Exception:
        return ""


def _traffic_goto_page_num(page, target, log):
    """跳到指定页码：优先直接点目标页码数字；点不到就点「不超过目标的最大可见页码」
    大步靠近（分页器只渲染当前页附近的页码），最后才用上/下一页步进。返回实际到达的页码。"""
    miss = 0
    for _ in range(200):
        if TRAFFIC_STOP.is_set():
            break
        st = _traffic_page_state(page)
        cur = st["pageNo"]
        if cur < 1:
            miss += 1
            if miss >= 10:
                log.append(f"⚠️ 列表 ~15 秒未加载（可能网络异常），放弃跳转到第 {target} 页")
                break
            page.wait_for_timeout(1500)
            continue
        miss = 0
        if cur == target or (cur < target and st["nextDisabled"]) or (cur > target and st["prevDisabled"]):
            break
        prev_sig = _traffic_list_signature(page)
        page.evaluate("""(args) => {
          const items = [...document.querySelectorAll("li[class*='PGT_pagerItem']")];
          const nums = items
            .map(li => ({li, n: parseInt((li.innerText||'').trim(), 10)}))
            .filter(x => !isNaN(x.n));
          const exact = nums.find(x => x.n === args.target);
          if (exact) { exact.li.click(); return 'num'; }
          if (args.target > args.cur) {
            const cand = nums.filter(x => x.n > args.cur && x.n <= args.target)
              .sort((a, b) => b.n - a.n)[0];
            if (cand) { cand.li.click(); return 'hop'; }
            const n = document.querySelector("li[class*='PGT_next']");
            if (n) { n.click(); return 'step'; }
          } else {
            const cand = nums.filter(x => x.n < args.cur && x.n >= args.target)
              .sort((a, b) => a.n - b.n)[0];
            if (cand) { cand.li.click(); return 'hop'; }
            const p = document.querySelector("li[class*='PGT_prev']");
            if (p) { p.click(); return 'step'; }
          }
          return null;
        }""", {"target": target, "cur": cur})
        # 轮询等页码变化，变了就立刻继续，不吃满固定等待
        changed = False
        for _ in range(6):
            page.wait_for_timeout(300)
            if _traffic_page_state(page)["pageNo"] != cur:
                changed = True
                break
        if changed:
            # 页码变了 ≠ 行内容刷新了（2026-08-08 实测旧页行连勾选状态残留 ~1-3s），
            # 等行签名变化再走，否则预筛/勾选会打在旧页的过期行上（点了 24 个实际勾中 0 个）
            for _ in range(14):
                if _traffic_list_signature(page) != prev_sig:
                    break
                page.wait_for_timeout(300)
    cur = _traffic_page_state(page)["pageNo"]
    if cur != target:
        log.append(f"⚠️ 未能到达第 {target} 页（当前第 {cur} 页），从当前页继续")
    return cur


def _traffic_drawer_confirm_dialog_js():
    """JS：在抽屉外找可见的弹窗（取消抽屉时可能弹「确定要取消吗」），点掉它的确认按钮。
    注意：「确认要批量开启流量加速器吗」是提交确认框，绝不能在这里点「确认」，直接跳过。
    返回点到的按钮文本，没弹窗返回 null。"""
    return """() => {
      const els = [...document.querySelectorAll("[class*='Modal'],[class*='modal'],[class*='Dialog'],[class*='dialog'],[class*='MDL'],[role='dialog']")];
      for (const el of els) {
        const r = el.getBoundingClientRect();
        if (r.width < 10 || r.height < 10) continue;
        if (el.querySelector("div[class*='Drawer_content']") || el.closest("div[class*='Drawer_content']")) continue;
        const text = el.innerText || '';
        if (text.includes('批量开启流量加速器')) continue;  // 提交确认框，不是关抽屉的确认框
        const btns = [...el.querySelectorAll('button')];
        const btn = btns.find(b => {
          const t = (b.innerText||'').trim();
          return t && t !== '取消' && /确定|确认|仍要取消|放弃|关闭/.test(t);
        });
        if (btn) { btn.click(); return (btn.innerText||'').trim(); }
      }
      return null;
    }"""


def _traffic_dump_dialogs(page, log):
    """诊断：把当前所有可见弹窗的文本和按钮写进日志，便于排查抽屉关不掉的原因。"""
    try:
        infos = page.evaluate("""() => {
          const out = [];
          const els = [...document.querySelectorAll("[class*='Modal'],[class*='modal'],[class*='Dialog'],[class*='dialog'],[class*='MDL'],[role='dialog']")];
          for (const el of els) {
            const r = el.getBoundingClientRect();
            if (r.width < 10 || r.height < 10) continue;
            const text = (el.innerText||'').replace(/\\s+/g, ' ').slice(0, 150);
            const btns = [...el.querySelectorAll('button')].map(b => {
              const t = (b.innerText||'').trim();
              const dis = b.disabled || (b.className||'').includes('disabled');
              return t + (dis ? '(禁用)' : '');
            }).filter(Boolean);
            out.push(text + ' ｜按钮: ' + btns.join('/'));
          }
          return out.slice(0, 5);
        }""")
        for info in infos:
            log.append("🔍 弹窗诊断: " + info)
        if not infos:
            log.append("🔍 弹窗诊断: 页面上没有可见弹窗（抽屉仍未关闭）")
    except Exception:
        pass


def _traffic_close_drawer(page, log, target_page=None):
    """三级兜底关抽屉：① Playwright 真实点击「取消」（不加 force，等按钮稳定/可见，最接近人手）
    → ② JS 点击 + 处理二次确认弹窗 → ③ 刷新页面。

    实测：抽屉按钮坐标会抖动，force 点击按旧坐标点空；JS 点击被页面的 isTrusted 校验忽略，
    所以把真实点击提到第一级。刷新后重新进「待开启」筛选并回到 target_page。返回 True/False。
    """
    def drawer_open():
        """抽屉关闭后元素滑出屏幕右侧但仍留在 DOM（rect 仍有 1400x1308），
        必须按「与视口的相交像素」判断：可见宽高都 >50px 才算真的开着。"""
        try:
            return page.evaluate("""() => {
              return [...document.querySelectorAll("div[class*='Drawer_content']")].some(d => {
                const r = d.getBoundingClientRect();
                const visW = Math.min(r.right, innerWidth) - Math.max(r.left, 0);
                const visH = Math.min(r.bottom, innerHeight) - Math.max(r.top, 0);
                return visW > 50 && visH > 50;
              });
            }""")
        except Exception:
            return False
    if not drawer_open():
        return True

    def handle_confirm():
        clicked = page.evaluate(_traffic_drawer_confirm_dialog_js())
        if clicked:
            log.append(f"ℹ️ 检测到二次确认弹窗，已点「{clicked}」")
            page.wait_for_timeout(1500)
        return clicked

    def wait_closed(max_ms=2500):
        """轮询等抽屉关闭，关了就立刻返回，不吃满固定等待。"""
        for _ in range(max_ms // 250):
            if not drawer_open():
                return True
            page.wait_for_timeout(250)
        return not drawer_open()

    # ① 真实鼠标点击（不加 force：Playwright 会等按钮稳定、滚动进视口再点，最接近手动点击）
    try:
        page.locator("div[class*='Drawer_content'] button", has_text="取消").last.click(timeout=6000)
    except Exception:
        pass
    if wait_closed():
        return True
    handle_confirm()
    if not drawer_open():
        return True
    # ② JS 点击取消
    page.evaluate("""() => {
      const drawer = """ + _TRAFFIC_VISIBLE_DRAWER_FIND_JS + """;
      const cancel = drawer && [...drawer.querySelectorAll("button")].find(b => (b.innerText||'').trim() === '取消');
      if (cancel) cancel.click();
    }""")
    if wait_closed():
        return True
    handle_confirm()
    if not drawer_open():
        return True
    # ③ 刷新页面兜底（僵尸抽屉唯一可靠的清理方式），刷新前先把弹窗状态写进日志
    _traffic_dump_dialogs(page, log)
    log.append("⚠️ 抽屉无法正常关闭，刷新页面重置状态…")
    try:
        page.reload(timeout=30000)
    except Exception:
        pass
    page.wait_for_timeout(4000)
    _traffic_apply_pending_filter(page, log)
    if target_page and target_page > 1:
        _traffic_goto_page(page, target_page, log)
    return True


def _traffic_select_spus(page, spus, log):
    """把主列表的勾选状态精确设置为：只勾选 spus 里的商品（按行内 SPU ID 匹配）。
    列表未加载时重试几次。返回最后一次的统计 dict。"""
    js = """(wanted) => {
      const set = new Set(wanted);
      const rows = [...document.querySelectorAll("tr")].filter(tr =>
        (tr.innerText||'').includes('SPU ID') && !tr.closest("div[class*='Drawer_content']"));
      let matched = 0, toggled = 0;
      for (const tr of rows) {
        const m = (tr.innerText||'').match(/SPU ID[：:]\\s*(\\d+)/);
        if (!m) continue;
        const lb = tr.querySelector("label[class*='CBX_outerWrapper']");
        if (!lb) continue;
        matched++;
        const checked = lb.getAttribute('data-checked') === 'true';
        if (checked !== set.has(m[1])) { lb.click(); toggled++; }
      }
      return {rows: rows.length, matched, toggled};
    }"""
    r = {"rows": 0, "matched": 0, "toggled": 0}
    for _ in range(4):
        r = page.evaluate(js, list(spus))
        if r["matched"] > 0:
            break
        page.wait_for_timeout(2000)
    log.append(f"✅ 已按通过清单重新勾选：本页 {r['matched']} 行，调整 {r['toggled']} 个复选框")
    if r["matched"] == 0:
        raise RuntimeError("主列表未加载出商品行，无法按通过清单勾选。")
    page.wait_for_timeout(400)
    # 勾选状态校验（2026-08-08 排障用）：区分「勾选本身错」和「重开抽屉快照错」
    chk = page.evaluate("""() => {
      const rows = [...document.querySelectorAll("tr")].filter(tr =>
        (tr.innerText||'').includes('SPU ID') && !tr.closest("div[class*='Drawer_content']"));
      let on = 0;
      for (const tr of rows) {
        const lb = tr.querySelector("label[class*='CBX_outerWrapper']");
        if (lb && lb.getAttribute('data-checked') === 'true') on++;
      }
      return {on, total: rows.length};
    }""")
    log.append(f"   勾选校验：主列表 {chk['total']} 行中已勾 {chk['on']} 行")
    return r


def _traffic_set_id_query(page, text, log, tag=""):
    """把 text 填进「商品ID查询」输入框（placeholder 含「多个查询」）并点「查询」。
    text 为空串 = 清空查询恢复完整列表。React 受控输入框必须用原生 setter + input 事件。"""
    r = page.evaluate("""(text) => {
      const ipt = [...document.querySelectorAll('input')].find(i => (i.placeholder||'').includes('多个查询'));
      if (!ipt) return 'no-input';
      const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
      setter.call(ipt, text);
      ipt.dispatchEvent(new Event('input', {bubbles: true}));
      ipt.dispatchEvent(new Event('change', {bubbles: true}));
      const btn = [...document.querySelectorAll('button')].find(b => (b.innerText||'').trim() === '查询');
      if (!btn) return 'no-btn';
      btn.click();
      return 'ok';
    }""", text)
    if r != 'ok':
        raise RuntimeError(f"商品ID查询操作失败（{r}），页面结构可能变化。")
    page.wait_for_timeout(3000)
    if tag:
        log.append(tag)


def _traffic_cleanup_zombies(page, log, threshold=15):
    """僵尸抽屉（关闭后滑出屏幕但留在 DOM）会随开窗次数无限堆积、拖慢页面，
    堆积超过 threshold 个时刷新页面清理（刷新是唯一可靠的清理方式），刷新后重选「待开启」筛选。"""
    try:
        n = page.evaluate("() => document.querySelectorAll(\"div[class*='Drawer_content']\").length")
    except Exception:
        return
    if not n or n <= threshold:
        return
    log.append(f"🧹 DOM 中已堆积 {n} 个抽屉元素，刷新页面清理…")
    try:
        page.reload(timeout=30000)
    except Exception:
        pass
    page.wait_for_timeout(5000)
    _traffic_apply_pending_filter(page, log)


def _traffic_loop(page, log):
    """主循环（2026-08-08 用户确认改为顺序逐页，取代 1/5/3/2/4 交错 + 窗口前移）：

    按 1→2→3→4→… 顺序逐页处理，越过末页后回第 1 页开始下一趟；
    连续两趟全列表无提交 → 收工。垃圾页有黑名单预筛兜底（~2s 跳过），
    不再需要交错回访/窗口前移来绕开垃圾；顺序页码也让抽屉数据紧跟当前页，
    减少深页分析抽屉拿到过期快照（幻影数据）的概率。
    每页流程（_traffic_batch_enable）：黑名单预筛 → 勾选非垃圾行开抽屉 → 只分析 →
    全通过直接选档提交 / 部分通过关抽屉重勾重开后选档提交（部分提交）。
    返回 (passed, failed, submitted_pages, manual_pages)。
    """
    all_passed, all_failed = [], []
    submitted_pages = 0
    manual_pages = 0
    _traffic_apply_pending_filter(page, log)

    target = 0           # 顺序逐页：每轮 +1
    load_fails = 0
    junk_pages = 0       # 整页都是滞留垃圾（价格不通过/被拒/读不出价/黑名单），不算「需人工」
    sweep_submitted = False  # 本趟扫描（第 1 页 → 越过末页）是否有提交
    empty_sweeps = 0
    for rnd in range(1, 2000):
        if TRAFFIC_STOP.is_set():
            log.append("⏹ 收到停止信号，已安全停止（已提交的页不受影响）。")
            break
        st = _traffic_page_state(page)
        if st["drawerOpen"]:
            _traffic_close_drawer(page, log, target_page=st["pageNo"])
            st = _traffic_page_state(page)
        _traffic_cleanup_zombies(page, log)
        if st["total"] == 0:
            # 列表接口抽风会间歇返回 0 条（2026-08-07 实测），刷新重查确认是真的空了再收工
            empty_confirmed = True
            for _ in range(2):
                log.append("⏳ 读到 0 条，刷新页面确认列表是否真的空了…")
                try:
                    page.reload(timeout=30000)
                except Exception:
                    pass
                page.wait_for_timeout(6000)
                _traffic_apply_pending_filter(page, log)
                page.wait_for_timeout(3000)
                if _traffic_page_state(page)["total"] != 0:
                    empty_confirmed = False
                    break
            if empty_confirmed:
                log.append("✅ 没有待开启的商品了，循环结束")
                break
            st = _traffic_page_state(page)
        if st["total"] < 0 or st["size"] < 1:
            log.append("❌ 未读到列表总数（页面可能未加载完）。循环停止，请刷新页面后重试。")
            break
        tp = -(-st["total"] // st["size"])  # 总页数，向上取整
        if target + 1 > tp:
            # 一趟扫描越过末页：这趟有提交 → 回第 1 页继续；连续两趟全列表无提交 → 收工
            if sweep_submitted:
                empty_sweeps = 0
            else:
                empty_sweeps += 1
                if empty_sweeps >= 2:
                    log.append("✅ 连续两趟全列表扫描都没有可提交的商品，循环结束")
                    break
                log.append("🔁 整趟扫描没有提交，回到第 1 页再扫一趟确认…")
            sweep_submitted = False
            target = 0
        target += 1
        log.append(f"—— 第 {rnd} 轮：处理第 {target} 页（待开启共 {st['total']} 条）——")

        try:
            _traffic_check_stop()
            _traffic_goto_page_num(page, target, log)
            pno = target
            # 列表未加载（Temu 限流/网络异常）时主动恢复：点「查询」→ 刷新页面 → 兜底等待
            loaded = False
            for attempt in range(3):
                stp = _traffic_page_state(page)
                if stp["pageNo"] >= 1 and stp["boxes"] > 1:
                    loaded = True
                    break
                if attempt == 0:
                    log.append(f"⏳ 第 {pno} 页列表未加载，点「查询」重新触发（1/3）…")
                    try:
                        _traffic_set_id_query(page, "", log)  # 空串 = 清空查询条件，纯重新触发查询
                    except Exception:
                        page.evaluate("""() => {
                          const b = [...document.querySelectorAll('button')].find(b => (b.innerText||'').trim() === '查询');
                          if (b) b.click();
                        }""")
                    page.wait_for_timeout(12000)
                    # 点「查询」后列表会重置回第 1 页，必须跳回原页码
                    _traffic_goto_page_num(page, pno, log)
                    page.wait_for_timeout(3000)
                elif attempt == 1:
                    log.append(f"⏳ 第 {pno} 页列表仍未加载，刷新页面重试（2/3）…")
                    try:
                        page.reload(timeout=30000)
                    except Exception:
                        pass
                    page.wait_for_timeout(6000)
                    try:
                        _traffic_apply_pending_filter(page, log)
                    except Exception:
                        pass
                    page.wait_for_timeout(3000)
                    _traffic_goto_page_num(page, pno, log)
                    page.wait_for_timeout(5000)
                else:
                    log.append(f"⏳ 第 {pno} 页列表未加载，等待 30 秒重试（3/3）…")
                    page.wait_for_timeout(30000)
            if not loaded:
                load_fails += 1
                if load_fails >= 3:
                    # 连续加载失败 = 大概率被限流，暂停 3 分钟等解除，不再空扫
                    log.append("⏸ 连续 3 页加载失败，暂停 3 分钟等限流解除…")
                    for _ in range(18):
                        _traffic_check_stop()
                        page.wait_for_timeout(10000)
                    load_fails = 0
                else:
                    log.append(f"⚠️ 第 {pno} 页列表持续未加载（可能被限流），跳过本页")
                continue
            load_fails = 0
            try:
                passed, failed, submitted = _traffic_batch_enable(page, log)
            except _TrafficStopped:
                raise
            except Exception as e:
                log.append(f"❌ 第 {pno} 页处理异常：{e}，跳过本页")
                try:
                    _traffic_close_drawer(page, log, target_page=pno)
                except Exception:
                    pass
                continue
            all_passed += passed
            all_failed += failed
            if submitted:
                submitted_pages += 1
                sweep_submitted = True
            else:
                _traffic_close_drawer(page, log, target_page=pno)
                if passed:
                    manual_pages += 1   # 有通过的商品但没提交成功，需要人工核对
                else:
                    junk_pages += 1     # 整页都是滞留垃圾（价格不通过/被拒/读不出价）
            # 页间间隔，降低触发 Temu 限流的频率
            page.wait_for_timeout(2500)
        except _TrafficStopped:
            log.append("⏹ 收到停止信号，已安全停止（已提交的页不受影响）。")
            try:
                _traffic_close_drawer(page, log)
            except Exception:
                pass
            break
    else:
        log.append("⚠️ 已达最大轮数上限（2000），循环停止")

    log.append(f"📊 汇总：自动提交 {submitted_pages} 页；未提交（需人工）{manual_pages} 页；"
               f"整页垃圾跳过 {junk_pages} 页；价格通过 {len(all_passed)} 个；价格不通过 {len(all_failed)} 个")
    log.append(f"📄 每条明细（SPU/站点/所选价格等）已记录到 {TRAFFIC_RECORD_FILE}")
    return all_passed, all_failed, submitted_pages, manual_pages


def _traffic_run(task_id):
    """后台线程：连接 Edge → 找流量加速器标签页 → 全选+批量开启 → 按核价底价规则逐行选档位。"""
    task = TRAFFIC_TASKS.get(task_id)
    if not task:
        return
    log = task["log"]
    lock = task["lock"]
    try:
        from playwright.sync_api import sync_playwright
        p = sync_playwright().start()
        try:
            try:
                browser, _ = _ensure_edge_cdp(p, log)
            except Exception as e:
                with lock:
                    task["error"] = str(e)
                    task["done"] = True
                return
            page, candidates = _traffic_find_tab(browser)
            if not page:
                temu_tabs = [u for u in candidates if "temu" in (u or "").lower()]
                with lock:
                    task["error"] = ("未找到「流量加速器」标签页。请先手动进入该页面再点「好了」。"
                                     "当前 Temu 标签页：" + ("；".join(temu_tabs) if temu_tabs else "（无）"))
                    task["done"] = True
                return
            try:
                page.bring_to_front()
            except Exception:
                pass
            try:
                title = page.title()
            except Exception:
                title = ""
            log.append(f"✅ 已定位流量加速器标签页：{title}")
            log.append(f"ℹ️ 页面 URL：{page.url}")
            page.wait_for_timeout(800)

            def _sku_map_sniffer(resp):
                """从列表接口的 JSON 响应里挖 skuId→goodsId(SPU) 映射
                （提交报错「submit higher custom price」只给 sku，抽屉行只有 SPU，靠它对应）。"""
                try:
                    if 'json' not in (resp.headers.get('content-type') or ''):
                        return
                    body = resp.json()
                except Exception:
                    return

                def walk(o, ctx_goods=None):
                    if isinstance(o, dict):
                        # productId = 页面展示的「SPU ID」（goodsId 是内部货号，对不上行）
                        g = (o.get('productId') or o.get('spuId') or o.get('goodsId') or ctx_goods)
                        skl = o.get('productSkuIdList') or o.get('skuIdList')
                        if g and isinstance(skl, list):
                            for s in skl:
                                TRAFFIC_SKU_MAP[str(s)] = str(g)
                        sku = o.get('skuId') or o.get('sku_id')
                        if sku and g:
                            TRAFFIC_SKU_MAP[str(sku)] = str(g)
                        for v in o.values():
                            walk(v, g)
                    elif isinstance(o, list):
                        for v in o:
                            walk(v, ctx_goods)
                walk(body)
            try:
                page.on("response", _sku_map_sniffer)
            except Exception:
                pass

            try:
                passed, failed, sub_pages, manual_pages = _traffic_loop(page, log)
            except Exception as e:
                with lock:
                    task["error"] = str(e)
                    task["done"] = True
                return
            with lock:
                note = (f"✅ 循环结束：自动提交 {sub_pages} 页；需人工处理 {manual_pages} 页；"
                        f"价格通过 {len(passed)} 个 / 不通过 {len(failed)} 个（清单见下方）。")
                task["result"] = {
                    "ok": True, "url": page.url,
                    "submitted_pages": sub_pages, "manual_pages": manual_pages,
                    "passed": passed, "failed": failed,
                    "note": note,
                }
                task["done"] = True
        finally:
            try:
                p.stop()
            except Exception:
                pass
    except Exception as e:
        with lock:
            task["error"] = str(e)
            task["done"] = True


@app.route('/api/traffic/start', methods=['POST'])
def api_traffic_start():
    """② 用户点「好了」：启动后台接管任务；立即返回 task_id，前端轮询 /api/traffic/status。"""
    import uuid
    TRAFFIC_STOP.clear()
    TRAFFIC_REJECTED_SPUS.clear()
    TRAFFIC_RECORDED_ONCE.clear()
    TRAFFIC_BLACKLIST.clear()
    n_rej = _traffic_load_rejected_spus()
    n_bl = _traffic_load_blacklist()
    task_id = uuid.uuid4().hex
    with TRAFFIC_TASKS_LOCK:
        TRAFFIC_TASKS[task_id] = {"lock": threading.Lock(), "log": [],
                                  "done": False, "result": None, "error": None}
        if n_rej:
            TRAFFIC_TASKS[task_id]["log"].append(
                f"ℹ️ 已从记录文件预载 {len(TRAFFIC_REJECTED_SPUS)} 个被 Temu 拒绝过的 SPU，本轮直接跳过不再尝试提交")
        if n_bl:
            TRAFFIC_TASKS[task_id]["log"].append(
                f"ℹ️ 已预载黑名单 {len(TRAFFIC_BLACKLIST)} 个已知垃圾（(SPU,站点)，{TRAFFIC_BLACKLIST_TTL_DAYS} 天过期），命中直接跳过不再分析")
    t = threading.Thread(target=_traffic_run, args=(task_id,), daemon=True)
    t.start()
    return jsonify({"ok": True, "task_id": task_id})


@app.route('/api/traffic/stop', methods=['POST'])
def api_traffic_stop():
    """随时停止当前流量加速任务（全局信号，无需 task_id；关掉页面后再打开也能停）。"""
    TRAFFIC_STOP.set()
    return jsonify({"ok": True})


@app.route('/api/traffic/status', methods=['GET'])
def api_traffic_status():
    """轮询接管任务进度：返回实时日志、是否完成、最终结果/错误。"""
    task_id = request.args.get('task_id')
    if not task_id or task_id not in TRAFFIC_TASKS:
        return jsonify({"ok": False, "error": "无效或已过期的任务ID"})
    task = TRAFFIC_TASKS[task_id]
    with task["lock"]:
        return jsonify({"ok": True, "done": task["done"], "error": task["error"],
                        "result": task["result"], "log": list(task["log"])})


@app.route('/api/activity/start', methods=['POST'])
def api_activity_start():
    """启动 Temu 报活动脚本。"""
    resp, code = _start_activity_script("报活动")
    return jsonify(resp), code


@app.route('/api/activity/stop', methods=['POST'])
def api_activity_stop():
    """停止当前报活动任务。"""
    with activity_lock:
        proc = activity_task.get("proc")
        if proc and proc.poll() is None:
            try:
                proc.terminate()
                try:
                    proc.wait(timeout=3)
                except subprocess.TimeoutExpired:
                    proc.kill()
            except Exception as e:
                return jsonify({"success": False, "message": f"停止失败: {e}"}), 500
        activity_task["status"] = "stopped"
        activity_task["completed_at"] = datetime.now().isoformat()
        activity_task["proc"] = None
    return jsonify({"success": True})


@app.route('/api/activity/status')
def api_activity_status():
    """获取报活动任务状态、增量日志和状态文件信息。"""
    with activity_lock:
        state_info = _read_activity_state()

        # 返回未读取过的日志（按 contract 返回字符串数组）
        idx = activity_task.get("log_index", 0)
        all_logs = activity_task.get("log", [])
        logs = []
        for entry in all_logs[idx:]:
            if isinstance(entry, dict):
                logs.append(entry.get("line", ""))
            else:
                logs.append(str(entry))
        activity_task["log_index"] = len(all_logs)

        # 计算运行时长
        elapsed = 0
        if activity_task.get("status") == "running" and activity_task.get("started_at"):
            try:
                elapsed = int((datetime.now() - datetime.fromisoformat(activity_task["started_at"])).total_seconds())
            except Exception:
                pass

        # 状态映射：stopped 对前端显示为 idle
        raw_status = activity_task.get("status", "idle")
        display_status = "idle" if raw_status == "stopped" else raw_status

        # state_info 不存在时返回 None，保持 graceful
        if not state_info:
            state_info = None

        # 等待用户勾选：引擎在 WAIT_USER_SELECT 步骤把候选活动放进 meta.candidates
        waiting_select = bool(state_info) and state_info.get("current_step") == "WAIT_USER_SELECT"
        candidates = (state_info.get("meta", {}) or {}).get("candidates", []) if state_info else []

        return jsonify({
            "status": display_status,
            "started_at": activity_task.get("started_at"),
            "elapsed_sec": elapsed,
            "log": logs,
            "waiting_select": waiting_select,
            "candidates": candidates if waiting_select else [],
            "state_info": {
                "current_step": state_info.get("current_step"),
                "completed_steps": state_info.get("completed_steps", []),
                "errors": state_info.get("errors", []),
                "meta": state_info.get("meta", {}),
            } if state_info else None,
        })


@app.route('/api/activity/select', methods=['POST'])
def api_activity_select():
    """提交用户勾选的活动主题，写入 user_selection.json 供引擎消费。"""
    data = request.get_json(silent=True) or {}
    themes = data.get("themes", [])
    if not isinstance(themes, list) or not themes or not all(isinstance(t, str) for t in themes):
        return jsonify({"success": False, "message": "themes 必须是非空字符串数组"}), 400

    # 引擎必须正在等待勾选，否则拒绝（防止写入后无人消费）
    state_info = _read_activity_state()
    if not state_info or state_info.get("current_step") != "WAIT_USER_SELECT":
        return jsonify({"success": False, "message": "引擎当前不在等待勾选状态（可能已超时或尚未列出活动）"}), 409

    # 校验勾选的活动必须在候选列表中
    valid = {c.get("name") for c in (state_info.get("meta", {}) or {}).get("candidates", [])}
    selected = []
    for t in themes:
        if t in valid and t not in selected:
            selected.append(t)
    if not selected:
        return jsonify({"success": False, "message": "勾选的活动都不在候选列表中，请刷新后重试"}), 400

    payload = {"themes": selected, "time": datetime.now().isoformat()}
    tmp = ACTIVITY_SELECTION_FILE + ".tmp"
    try:
        os.makedirs(os.path.dirname(ACTIVITY_SELECTION_FILE), exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        os.replace(tmp, ACTIVITY_SELECTION_FILE)  # 原子写入，避免引擎读到半截文件
    except Exception as e:
        return jsonify({"success": False, "message": f"写入勾选文件失败: {e}"}), 500

    return jsonify({"success": True, "message": f"已提交 {len(selected)} 个活动，继续执行", "themes": selected})


# ============================================================================
# 商品信息同步（报活动前置：抓取最新商品信息，用户指定页数）
# ============================================================================
ACTIVITY_SYNC_SCRIPT = ACTIVITY_DIR + '/sync_products.py'
ACTIVITY_SYNC_RESULT = ACTIVITY_DIR + '/state/product_sync/latest_sync_result.json'
ACTIVITY_GO_SIGNAL = ACTIVITY_DIR + '/go.signal'  # 用户『好了』信号（与核价同机制）

sync_task = {
    "status": "idle",          # idle | running | completed | error
    "started_at": None,
    "completed_at": None,
    "proc": None,
    "log": [],
    "log_index": 0,
}
sync_lock = threading.Lock()


def _sync_log_reader(proc):
    """后台线程：读取同步脚本 stdout/stderr 写入 sync_task 日志。"""
    def _read_stream(stream, kind):
        try:
            for raw in iter(stream.readline, b""):
                line = None
                for enc in ("utf-8", "gbk", "gb2312"):
                    try:
                        line = raw.decode(enc, errors="strict").rstrip("\r\n")
                        break
                    except Exception:
                        continue
                if line is None:
                    line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
                if not line:
                    continue
                with sync_lock:
                    sync_task["log"].append(line)
        except Exception:
            pass

    try:
        _read_stream(proc.stdout, "out")
        _read_stream(proc.stderr, "err")
    finally:
        rc = proc.wait()
        for t in (proc.stdout, proc.stderr):
            try:
                t.close()
            except Exception:
                pass
        with sync_lock:
            if sync_task["status"] == "running":
                sync_task["status"] = "completed" if rc == 0 else "error"
            sync_task["completed_at"] = datetime.now().isoformat()
            sync_task["proc"] = None


@app.route('/api/activity/sync-products', methods=['POST'])
def api_activity_sync_products():
    """启动商品信息同步（报活动前置：自动开 Edge → 等『好了』→ 抓最新商品信息，可指定页范围）。"""
    data = request.get_json(silent=True) or {}
    start_pg = data.get("start")
    end_pg = data.get("end")
    pages = str(data.get("pages") or "").strip() or "1-5"

    if start_pg is not None and end_pg is not None:
        pages = f"{int(start_pg)}-{int(end_pg)}"

    with sync_lock:
        if sync_task.get("status") == "running" and sync_task.get("proc") and sync_task["proc"].poll() is None:
            return jsonify({"success": False, "message": "商品信息同步已在运行，请等待完成"}), 409
        sync_task["status"] = "running"
        sync_task["started_at"] = datetime.now().isoformat()
        sync_task["completed_at"] = None
        sync_task["log"] = [f"[{datetime.now().strftime('%H:%M:%S')}] 启动: 商品信息同步 (页 {pages})"]
        sync_task["log_index"] = 0

    if not os.path.exists(ACTIVITY_SYNC_SCRIPT):
        with sync_lock:
            sync_task["status"] = "error"
            sync_task["completed_at"] = datetime.now().isoformat()
        return jsonify({"success": False, "message": f"同步脚本不存在: {ACTIVITY_SYNC_SCRIPT}"}), 404

    # 清除上一轮残留的 go.signal，避免旧信号被立即消费
    try:
        if os.path.exists(ACTIVITY_GO_SIGNAL):
            os.remove(ACTIVITY_GO_SIGNAL)
    except Exception as e:
        print(f"[activity sync] 清除 go.signal 失败: {e}", flush=True)

    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env.setdefault("PYTHONIOENCODING", "utf-8")
    try:
        proc = subprocess.Popen(
            [get_python(), ACTIVITY_SYNC_SCRIPT, "--pages", pages],
            cwd=ACTIVITY_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,
            env=env,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    except Exception as e:
        with sync_lock:
            sync_task["status"] = "error"
            sync_task["completed_at"] = datetime.now().isoformat()
            sync_task["proc"] = None
        return jsonify({"success": False, "message": f"启动同步脚本失败: {e}"}), 500

    with sync_lock:
        sync_task["proc"] = proc
    threading.Thread(target=_sync_log_reader, args=(proc,), daemon=True).start()
    return jsonify({"success": True, "message": f"已启动商品信息同步（{pages} 页），请在 Edge 确认上新页面后点『好了』"})


@app.route('/api/activity/sync-signal', methods=['POST'])
def api_activity_sync_signal():
    """创建 go.signal 文件，通知同步脚本用户已在 Edge 确认上新页面。"""
    try:
        os.makedirs(os.path.dirname(ACTIVITY_GO_SIGNAL), exist_ok=True)
        with open(ACTIVITY_GO_SIGNAL, "w", encoding="utf-8") as f:
            f.write("go")
        return jsonify({"ok": True, "msg": "已发送『好了』信号，开始获取商品信息"})
    except Exception as e:
        return jsonify({"error": f"创建 signal 文件失败: {e}"}), 500


@app.route('/api/activity/sync-status')
def api_activity_sync_status():
    """获取商品信息同步状态（增量日志 + 结果汇总）。"""
    with sync_lock:
        idx = sync_task.get("log_index", 0)
        all_logs = sync_task.get("log", [])
        logs = all_logs[idx:]
        sync_task["log_index"] = len(all_logs)

        elapsed = 0
        if sync_task.get("status") == "running" and sync_task.get("started_at"):
            try:
                elapsed = int((datetime.now() - datetime.fromisoformat(sync_task["started_at"])).total_seconds())
            except Exception:
                pass

        result = None
        try:
            if os.path.exists(ACTIVITY_SYNC_RESULT):
                result = json.loads(Path(ACTIVITY_SYNC_RESULT).read_text(encoding="utf-8"))
        except Exception:
            result = None

        return jsonify({
            "status": sync_task.get("status", "idle"),
            "started_at": sync_task.get("started_at"),
            "completed_at": sync_task.get("completed_at"),
            "elapsed_sec": elapsed,
            "log": logs,
            "result": result,
        })


@app.route('/api/activity/sync-download')
def api_activity_sync_download():
    """下载商品信息同步生成的 Excel。"""
    filename = request.args.get("file", "").strip()
    if not filename:
        return jsonify({"error": "请指定文件名"}), 400
    filename = os.path.basename(filename)
    if not filename.endswith(".xlsx"):
        return jsonify({"error": "仅支持 .xlsx 文件"}), 400
    path = Path(ACTIVITY_DIR) / "state" / "product_sync" / filename
    if not path.exists():
        return jsonify({"error": "文件不存在"}), 404
    return send_file(str(path), as_attachment=True, download_name=filename)


@app.route('/retail_price')
def retail_price_page():
    """Temu 建议零售价填写页面。"""
    return send_file(str(Path(__file__).parent / 'retail_price.html'))


@app.route('/api/retail_price/start', methods=['POST'])
def api_retail_price_start():
    """启动 Temu 建议零售价填写脚本。"""
    resp, code = _start_retail_price_script("建议零售价填写")
    return jsonify(resp), code


@app.route('/api/retail_price/start_diagnose', methods=['POST'])
def api_retail_price_start_diagnose():
    """启动 Temu 建议零售价诊断：仅 dump 抽屉结构，不填写/不提交。复用「👌 好了」信号触发。"""
    resp, code = _start_retail_price_script("建议零售价诊断", diagnose=True)
    return jsonify(resp), code


@app.route('/api/retail_price/stop', methods=['POST'])
def api_retail_price_stop():
    """停止当前建议零售价填写任务。"""
    with retail_price_lock:
        proc = retail_price_task.get("proc")
        if proc and proc.poll() is None:
            try:
                proc.terminate()
                try:
                    proc.wait(timeout=3)
                except subprocess.TimeoutExpired:
                    proc.kill()
            except Exception as e:
                retail_price_task["status"] = "error"
                retail_price_task["task_label"] = "停止失败"
                retail_price_task["proc"] = None
                retail_price_task["log"].append({"line": f"[{datetime.now().strftime('%H:%M:%S')}] 停止失败: {e}", "kind": "error"})
                return jsonify({"error": f"停止失败: {e}"}), 500
        retail_price_task["status"] = "stopped"
        retail_price_task["task_label"] = "已停止"
        retail_price_task["completed_at"] = datetime.now().isoformat()
        retail_price_task["proc"] = None
    return jsonify({"ok": True, "msg": "已停止"})


@app.route('/api/retail_price/status')
def api_retail_price_status():
    """获取建议零售价填写任务状态与增量日志。"""
    with retail_price_lock:
        idx = retail_price_task.get("log_index", 0)
        all_logs = retail_price_task.get("log", [])
        logs = all_logs[idx:]
        retail_price_task["log_index"] = len(all_logs)

        elapsed = 0
        if retail_price_task.get("status") == "running" and retail_price_task.get("started_at"):
            try:
                elapsed = int((datetime.now() - datetime.fromisoformat(retail_price_task["started_at"])).total_seconds())
            except Exception:
                pass

        raw_status = retail_price_task.get("status", "idle")
        display_status = "idle" if raw_status == "stopped" else raw_status

        return jsonify({
            "status": display_status,
            "task_label": retail_price_task.get("task_label", ""),
            "started_at": retail_price_task.get("started_at"),
            "completed_at": retail_price_task.get("completed_at"),
            "elapsed_sec": elapsed,
            "log": logs,
        })


@app.route('/api/retail_price/signal', methods=['POST'])
def api_retail_price_signal():
    """创建 go.signal 文件，通知脚本用户已准备好。"""
    signal_path = RETAIL_PRICE_DIR / "go.signal"
    try:
        signal_path.write_text("", encoding="utf-8")
        return jsonify({"ok": True, "msg": "已发送'好了'信号"})
    except Exception as e:
        return jsonify({"error": f"创建信号文件失败: {e}"}), 500


# ============================================================================
# 胚衣制作（素材库）
# ============================================================================

@app.route('/peiyi')
def peiyi_page():
    """胚衣制作页面：分类上传素材，自动处理为 1340×1785 @ 72DPI。"""
    return send_file(str(Path(__file__).parent / 'peiyi.html'))


def _peiyi_max_index(dest_dir, prefix: str) -> int:
    """返回该分类文件夹中已存在的最大序号（黑W12.jpg -> 12），用于按进入顺序命名。"""
    max_idx = 0
    plen = len(prefix)
    if dest_dir.exists():
        for fn in os.listdir(dest_dir):
            low = fn.lower()
            if low.endswith('.jpg') and low.startswith(prefix.lower()):
                num_part = fn[plen:-4]
                if num_part.isdigit():
                    max_idx = max(max_idx, int(num_part))
    return max_idx


def _peiyi_process_image(src_path: str, category: str, dest_path: str):
    """读取源图，强制拉伸到 1340×1785，合成底色，以 72 DPI 存为 JPG。"""
    from PIL import Image, ImageOps
    bg = PEIYI_BG.get(category, (255, 255, 255))
    with Image.open(src_path) as im:
        try:
            im = ImageOps.exif_transpose(im)
        except Exception:
            pass
        if im.mode in ('RGBA', 'LA') or (im.mode == 'P' and 'transparency' in im.info):
            base = Image.new('RGB', im.size, bg)
            if im.mode == 'P':
                im = im.convert('RGBA')
            base.paste(im, (0, 0), im)
            im = base
        else:
            im = im.convert('RGB')
        # 强制拉伸铺满目标尺寸（用户确认：允许变形）
        im = im.resize(PEIYI_SIZE, Image.LANCZOS)
        im.save(dest_path, 'JPEG', dpi=PEIYI_DPI, quality=92)


@app.route('/api/peiyi/upload', methods=['POST'])
def api_peiyi_upload():
    """批量上传素材：按 category 自动处理并存入对应文件夹。"""
    category = request.form.get('category', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': f'未知分类: {category}'}), 400
    files = request.files.getlist('files')
    if not files:
        return jsonify({'ok': False, 'error': '未收到文件'}), 400

    dest_dir = peiyi_dirs[category]
    dest_dir.mkdir(parents=True, exist_ok=True)

    # 按进入顺序命名：颜色 + 面 + 序号（黑W1, 黑W2 ...；白B1 ...）
    prefix = (category[1] if len(category) > 1 else '') + category[0]
    next_idx = _peiyi_max_index(dest_dir, prefix) + 1

    results = []
    ok_count = 0
    for f in files:
        orig = f.filename or 'material'
        if not orig.lower().endswith(PEIYI_ALLOWED_EXT):
            results.append({'file': orig, 'ok': False, 'error': '不支持的图片格式'})
            continue
        tmp = dest_dir / (f'_tmp_{datetime.now().strftime("%H%M%S%f")}_{os.path.splitext(orig)[1]}')
        f.save(str(tmp))
        try:
            out_name = f'{prefix}{next_idx}.jpg'
            next_idx += 1
            out_path = dest_dir / out_name
            # 保险：序号理论上递增不会撞，仍做兜底顺延
            while out_path.exists():
                out_name = f'{prefix}{next_idx}.jpg'
                next_idx += 1
                out_path = dest_dir / out_name
            _peiyi_process_image(str(tmp), category, str(out_path))
            results.append({'file': orig, 'ok': True, 'saved': out_path.name})
            ok_count += 1
        except Exception as e:
            results.append({'file': orig, 'ok': False, 'error': str(e)})
        finally:
            try:
                os.remove(str(tmp))
            except OSError:
                pass

    return jsonify({'ok': True, 'category': category, 'saved': ok_count, 'results': results})


def _peiyi_read_meta(dest_dir, name):
    """读取与图片同名的 .meta.json 侧车，返回5个参数字典（含可选 bw 第二套）；
    不存在/损坏返回 None。

    双组参数约定（仅正面 W白/W黑 需要）：
      - 顶层 width/height/rotation/highest_y/center_x = ① 单面款（只有W贴图）
      - "bw" 子块同名五参                         = ② 双面款（有W+B贴图）
    """
    stem, _ = os.path.splitext(name)
    mp = dest_dir / (stem + '.meta.json')
    if not mp.exists():
        return None
    try:
        data = json.loads(mp.read_text(encoding='utf-8'))
        out = {k: data.get(k) for k in PEIYI_META_KEYS}
        bw = data.get('bw')
        if isinstance(bw, dict):
            out['bw'] = {k: bw.get(k) for k in PEIYI_META_KEYS}
        return out
    except Exception:
        return None


def _meta_filled(meta):
    """判断一组5个贴图参数是否全部已填写（非 None、非空字符串）。"""
    if not isinstance(meta, dict):
        return False
    return all(meta.get(k) not in (None, '') for k, _, _ in PEIYI_META_FIELDS)


@app.route('/api/peiyi/list')
def api_peiyi_list():
    """列出各分类已存素材（用于画廊预览），含每张图的贴图参数 meta。"""
    category = request.args.get('category', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category and category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    cats = [category] if category else list(peiyi_dirs.keys())
    out = {}
    for c in cats:
        d = peiyi_dirs[c]
        items = []
        if d.exists():
            for fn in os.listdir(d):
                # 只把“原图”纳入画廊：① 允许的图像格式；② 非临时文件；③ 非遮罩侧车
                lfn = fn.lower()
                ext = os.path.splitext(lfn)[1]
                if ext not in PEIYI_ALLOWED_EXT:
                    continue
                if fn.startswith('_tmp_'):
                    continue
                if any(lfn.endswith(s) for s in PEIYI_MASK_SUFFIXES):
                    continue
                fp = d / fn
                try:
                    st = fp.stat()
                    meta = _peiyi_read_meta(d, fn) or {}
                    # 保证顶层与 bw 第二套字段存在（None 表示未填），便于前端渲染空输入框
                    for k, _, _ in PEIYI_META_FIELDS:
                        meta.setdefault(k, None)
                    if not isinstance(meta.get('bw'), dict):
                        meta['bw'] = {k: None for k, _, _ in PEIYI_META_FIELDS}
                    else:
                        for k, _, _ in PEIYI_META_FIELDS:
                            meta['bw'].setdefault(k, None)
                    # 只有五项全部填写才算完成；部分填写仍视为待填，排在最前
                    unfilled = not _meta_filled(meta)
                    # 双面款(W+B)第二组参数缺失标记（所有正面 W* 分类需要两组；B面只一组）
                    bw_missing = False
                    if category and category[0] == 'W':
                        bw_missing = not _meta_filled(meta.get('bw'))
                    # 遮罩状态（body_mask / occluder_mask / occluder / parse）
                    stem, _ = os.path.splitext(fn)
                    occ_mask_path = d / (stem + '_occluder_mask.png')
                    occ_path = d / (stem + '_occluder.png')
                    occ_px = None
                    has_mask = False
                    if occ_mask_path.exists():
                        has_mask = True
                        try:
                            occ_arr = np.array(Image.open(str(occ_mask_path)))
                            occ_px = int((occ_arr > 0).sum())
                        except Exception:
                            occ_px = None
                    elif occ_path.exists():
                        has_mask = True
                    mask_urls = {}
                    for suffix, key in [
                        ('_occluder.png', 'occluder'),
                        ('_occluder_mask.png', 'occluder_mask'),
                        ('_body_mask.png', 'body_mask'),
                        ('_parse.png', 'parse'),
                    ]:
                        mp = d / (stem + suffix)
                        if mp.exists():
                            mask_urls[key] = (f'/api/peiyi/material/{urllib.parse.quote(c)}/{urllib.parse.quote(mp.name)}'
                                              + _peiyi_cat_qs(_cat))
                    # 最新一版遮罩评分（用于图片墙角标），读不到则 None
                    score_info = _peiyi_latest_score(d, stem)
                    items.append({
                        'name': fn,
                        'size': st.st_size,
                        # URL 编码文件名/分类，避免中文路径导致浏览器无法加载图片
                        'url': (f'/api/peiyi/material/{urllib.parse.quote(c)}/{urllib.parse.quote(fn)}'
                                + _peiyi_cat_qs(_cat)),
                        'modified': datetime.fromtimestamp(st.st_mtime).isoformat(),
                        'meta': meta,
                        'unfilled': unfilled,
                        'bw_missing': bw_missing,
                        'has_mask': has_mask,
                        'occluder_px': occ_px,
                        'mask_urls': mask_urls,
                        'score': score_info,
                        '_mtime': st.st_mtime,
                    })
                except OSError:
                    pass
            # 排序：① 未填写五项数据的排最前；② 同组内按修改时间倒序（后进入排前面）
            items.sort(key=lambda e: (0 if e['unfilled'] else 1, -e['_mtime']))
            for e in items:
                e.pop('_mtime', None)
        out[c] = items
    return jsonify({'ok': True, 'categories': out})


def _peiyi_latest_score(category_dir, stem):
    """读取某胚衣最新一版的评分（来自 _mask_versions/<stem>/latest.txt → vNNN/score.json）。
    返回 dict 或 None（尚无版本/读取失败）。任何异常都吞掉，绝不影响列表页。"""
    try:
        vroot = Path(category_dir) / "_mask_versions" / stem
        latest_f = vroot / "latest.txt"
        if not latest_f.exists():
            return None
        latest = latest_f.read_text(encoding="utf-8").strip()
        sf = vroot / latest / "score.json"
        if not sf.exists():
            return None
        import json as _json
        data = _json.loads(sf.read_text(encoding="utf-8"))
        m = data.get("metrics", {}) or {}
        return {
            "version": data.get("version", latest),
            "timestamp": data.get("timestamp", ""),
            "algorithm_version": data.get("algorithm_version", ""),
            "score": data.get("score"),
            "occ_ratio": m.get("occ_ratio"),
            "body_coverage": m.get("body_coverage"),
            "is_person": m.get("is_person"),
            "flat_lay": m.get("flat_lay"),
            "flags": data.get("flags", []) or [],
        }
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 抽绳遮罩工具服务 (8777) —— 由「卫衣·胚衣制作」页面生命周期控制：
# 页面加载时 ensure 启动，页面关闭时 stop。不再由 bridge 启动脚本拉起。
# ---------------------------------------------------------------------------
_DS_PROC = None
_DS_PY = r"C:\Users\Administrator\AppData\Local\Programs\Python\Python311\python.exe"
_DS_SERVER = r"D:\Semems Hoodie\drawstring_tool\drawstring_server.py"
_DS_PORT = 8777


def _ds_port_listening():
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        return s.connect_ex(("127.0.0.1", _DS_PORT)) == 0
    finally:
        s.close()


def _ds_kill_by_port():
    """兜底：杀掉 8777 端口上任何残留监听进程（含非本进程启动的）。"""
    killed = []
    try:
        out = subprocess.check_output(["netstat", "-ano"],
                                      stderr=subprocess.DEVNULL).decode("gbk", "ignore")
        for line in out.splitlines():
            if (":%d" % _DS_PORT) in line and "LISTENING" in line:
                pid = line.split()[-1]
                if pid.isdigit():
                    try:
                        subprocess.call(["taskkill", "/PID", pid, "/F"],
                                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                        killed.append(int(pid))
                    except Exception:
                        pass
    except Exception:
        pass
    return killed


@app.route("/api/drawstring/ensure", methods=["POST"])
def api_drawstring_ensure():
    """确保抽绳工具服务在运行；未运行则拉起。"""
    global _DS_PROC
    if _ds_port_listening():
        return jsonify({"ok": True, "already": True})
    try:
        _DS_PROC = subprocess.Popen(
            [_DS_PY, _DS_SERVER],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            creationflags=0x00000008,  # DETACHED_PROCESS：不弹控制台窗口
        )
        # 轮询等待服务起来（最多 ~4s）
        import time
        for _ in range(40):
            if _ds_port_listening():
                return jsonify({"ok": True, "started": True, "pid": _DS_PROC.pid})
            time.sleep(0.1)
        return jsonify({"ok": False, "error": "started but 8777 not listening"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/drawstring/stop", methods=["POST"])
def api_drawstring_stop():
    """关闭抽绳工具服务（页面关闭时调用）。"""
    global _DS_PROC
    killed = []
    if _DS_PROC is not None and _DS_PROC.poll() is None:
        try:
            _DS_PROC.kill()
            killed.append(_DS_PROC.pid)
        except Exception:
            pass
    _DS_PROC = None
    killed += _ds_kill_by_port()  # 兜底清理端口上残留
    return jsonify({"ok": True, "killed": killed})


@app.route('/api/peiyi/scores')
def api_peiyi_scores():
    """汇总所有胚衣最新一版的评分（低分可一眼标红）。
    评分在“生成遮罩”时写入 _mask_versions/<stem>/vNNN/score.json。"""
    category = request.args.get('category', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category and category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    cats = [category] if category else list(peiyi_dirs.keys())
    rows = []
    for c in cats:
        d = peiyi_dirs[c]
        if not d.exists():
            continue
        for fn in os.listdir(d):
            lfn = fn.lower()
            ext = os.path.splitext(lfn)[1]
            if ext not in PEIYI_ALLOWED_EXT:
                continue
            if fn.startswith('_tmp_'):
                continue
            if any(lfn.endswith(s) for s in PEIYI_MASK_SUFFIXES):
                continue
            stem, _ = os.path.splitext(fn)
            info = _peiyi_latest_score(d, stem)
            row = {'category': c, 'name': fn, 'stem': stem, 'has_score': info is not None}
            if info:
                row.update(info)
            rows.append(row)
    # 排序：有分数按分数升序（低分排最前，问题胚衣一眼可见），无分数排最后
    def _key(r):
        sc = r.get('score')
        return (0, sc) if (r.get('has_score') and isinstance(sc, (int, float))) else (1, 0)
    rows.sort(key=_key)
    return jsonify({'ok': True, 'rows': rows})


@app.route('/api/peiyi/material/<category>/<path:filename>')
def api_peiyi_material(category, filename):
    """返回已存素材（原图 / 遮罩侧车）。

    遮罩文件会被“重新生成遮罩”覆盖更新，因此此处禁用浏览器缓存
    （Cache-Control: no-store, no-cache），确保预览/画廊实时反映最新内容；
    并按真实扩展名返回正确 MIME（PNG 遮罩不再被当成 JPEG，避免渲染异常）。
    """
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None or category not in peiyi_dirs:
        abort(404)
    safe = os.path.basename(filename)
    fp = peiyi_dirs[category] / safe
    if not fp.exists():
        abort(404)
    ext = os.path.splitext(safe)[1].lower().lstrip('.')
    mime_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                'png': 'image/png', 'webp': 'image/webp', 'bmp': 'image/bmp'}
    mime = mime_map.get(ext, 'image/jpeg')
    resp = send_file(str(fp), mimetype=mime, max_age=0)
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


# 版本文件名后缀 → 预览键（与 _mask_versions/<stem>/vNNN/ 内的侧车一致）
_PEIYI_VERSION_FILE_KEYS = [
    ('_occluder.png', 'occluder'),
    ('_occluder_mask.png', 'occluder_mask'),
    ('_body_mask.png', 'body_mask'),
    ('_parse.png', 'parse'),
    ('_alpha.png', 'alpha'),
]


@app.route('/api/peiyi/versions/<category>/<stem>')
def api_peiyi_versions(category, stem):
    """列出某胚衣的所有遮罩版本（每个版本的分数/时间/指标/各层遮罩图URL/是否当前）。
    数据来自 _mask_versions/<stem>/vNNN/ + latest.txt。"""
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    stem = os.path.basename(stem)
    d = peiyi_dirs[category]
    vroot = d / "_mask_versions" / stem
    current = None
    versions = []
    if vroot.exists():
        latest_f = vroot / "latest.txt"
        if latest_f.exists():
            try:
                current = latest_f.read_text(encoding="utf-8").strip()
            except Exception:
                current = None
        for vd in sorted(vroot.iterdir()):
            if not (vd.is_dir() and vd.name.startswith('v')):
                continue
            info = {}
            sf = vd / "score.json"
            if sf.exists():
                try:
                    info = json.loads(sf.read_text(encoding="utf-8"))
                except Exception:
                    info = {}
            urls = {}
            for suffix, key in _PEIYI_VERSION_FILE_KEYS:
                if (vd / (stem + suffix)).exists():
                    urls[key] = (f'/api/peiyi/version_file/{urllib.parse.quote(category)}'
                                 f'/{urllib.parse.quote(stem)}/{urllib.parse.quote(vd.name)}'
                                 f'/{urllib.parse.quote(stem + suffix)}'
                                 + _peiyi_cat_qs(_cat))
            versions.append({
                'version': vd.name,
                'score': info.get('score'),
                'timestamp': info.get('timestamp', ''),
                'algorithm_version': info.get('algorithm_version', ''),
                'flags': info.get('flags', []) or [],
                'metrics': info.get('metrics', {}) or {},
                'is_current': (vd.name == current),
                'urls': urls,
            })
    return jsonify({'ok': True, 'category': category, 'stem': stem,
                    'current': current, 'versions': versions})


@app.route('/api/peiyi/version_file/<category>/<stem>/<version>/<path:filename>')
def api_peiyi_version_file(category, stem, version, filename):
    """返回某胚衣某版本目录里的遮罩图片（禁用缓存，按真实扩展名给 MIME）。"""
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None or category not in peiyi_dirs:
        abort(404)
    stem = os.path.basename(stem)
    version = os.path.basename(version)
    safe = os.path.basename(filename)
    fp = peiyi_dirs[category] / "_mask_versions" / stem / version / safe
    if not fp.exists():
        abort(404)
    ext = os.path.splitext(safe)[1].lower().lstrip('.')
    mime_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                'png': 'image/png', 'webp': 'image/webp', 'bmp': 'image/bmp'}
    mime = mime_map.get(ext, 'image/png')
    resp = send_file(str(fp), mimetype=mime, max_age=0)
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/api/peiyi/use_version', methods=['POST'])
def api_peiyi_use_version():
    """把选中版本的遮罩文件复制回素材库标准路径（=退回/切换到该版本），并更新 latest.txt。
    生产贴图读的是标准路径，因此这一步立即决定以后用哪一版。"""
    data = request.get_json(force=True, silent=True) or {}
    category = data.get('category', '')
    stem = os.path.basename(data.get('stem', ''))
    version = os.path.basename(data.get('version', ''))
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    if not stem or not version:
        return jsonify({'ok': False, 'error': '缺少 stem/version'}), 400
    d = peiyi_dirs[category]
    vdir = d / "_mask_versions" / stem / version
    if not vdir.exists():
        return jsonify({'ok': False, 'error': '版本不存在'}), 404
    copied = []
    for suffix, _ in _PEIYI_VERSION_FILE_KEYS:
        src = vdir / (stem + suffix)
        if src.exists():
            try:
                shutil.copy2(str(src), str(d / (stem + suffix)))
                copied.append(suffix)
            except Exception as e:
                return jsonify({'ok': False, 'error': f'复制失败 {suffix}: {e}'}), 500
    try:
        (d / "_mask_versions" / stem / "latest.txt").write_text(version, encoding="utf-8")
    except Exception as e:
        return jsonify({'ok': False, 'error': f'更新 latest.txt 失败: {e}'}), 500
    return jsonify({'ok': True, 'category': category, 'stem': stem,
                    'version': version, 'copied': copied})


def _explorer_select_file(path):
    """在资源管理器中打开 path 所在文件夹并【选中】该文件（Windows 最可靠方式）。

    为什么不用 explorer /select,：该命令对带空格/中文的路径解析极不稳定，
    一旦解析失败就会回退到“文档库”。这里改用系统底层
    SHOpenFolderAndSelectItems（专业软件通用做法），彻底避开该坑。
    返回 True 表示已成功触发；全部失败时返回 False。
    """
    import os
    import ctypes
    from ctypes import wintypes, POINTER, byref, c_void_p, c_uint
    path = os.path.normpath(str(path))
    folder = os.path.dirname(path)
    # 方法1：ctypes 调用 Shell 接口精准选中文件（无空格/中文坑）
    try:
        shell32 = ctypes.windll.shell32
        ole32 = ctypes.windll.ole32
        # 必须先初始化 COM（STA），否则 Shell 接口会报“尚未调用 CoInitialize”
        init_hr = ole32.CoInitialize(None)
        try:
            shell32.SHParseDisplayName.argtypes = [
                wintypes.LPCWSTR, c_void_p, POINTER(c_void_p), c_uint, POINTER(c_uint)
            ]
            shell32.SHParseDisplayName.restype = ctypes.HRESULT
            shell32.SHOpenFolderAndSelectItems.argtypes = [
                c_void_p, c_uint, POINTER(c_void_p), c_uint
            ]
            shell32.SHOpenFolderAndSelectItems.restype = ctypes.HRESULT
            pidl = c_void_p()
            attrs = c_uint()
            hr = shell32.SHParseDisplayName(path, None, byref(pidl), 0, byref(attrs))
            if hr == 0 and pidl:
                try:
                    shell32.SHOpenFolderAndSelectItems(pidl, 0, None, 0)
                    return True
                finally:
                    ole32.CoTaskMemFree(pidl)
        finally:
            # 仅有本函数成功初始化时才反初始化，避免误关别的模块已初始化的 COM
            if init_hr == 0:
                ole32.CoUninitialize()
    except Exception:
        pass
    # 方法2：兜底 explorer /select,（带引号，正确处理空格/中文）
    try:
        subprocess.Popen(f'explorer.exe /select,"{path}"', shell=True)
        return True
    except Exception:
        pass
    # 方法3：兜底只打开正确文件夹（至少不会跑到“文档”）
    try:
        os.startfile(folder)
        return True
    except Exception:
        pass
    try:
        subprocess.Popen(f'explorer.exe "{folder}"', shell=True)
        return True
    except Exception:
        return False


@app.route('/api/peiyi/open', methods=['POST'])
def api_peiyi_open():
    """在文件资源管理器中打开该素材所在文件夹并选中该文件（仅本机/localhost 生效）。"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    fp = peiyi_dirs[category] / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404
    try:
        ok = _explorer_select_file(str(fp))
        return jsonify({'ok': ok, 'path': str(fp)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/peiyi/delete', methods=['POST'])
def api_peiyi_delete():
    """删除某个已存素材。"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    fp = peiyi_dirs[category] / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404
    try:
        os.remove(str(fp))
        # 同时删除该素材的遮罩/参数侧车文件
        stem, _ = os.path.splitext(safe)
        for suffix in ['.meta.json', '_occluder.png', '_occluder_mask.png', '_body_mask.png', '_parse.png', '_alpha.png']:
            try:
                (fp.parent / (stem + suffix)).unlink(missing_ok=True)
            except Exception:
                pass
        return jsonify({'ok': True, 'msg': f'{safe} 已删除'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/peiyi/reindex', methods=['POST'])
def api_peiyi_reindex():
    """把某分类文件夹内所有图片按进入顺序（修改时间）重新编号为 黑W1, 黑W2 ...。
    用于手动拖入、尚未按规则命名的图片。两遍重命名避免同名冲突。"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    d = peiyi_dirs[category]
    if not d.exists():
        return jsonify({'ok': True, 'renamed': 0, 'msg': '空文件夹'})
    prefix = (category[1] if len(category) > 1 else '') + category[0]

    files = [fn for fn in os.listdir(d)
             if fn.lower().endswith('.jpg') and not fn.startswith('_tmp_')]
    files.sort(key=lambda fn: os.path.getmtime(str(d / fn)))

    # 第一遍：全部移到临时名，腾出目标名
    tmp_map = []
    for i, fn in enumerate(files):
        tmp = f'_re_{i}_{fn}'
        os.rename(str(d / fn), str(d / tmp))
        tmp_map.append((tmp, i + 1))
    # 第二遍：临时名 -> 黑W1, 黑W2 ...
    renamed = 0
    for tmp, idx in tmp_map:
        new_name = f'{prefix}{idx}.jpg'
        dest = d / new_name
        n = idx
        while dest.exists():
            n += 1
            dest = d / f'{prefix}{n}.jpg'
        os.rename(str(d / tmp), str(dest))
        renamed += 1
    return jsonify({'ok': True, 'renamed': renamed, 'prefix': prefix,
                    'msg': f'已重新编号为 {prefix}1..{prefix}{renamed}'})


@app.route('/api/peiyi/meta', methods=['POST'])
def api_peiyi_meta():
    """保存单张素材的贴图参数到同名 .meta.json 侧车。

    支持双组（仅正面 W白/W黑 需要）：
      - 顶层 width/height/rotation/highest_y/center_x = ① 单面款（只有W贴图）
      - payload 中的 "bw" 对象同名五参               = ② 双面款（有W+B贴图）
    合并已有文件，避免只保存一组时清掉另一组。
    """
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    d = peiyi_dirs[category]
    if not (d / safe).exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404
    stem, _ = os.path.splitext(safe)
    mp = d / (stem + '.meta.json')
    # 合并已有文件，避免只写部分字段时清掉另一套
    existing = {}
    if mp.exists():
        try:
            existing = json.loads(mp.read_text(encoding='utf-8'))
        except Exception:
            existing = {}

    def _num(v):
        """空值/空字符串存为 None；有值则转 float。"""
        if v is None or v == '':
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    meta = {}
    for k, _, _ in PEIYI_META_FIELDS:
        # payload 中带该字段则按新值存（含空值=清空），未带则保留已有值
        meta[k] = _num(data.get(k)) if k in data else existing.get(k)
    # 第二组（双面款 W+B）：来自 payload 的 bw 对象；payload 未带则保留原 bw
    bw_in = data.get('bw')
    if isinstance(bw_in, dict):
        bw = {}
        for k, _, _ in PEIYI_META_FIELDS:
            bw[k] = _num(bw_in.get(k)) if k in bw_in else existing.get('bw', {}).get(k)
        meta['bw'] = bw
    elif isinstance(existing.get('bw'), dict):
        meta['bw'] = existing['bw']
    mp.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
    return jsonify({'ok': True, 'meta': meta})


@app.route('/api/peiyi/mask', methods=['POST'])
def api_peiyi_mask():
    """为单张胚衣素材生成三层遮罩 + _tpl 扭曲素材。

    返回 JSON 在原有 masks 基础上新增 tpl 字段（_tpl 生成状态）。
    """
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    d = peiyi_dirs[category]
    fp = d / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404
    # 隔离到子进程执行（cv2/OpenMP 偶发崩溃会拖垮主服务，故必须隔离）
    try:
        MOCKUP_OUT.mkdir(parents=True, exist_ok=True)
        env = _single_thread_env(os.environ)
        env["PYTHONPATH"] = f"{ZCODE_PROJECT};{PY_PACKAGES}"
        out_log = MOCKUP_OUT / "_mask_stdout.log"
        err_log = MOCKUP_OUT / "_mask_stderr.log"
        startupinfo = None
        if os.name == "nt":
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE
        with open(out_log, "w", encoding="utf-8", errors="replace") as of, \
             open(err_log, "w", encoding="utf-8", errors="replace") as ef:
            r = subprocess.run(
                [str(MOCKUP_PY), str(ZCODE_PROJECT / "_peiyi_worker.py"), "mask", str(fp), category,
                 str(out_log) + ".json"],
                cwd=str(ZCODE_PROJECT), env=env,
                stdin=subprocess.DEVNULL, stdout=of, stderr=ef,
                timeout=600,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
                startupinfo=startupinfo,
            )
        raw = out_log.read_text(encoding="utf-8", errors="replace").strip()
        err = err_log.read_text(encoding="utf-8", errors="replace").strip()
        res_path = str(out_log) + ".json"
        if r.returncode != 0 or not Path(res_path).exists():
            return jsonify({'ok': False,
                            'error': f'subprocess rc={r.returncode}: {err[-800:]}',
                            'trace': raw[-800:]}), 500
        res = json.loads(Path(res_path).read_text(encoding="utf-8", errors="replace"))
        return jsonify(res), (200 if res.get('ok') else 500)
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}', 'trace': _tb.format_exc()[-1500:]}), 500


@app.route('/api/peiyi/correct_preview', methods=['POST'])
def api_peiyi_correct_preview():
    """预览校正效果（不保存，返回临时区状态）"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')
    click_x = data.get('x')
    click_y = data.get('y')
    mode = data.get('mode', 'add_occ')

    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    if click_x is None or click_y is None:
        return jsonify({'ok': False, 'error': '缺少点击坐标 x, y'}), 400
    if mode not in ('add_occ', 'remove_occ', 'add_body'):
        return jsonify({'ok': False, 'error': f'未知模式: {mode}'}), 400

    safe = os.path.basename(name)
    d = peiyi_dirs[category]
    fp = d / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404

    try:
        import peiyi_correct
        result = peiyi_correct.preview_correction(str(fp), click_x, click_y, mode=mode)
        if result.get('ok'):
            return jsonify(result)
        return jsonify(result), 400
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/api/peiyi/correct_confirm', methods=['POST'])
def api_peiyi_correct_confirm():
    """确认临时遮罩并归档为新版本"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')

    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    d = peiyi_dirs[category]
    fp = d / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404

    try:
        import peiyi_correct
        result = peiyi_correct.confirm_correction(str(fp))
        if result.get('ok'):
            return jsonify(result)
        return jsonify(result), 400
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/api/peiyi/correct_cancel', methods=['POST'])
def api_peiyi_correct_cancel():
    """放弃临时修改"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')

    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    d = peiyi_dirs[category]
    fp = d / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404

    try:
        import peiyi_correct
        result = peiyi_correct.cancel_correction(str(fp))
        return jsonify(result)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/api/peiyi/correct_check', methods=['POST'])
def api_peiyi_correct_check():
    """检查是否有未确认的临时修改"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')

    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    d = peiyi_dirs[category]
    fp = d / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404

    try:
        import peiyi_correct
        result = peiyi_correct.check_working_status(str(fp))
        return jsonify(result)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/api/peiyi/working_file/<category>/<stem>/<path:filename>')
def api_peiyi_working_file(category, stem, filename):
    """提供 _working 临时目录的预览图"""
    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(filename)
    fp = peiyi_dirs[category] / "_mask_versions" / stem / "_working" / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404
    return send_file(str(fp))


@app.route('/api/peiyi/delete_version', methods=['POST'])
def api_peiyi_delete_version():
    """删除指定版本（不能删除当前正在使用的版本）"""
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    stem = data.get('stem', '')
    version = data.get('version', '')

    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    if not stem or not version:
        return jsonify({'ok': False, 'error': '缺少 stem 或 version'}), 400

    d = peiyi_dirs[category]
    try:
        import peiyi_correct
        result = peiyi_correct.delete_version(d, stem, version)
        if result.get('ok'):
            return jsonify(result)
        return jsonify(result), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/api/peiyi/import_manual', methods=['POST'])
def api_peiyi_import_manual():
    """导入 PS 手动遮罩，与 AI 遮罩合并。

    POST JSON: { "category": "W白", "name": "白W2.jpg" }
    手动遮罩文件位置: 素材目录/白W2_manual.png 或 _mask_versions/白W2/白W2.png
    """
    data = request.get_json(silent=True) or {}
    category = data.get('category', '')
    name = data.get('name', '')

    _cat, peiyi_dirs = _peiyi_request_dirs()
    if peiyi_dirs is None:
        return jsonify({'ok': False, 'error': f'未知品类: {_cat}'}), 400
    if category not in peiyi_dirs:
        return jsonify({'ok': False, 'error': '未知分类'}), 400
    safe = os.path.basename(name)
    d = peiyi_dirs[category]
    fp = d / safe
    if not fp.exists():
        return jsonify({'ok': False, 'error': '文件不存在'}), 404

    try:
        import peiyi_correct
        result = peiyi_correct.import_manual_mask(str(fp))
        if result.get('ok'):
            return jsonify(result)
        return jsonify(result), 400
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500


# ============================================================================
# 贴图（AI 去背贴图）：自动按胚衣数据 + 遮罩 + 扭曲精准贴入
# ============================================================================
def _resolve_peiyi_embryo(category, name, dirs=None):
    """素材库图片路径、款名 stem、衫色（按分类名含 黑/白 推断）。dirs 缺省 = wb 四大分类。"""
    dirs = dirs or PEIYI_CATEGORIES
    safe = os.path.basename(name)
    fp = dirs[category] / safe
    stem = fp.stem
    color = "black" if "黑" in category else "white"
    return fp, stem, color


def _find_category_for_stem(stem, dirs=None):
    """按款名 stem 在四大分类里反查 category + 文件名。dirs 缺省 = wb 四大分类。"""
    dirs = dirs or PEIYI_CATEGORIES
    for cat, d in dirs.items():
        if not d.exists():
            continue
        cand = d / f"{stem}.jpg"
        if cand.exists():
            return cat, cand.name
        cand = d / stem
        if cand.exists():
            return cat, cand.name
    return None, None


def _load_presets(cat=None):
    """同步 CSV→presets.json（若 CSV 更新）并读取 templates。"""
    try:
        sys.path.insert(0, str(KIMI_SCRIPTS_DIR))
        import sync_presets_from_csv
        sync_presets_from_csv.sync_if_stale()
    except Exception:
        pass
    p = white_t_presets_for(cat)
    try:
        return json.loads(p.read_text(encoding="utf-8")).get("templates", {})
    except Exception:
        return {}


def _preset_key_for_stem(stem, presets):
    if stem in presets:
        return stem
    for k, v in presets.items():
        if Path(v.get("path", "")).stem == stem:
            return k
    return None


def _embryo_fields(category, name, presets, dirs=None):
    """读取5个贴图字段：素材库 .meta.json 优先，CSV→presets 兜底。

    注意：meta 未填写的字段必须显式视为「缺失」（None），不能用 670 这类
    魔法默认值占位，否则会挡住 presets 的同名字段（如中心点x）。
    """
    dirs = dirs or PEIYI_CATEGORIES
    meta = _peiyi_read_meta(dirs[category], name) or {}

    def _num(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    fw = _num(meta.get("width"))
    fh = _num(meta.get("height"))
    rot = _num(meta.get("rotation"))
    ty = _num(meta.get("highest_y"))
    cx = _num(meta.get("center_x"))

    stem = Path(name).stem
    pkey = _preset_key_for_stem(stem, presets)
    p = presets.get(pkey) if pkey else None
    if p:
        if fw is None:
            fw = _num(p.get("final_w"))
        if fh is None:
            fh = _num(p.get("final_h"))
        if rot is None:
            rot = _num(p.get("rotation_degrees"))
        if ty is None:
            ty = _num(p.get("effective_top_y"))
        if cx is None:
            cx = _num(p.get("effective_center_x"))
    # 兜底默认值（仅在所有来源都缺失时）
    fw = fw or 0.0
    fh = fh or 0.0
    rot = rot or 0.0
    ty = ty or 0.0
    cx = cx if cx is not None else 670.0
    return {"final_w": fw, "final_h": fh, "rotation": rot, "top_y": ty, "center_x": cx}


def _ensure_tpl(stem, fp, cat=None):
    """确保 _tpl/<款名>/ 存在（自动生成扭曲素材）。返回 tpl_dir 或 None。

    实现已迁入 engine/t_shirt.py（TShirtPlugin.load_template），此处仅按品类分发；
    wb 路径与改造前逐字节一致。
    """
    return _garment_plugin_for(cat, cfg=_mockup_cfg(cat)).load_template(stem, fp)


def _ensure_occluder(fp, category):
    """若素材库图片尚未生成 body/occluder 遮罩，则生成（best-effort，超时/失败不阻塞贴图）。"""
    occ = fp.parent / (fp.stem + "_occluder.png")
    if occ.exists():
        return occ
    try:
        env = _single_thread_env(os.environ)
        env["PYTHONPATH"] = f"{ZCODE_PROJECT};{PY_PACKAGES}"
        code = (
            "import peiyi_mask,sys\n"
            "r=peiyi_mask.generate_masks(sys.argv[1], category=sys.argv[2])\n"
            "print('OK' if r.get('ok') else 'FAIL', str(r.get('error',''))[:200])\n"
        )
        r = subprocess.run(
            [str(MOCKUP_PY), "-c", code, str(fp), category],
            cwd=str(MOCKUP_ROOT), env=env, capture_output=True, text=True, timeout=600,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )
        print(f"[贴图] 遮罩生成: {r.stdout.strip()[-200:]} {r.stderr.strip()[-200:]}", flush=True)
    except Exception as e:
        print(f"[贴图] 遮罩生成跳过 {fp.name}: {e}", flush=True)
    return occ if occ.exists() else None


def _remove_white_bg(path):
    """把近白底的 PNG 转透明底（适用于白底/纯色底设计图）。"""
    try:
        from PIL import Image
        import numpy as np
        im = Image.open(path).convert("RGBA")
        a = np.array(im)
        rgb = a[..., :3].astype(np.int16)
        white = (rgb[:, :, 0] > 240) & (rgb[:, :, 1] > 240) & (rgb[:, :, 2] > 240)
        a[white, 3] = 0
        Image.fromarray(a).save(path)
    except Exception:
        pass


def _run_white_t_mockup(design_path, out_path, preset_key, fp, fields, tpl_dir, color, occluder, cat=None):
    """实现已迁入 engine/t_shirt.py（TShirtPlugin.place_design），此处仅按品类分发；

    wb 路径与改造前逐字节一致；hoodie 由桩插件抛 NotImplementedError（等待卫衣模板标定）。
    """
    return _garment_plugin_for(cat, cfg=_mockup_cfg(cat)).place_design(
        design_path, out_path, preset_key, fp, fields, tpl_dir, color, occluder)


@app.route('/api/mockup', methods=['POST'])
def api_mockup():
    """贴图：自动读取胚衣5字段 + 遮罩 + 扭曲，精准贴入。

    入参（multipart）：
      design         : 贴图素材文件（建议透明底 PNG；白底图可勾选 auto_remove_bg）
      template       : 胚衣标识，支持 "分类/文件名"（如 W白/白W3.jpg）或款名；
                        多个用逗号分隔即批量贴图
      auto_remove_bg : '1' 表示把白底设计图去背
    """
    try:
        design = request.files.get('design')
        if not design:
            return jsonify({'ok': False, 'error': '未收到贴图素材'}), 400
        templates = (request.form.get('template') or '').strip()
        if not templates:
            return jsonify({'ok': False, 'error': '未指定胚衣'}), 400
        names = [t.strip() for t in templates.split(',') if t.strip()]
        auto_bg = request.form.get('auto_remove_bg') == '1'

        # 品类化（第4步）：cat 缺省 wb，wb 路径与改造前逐字节一致
        cat = _request_cat()
        peiyi_dirs = _peiyi_dirs(cat)
        if peiyi_dirs is None:
            return jsonify({'ok': False, 'error': f'未知品类: {cat}'}), 400
        cfg = _mockup_cfg(cat)

        cfg.mockup_out.mkdir(parents=True, exist_ok=True)
        des_path = cfg.mockup_out / f"_des_{datetime.now().strftime('%H%M%S%f')}.png"
        design.save(str(des_path))
        if auto_bg:
            _remove_white_bg(des_path)

        presets = _load_presets(cat)
        results = []
        for t in names:
            if '/' in t:
                category, name = t.split('/', 1)
            else:
                category, name = _find_category_for_stem(t, peiyi_dirs)
            if not category or category not in peiyi_dirs:
                results.append({'template': t, 'ok': False, 'error': '未找到该胚衣（请检查素材库分类）'})
                continue
            fp, stem, color = _resolve_peiyi_embryo(category, name, peiyi_dirs)
            if not fp.exists():
                results.append({'template': t, 'ok': False, 'error': f'素材库图片不存在: {fp.name}'})
                continue
            fields = _embryo_fields(category, name, presets, peiyi_dirs)
            if not fields['final_w'] or not fields['final_h']:
                results.append({'template': t, 'ok': False,
                                'error': '该胚衣缺少缩放后宽/高（请在素材库填写，或检查胚衣参数表）'})
                continue
            pkey = _preset_key_for_stem(stem, presets)
            try:
                tpl_dir = _ensure_tpl(stem, fp, cat)
                occ = _ensure_occluder(fp, category)
                out_path = cfg.mockup_out / f"{stem}_{datetime.now().strftime('%H%M%S%f')}.jpg"
                r = _run_white_t_mockup(des_path, out_path, pkey, fp, fields, tpl_dir, color, occ, cat)
            except NotImplementedError as e:
                # 桩品类（如 hoodie）：等待对应模板标定
                results.append({'template': t, 'ok': False, 'error': str(e)})
                continue
            if r.returncode != 0:
                err = (r.stderr or r.stdout or '')[-600:]
                results.append({'template': t, 'ok': False, 'error': err})
                continue
            results.append({
                'template': t, 'ok': True,
                'url': f"/api/mockup/result/{out_path.name}" + _peiyi_cat_qs(cat),
                'fields': fields, 'color': color,
                'used_tpl': tpl_dir is not None, 'used_occluder': occ is not None,
                'preset': pkey,
            })
        return jsonify({'ok': any(x['ok'] for x in results), 'results': results})
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}', 'trace': _tb.format_exc()[-1500:]}), 500


@app.route('/api/mockup/result/<path:filename>')
def api_mockup_result(filename):
    # 品类化（第4步）：cat 缺省 wb，wb 与改造前一致（MOCKUP_OUT）
    mdir = mockup_out_for(_request_cat())
    safe = os.path.basename(filename)
    fp = mdir / safe
    if not fp.exists() or fp.resolve().parent != mdir.resolve():
        abort(404)
    return send_file(str(fp), mimetype='image/jpeg', max_age=0)


# ============================================================================
# 后台生图任务
# ============================================================================

def _run_generation(selected_files: list, task_id: str, reuse_dx: str = None, cat: str = None):
    """后台执行 Lovart 管线。cat 缺省 wb（行为与改造前一致）；非 wb 品类走品类独立目录/注册表。"""
    global task_state
    start_ts = datetime.now()

    # 品类上下文（wb 下全部为既有全局常量，行为不变）
    gp = _gen_paths(cat)
    g_cat        = gp["cat"]
    g_prefix     = gp["prefix"]
    g_inbox      = gp["inbox"]
    g_projects   = gp["projects"]
    g_registry   = gp["registry"]
    g_manifest   = gp["uid_manifest"]
    g_wb_reg     = gp["wb_registry"]
    g_prompt     = gp["prompt"]

    reg = load_registry(g_registry)
    reg = ensure_registry_v4(reg)

    try:
        log(f"▶ 任务 {task_id} 开始")
        log(f"选中文件: {', '.join(selected_files)}")

        # ── 1. 分配 UID / group_id ──────────────────────────────
        selected_set = set(selected_files)
        inbox_groups = group_inbox_files(g_inbox)

        uid_map = {}       # filename → uid
        group_map = {}     # group_number → group_id
        reused_groups = 0  # 并入已有组的次数（同组链接）
        dx_join_map = {}   # filename → 已有 DX 文件夹名（并入旧组时传给 Lovart）
        matched = [g for g in inbox_groups
                   if any(f["filename"] in selected_set for f in g["images"])]

        task_state["groups_total"] = len(matched)
        log(f"识别到 {len(matched)} 个图片组")

        for g in matched:
            # 同组链接：编号已有旧组且角色不冲突 → 并入旧组（后到的另一半进同一 DX 文件夹）；
            # 角色冲突（同角色不同内容）= 新设计复用编号 → 开新组
            sel_imgs = [img for img in g["images"] if img["filename"] in selected_set]
            roles_md5 = {img["suffix"]: compute_md5(str(g_inbox / img["filename"])) for img in sel_imgs}
            gid, join_dx = find_reusable_group(reg, g["group_number"], roles_md5, g_projects)
            if gid:
                log(f"🔗 同组链接: 编号 {g['group_number']} 并入已有组 {gid}（→ {join_dx}）")
                reused_groups = reused_groups + 1
            else:
                gid = get_next_group_id(reg)
                reg["groups"][gid] = {
                    "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "images": [],
                    "source_files": [],
                    "dx_folder": "",
                    "status": "pending",
                }
            group_map[g["group_number"]] = gid

            for img in sel_imgs:
                fname = img["filename"]

                uid = get_next_uid(reg, g_prefix)
                uid_map[fname] = uid
                md5_val = roles_md5[img["suffix"]]

                entry = {
                    "md5": md5_val,
                    "src_id": "",
                    "design_number": g["group_number"],
                    "role": img["suffix"],
                    "original_name": fname,
                    "current_name": fname,
                    "current_path": f"01_INBOX/{fname}",
                    "paired_with": "",
                    "paired_name": "",
                    "cut_path": "",
                    "uid": uid,
                    "group_id": gid,
                    "inbox_original_name": fname,
                    "events": [{
                        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "event": "bridge_generate_start",
                        "detail": f"UID={uid}, group={gid}",
                    }],
                }

                reg["images"][md5_val] = entry
                reg["uid_index"][uid] = md5_val
                reg["name_index"][fname] = md5_val
                reg["groups"][gid]["images"].append(uid)
                reg["groups"][gid]["source_files"].append(fname)
                uid_map[fname] = uid
                if join_dx:
                    dx_join_map[fname] = join_dx

        save_registry(reg, g_registry)
        log(f"已分配 {len(uid_map)} 个 UID，{len(group_map)} 个 group_id" +
            (f"（并入旧组 {reused_groups} 个）" if reused_groups else ""))

        # ── 1b. 写入 UID manifest ─────────────────
        try:
            manifest = {"version": 1, "generated_at": datetime.now().isoformat(), "items": {}}
            for fname, uid in uid_map.items():
                gid = None
                role = ""
                for g in matched:
                    for img in g["images"]:
                        if img["filename"] == fname:
                            gid = group_map[g["group_number"]]
                            role = img["suffix"]
                            break
                    if gid is not None:
                        break
                manifest["items"][fname] = {
                    "uid": uid,
                    "group_id": gid,
                    "role": role,
                }
            g_manifest.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
            log(f"已写入 UID manifest: {g_manifest.name}")
        except Exception as e:
            log(f"WARN: UID manifest 写入失败: {e}")

        # INBOX sidecar：wb_meta 以 DX 为根目录，INBOX 文件无法推断 DX，跳过。
        # 元数据已通过 UID manifest 传给 Lovart，不影响溯源。

            # ── 2. 运行 Lovart 管线（不移走未选中文件，Lovart 自带 SHA256 去重） ──
        task_state["status"] = "running"
        task_state["progress"] = "正在运行 Lovart 生图管线..."
        log("启动 Lovart 管线...")
        _save_state()

        env = os.environ.copy()
        env["PYTHONPATH"] = PYTHONPATH
        env["LOVART_INSECURE_SSL"] = "1"
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUNBUFFERED"] = "1"
        env["BRIDGE_UID_MANIFEST"] = str(g_manifest)
        # 品类注入：数据根 + 款号前缀（wb 下与脚本内缺省值相同，行为不变）
        if g_cat == _DEFAULT_CAT:
            env["SEMEMS_ROOT"] = str(BASE_DIR)      # 便携包=包根 data/，本机=D:\Semems WB，与脚本缺省一致
        else:
            env["SEMEMS_ROOT"] = str(_cat_root(g_cat))
        env["LOVART_ID_PREFIX"] = g_prefix
        # 非 wb 品类：始终注入品类提示词（覆盖视为完整 prompt，文件自带 concrete request）
        if gp["always_prompt"] and g_prompt.exists():
            env["LOVART_PROMPT_FILE"] = str(g_prompt)
        # 强制生成：用户点“开始 Lovart 生图”即明确要生图，忽略去重
        # （原图编号会复用, 每批从1开始, 否则正常生图会被旧记录误拦）
        env["LOVART_FORCE"] = "1"
        # 同组链接：并入已有 DX 文件夹的映射（后到的另一半进同一文件夹，编号可复用）
        if dx_join_map:
            env["LOVART_DX_MAP"] = json.dumps(dx_join_map, ensure_ascii=False)
            log(f"同组链接: {len(dx_join_map)} 个文件并入已有文件夹: {sorted(set(dx_join_map.values()))}")
        # 重新生图时使用统一提示词文件，并传入目标 DX 复用映射
        if task_id and task_id.startswith("TASK_REGEN_"):
            prompt_path = g_prompt
            if prompt_path.exists():
                env["LOVART_PROMPT_FILE"] = str(prompt_path)
            # reuse_dx 可以是单个 DX（str）或 filename -> dx 映射（dict）
            if reuse_dx:
                if isinstance(reuse_dx, dict):
                    regen_map = {fname: reuse_dx[fname] for fname in selected_files if fname in reuse_dx}
                else:
                    regen_map = {fname: reuse_dx for fname in selected_files}
                if regen_map:
                    env["LOVART_REGEN_DX_MAP"] = json.dumps(regen_map)

        proc = subprocess.Popen(
            [get_python(), "run_official_v53.py"],
            cwd=str(LOVART_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding='utf-8',
            env=env,
        )

        # 逐行读取输出，更新进度（完整透传 Lovart 输出，避免“静默完成”看不出原因）
        for line in proc.stdout:
            line = line.rstrip()
            if not line:
                continue
            log(line[:400])
            task_state["progress"] = line[:200]

        proc.wait()

        # ── 4. 更新 registry ────────────────────────────────────
        log("更新注册表，建立溯源关系...")
        reg = load_registry(g_registry)
        reg = ensure_registry_v4(reg)

        # 扫描 Lovart 生成的 {prefix} 文件夹，关联 group + 建立溯源
        if g_projects.exists():
            cutoff = start_ts.timestamp()
            for d in sorted(os.listdir(g_projects)):
                if not d.startswith(g_prefix):
                    continue
                ai_dir = g_projects / d / "01_AI"
                if not ai_dir.exists():
                    continue

                dir_mtime = ai_dir.stat().st_mtime
                if dir_mtime < cutoff:
                    continue

                sm_path = g_projects / d / "source_map.json"
                if not sm_path.exists():
                    continue

                try:
                    with open(sm_path, 'r', encoding='utf-8') as f:
                        sm = json.load(f)
                except Exception:
                    continue

                for src in sm.get("sources", []):
                    src_id = src.get("src_id", "")
                    role = src.get("role", "")
                    target_file = src.get("file", "")
                    uid = src.get("uid", "")
                    gid = src.get("group_id", "")

                    # 优先按 uid 匹配注册表；否则回退 role+group_id
                    img_info = None
                    if uid:
                        md5_key = reg.get("uid_index", {}).get(uid, "")
                        img_info = reg.get("images", {}).get(md5_key)
                    if not img_info and gid and role:
                        for mk, info in reg.get("images", {}).items():
                            if info.get("role") == role and info.get("group_id") == gid:
                                img_info = info
                                break
                    if not img_info:
                        for mk, info in reg.get("images", {}).items():
                            if info.get("role") == role and \
                               info.get("group_id") in group_map.values():
                                img_info = info
                                break

                    if img_info:
                        # 更新注册表
                        img_info["current_name"] = target_file
                        img_info["current_path"] = f"02_PROJECTS/{d}/01_AI/{target_file}"
                        img_info["src_id"] = src_id
                        img_info["events"].append({
                            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "event": "bridge_generate_complete",
                            "detail": f"输出到 {d}/01_AI/{target_file}",
                        })
                        # 更新对应的 group
                        gid = img_info.get("group_id")
                        if gid in reg.get("groups", {}):
                            reg["groups"][gid]["dx_folder"] = d
                            reg["groups"][gid]["status"] = "generated"

                        # 写入 AI sidecar 与 uid_map（wb_meta 的 05_META 根为 wb 专用，仅 wb 写入；
                        # 卫衣等品类的 sidecar 体系待标定，注册表已记录完整溯源，不影响主流程）
                        if wb_meta and g_cat == _DEFAULT_CAT:
                            try:
                                ai_path = g_projects / d / "01_AI" / target_file
                                if ai_path.exists():
                                    inbox_name = img_info.get("inbox_original_name", "")
                                    wb_meta.register_ai(
                                        ai_path,
                                        uid=img_info.get("uid", uid),
                                        group_id=img_info.get("group_id", gid),
                                        role=role,
                                        parent_uid=img_info.get("uid", uid),
                                        inbox_file=f"01_INBOX/{inbox_name}" if inbox_name else None,
                                    )
                            except Exception as e:
                                log(f"WARN: AI sidecar 写入失败 {target_file}: {e}")

        # 建立溯源关系（AI 图 → INBOX 原图）
        lovart_reg_path = g_wb_reg
        lovart_reg = {}
        if lovart_reg_path.exists():
            try:
                with open(lovart_reg_path, 'r', encoding='utf-8') as f:
                    lovart_reg = json.load(f)
            except Exception:
                pass

        for md5_key, img_info in reg.get("images", {}).items():
            if img_info.get("source_type") or not img_info.get("group_id"):
                continue
            gid = img_info.get("group_id")
            if gid not in group_map.values():
                continue
            # 这张 AI 图刚生成，找它的 INBOX 源图
            inbox_name = img_info.get("inbox_original_name", "")
            src_md5 = reg.get("name_index", {}).get(inbox_name, "")
            if src_md5 and src_md5 in reg.get("images", {}):
                _register_provenance(reg, md5_key, src_md5, "ai_gen")

        save_registry(reg, g_registry)
        log("注册表更新完成（含溯源关系）")

        # 通知对应品类的 check_rem 刷新缓存，确保新款立即在去背预览页显示
        # （各品类独立实例，按 g_cat 找端口通知，不再只通知 wb 的 8766）
        _refresh_port = _CHECK_REM_PORT_FOR_CAT.get(g_cat)
        if _refresh_port:
            try:
                urlopen(f"http://127.0.0.1:{_refresh_port}/refresh", timeout=3)
            except Exception:
                pass

        # ── 6. 完成 ─────────────────────────────────────────────
        task_state["status"] = "completed"
        task_state["completed_at"] = datetime.now().isoformat()
        task_state["progress"] = f"任务结束：处理 {len(matched)} 组 / {len(uid_map)} 张"
        if proc.returncode != 0:
            task_state["progress"] += f" (管线退出码: {proc.returncode})"
        log(f"⏹ 任务 {task_id} 结束 (处理 {len(uid_map)} 张, {len(matched)} 组)")
        _save_state()

    except Exception as e:
        task_state["status"] = "error"
        task_state["progress"] = f"错误: {str(e)}"
        log(f"✘ 错误: {str(e)}")
        _save_state()
        import traceback
        log(traceback.format_exc()[:300])

        # 尝试恢复文件
        try:
            td = g_inbox / EXCLUDE_DIR
            if td.exists():
                for fname in list(os.listdir(td)):
                    shutil.move(str(td / fname), str(g_inbox / fname))
                td.rmdir()
        except Exception:
            pass


# ============================================================================
# 入口
# ============================================================================

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Y2 Bridge Server")
    parser.add_argument("--port", type=int, default=8765, help="Bridge 服务端口 (默认 8765)")
    parser.add_argument("--host", type=str, default="127.0.0.1", help="监听地址 (默认 127.0.0.1)")
    args = parser.parse_args()

    # 恢复上次的任务状态（如果是已完成/错误状态）
    _load_state()

    # 自动大写 INBOX 文件名后缀
    renamed = auto_uppercase_inbox()
    if renamed:
        # 重新扫描分组（更新 registry 的 name_index）
        reg = load_registry()
        reg = ensure_registry_v4(reg)
        for fname in os.listdir(INBOX_DIR):
            if fname.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')) and not fname.startswith('_'):
                reg["name_index"][fname] = ""
        save_registry(reg)

    print("╔══════════════════════════════════════════╗")
    print("║   Y2 Bridge Server v2.6.1              ║")
    if renamed:
        print(f"║   AutoUppercase: {renamed} files          ║")
    print("║                                         ║")
    print(f"║   INBOX:   {INBOX_DIR}")
    print(f"║   Output:  {PROJECTS_DIR}")
    print(f"║   Lovart:  {LOVART_SCRIPT}")
    print("║                                         ║")
    print(f"║   Open:  http://{args.host}:{args.port}")
    print("║   AutoScan: every 60s                   ║")
    print("╚══════════════════════════════════════════╝")

    # 后台自动溯源扫描（每 60 秒）
    def _auto_scan_loop():
        while True:
            time.sleep(60)
            try:
                n = scan_provenance()
                if n:
                    print(f"  [AutoScan] 新增 {n} 条血缘关系", flush=True)
            except Exception:
                pass

    t = threading.Thread(target=_auto_scan_loop, daemon=True)
    t.start()

    # 后台守护 check_rem.py（端口 8766），让「去背预览」点击即开
    t2 = threading.Thread(target=_check_rem_daemon, daemon=True)
    t2.start()

    # 写入 PID 文件，供启动脚本优雅停止服务
    pid_path = Path(__file__).resolve().parent / "bridge.pid"
    try:
        pid_path.write_text(str(os.getpid()), encoding="utf-8")
    except Exception:
        pass

    try:
        app.run(host=args.host, port=args.port, debug=False, threaded=True)
    finally:
        try:
            if pid_path.exists():
                pid_path.unlink()
        except Exception:
            pass
